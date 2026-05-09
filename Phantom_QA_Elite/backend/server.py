"""
server.py — Phantom QA Elite Backend Server
=============================================
Autonomous Quality Assurance Command Center
Port: 5030 | Antigravity-AI

Multi-agent test bench: Architect + Ghost User + Skeptic
"""

import os
import sys
import json
import time
import asyncio
import logging
from pathlib import Path
from datetime import datetime
from typing import AsyncGenerator

from fastapi import FastAPI, Request
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import uvicorn
from dotenv import load_dotenv

# ═══════════════════════════════════════════════════════════
#  ENVIRONMENT
# ═══════════════════════════════════════════════════════════

ROOT = Path(__file__).parent.parent
BACKEND_DIR = Path(__file__).parent

# Load .env from multiple possible locations
for env_path in [
    ROOT / ".env",
    ROOT.parent / ".env",
    ROOT.parent.parent / ".env",
]:
    if env_path.exists():
        load_dotenv(env_path)
        break

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(name)s] %(message)s")
logger = logging.getLogger("PhantomQA")

# Add backend to path for imports
sys.path.insert(0, str(BACKEND_DIR))

from memory_store import init_db, save_test_run, save_test_result, \
    get_test_runs, get_test_run, get_dashboard_stats, \
    save_repair_dispatch, mark_repair_complete, get_pending_repairs, get_repair_dispatches
from agents.architect import run_architect
from agents.ghost_user import run_ghost_user
from agents.skeptic import run_skeptic
from warroom_interface import warroom_respond

# ── Ghost Stream Event Queue ─────────────────────────────
_ghost_stream_clients: list[asyncio.Queue] = []

# ── QA Telemetry Stream Queue (Phase 4 — Native Ghost Stream) ──
# Separate channel so QA Architect + Auto-Heal events don't
# collide with Playwright Ghost User activity events.
_qa_stream_clients: list[asyncio.Queue] = []

# ── QA Event Replay Buffer (last 100 events) ──────────────────────
# New SSE connections replay this buffer immediately so the Ghost
# Stream tab never starts blank even mid-run.
_QA_BUFFER_MAX = 100
_qa_event_buffer: list[dict] = []

def ghost_event_callback(event: dict):
    """Push event to all connected Ghost Stream SSE clients."""
    for q in _ghost_stream_clients[:]:
        try:
            q.put_nowait(event)
        except asyncio.QueueFull:
            pass

# ═══════════════════════════════════════════════════════════
#  APP INITIALIZATION
# ═══════════════════════════════════════════════════════════

app = FastAPI(
    title="Phantom QA Elite — Autonomous Quality Assurance",
    version="1.0.0",
    docs_url="/docs",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Serve Frontend ────────────────────────────────────────
FRONTEND_DIR = ROOT / "frontend"
if FRONTEND_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")


# ── No-cache for development ─────────────────────────────
@app.middleware("http")
async def no_cache_static(request: Request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/static/"):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


# ── Global Exception Handler ─────────────────────────────
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    error_msg = str(exc)[:300]
    logger.error(f"[ERROR] {request.method} {request.url.path}: {error_msg}")
    return JSONResponse(
        {"error": f"Engine error: {error_msg}", "agent": "Phantom_QA_Elite"},
        status_code=500
    )


# ── Safe Body Parser ─────────────────────────────────────
async def safe_parse_body(request: Request) -> dict:
    """Parse JSON body safely — returns {} on failure."""
    try:
        body = await request.body()
        if not body or body.strip() == b"":
            return {}
        return await request.json()
    except Exception:
        return {}


# ═══════════════════════════════════════════════════════════
#  ROUTES — Frontend
# ═══════════════════════════════════════════════════════════

@app.get("/")
async def serve_frontend():
    index_path = FRONTEND_DIR / "index.html"
    if index_path.exists():
        return FileResponse(str(index_path))
    return JSONResponse({"error": "Frontend not found"})


# ═══════════════════════════════════════════════════════════
#  ROUTES — Health & Dashboard
# ═══════════════════════════════════════════════════════════

@app.get("/api/health")
async def health():
    # Check Playwright availability
    pw_status = "unknown"
    try:
        import playwright
        pw_status = "installed"
    except ImportError:
        pw_status = "not_installed"

    return {
        "status": "online",
        "agent": "Phantom_QA_Elite",
        "version": "1.0.0",
        "port": 5030,
        "gemini": "configured" if os.environ.get("GEMINI_API_KEY") else "missing",
        "playwright": pw_status,
        "timestamp": datetime.now().isoformat(),
    }


@app.get("/api/dashboard")
async def dashboard():
    stats = get_dashboard_stats()
    return {
        "agent": "Phantom_QA_Elite",
        "stats": stats,
        "timestamp": datetime.now().isoformat(),
    }


# ═══════════════════════════════════════════════════════════
#  ROUTES — Pulse & Auto-Healer
# ═══════════════════════════════════════════════════════════

# Track active reboots so we don't spawn 50 processes while waiting for boot
healing_locks = {}

def get_dynamic_ports():
    ports = {
        "Meta_App_Factory": {
            "url": "http://localhost:5000", 
            "health": "/api/health",
            "cwd": str(ROOT.parent),
            "command": "python api.py",
            "manifest_status": "ACTIVE"
        },
        "Phantom_QA_Elite": {
            "url": "http://localhost:5030", 
            "health": "/api/health",
            "manifest_status": "ACTIVE"
        }
    }
    try:
        import json
        # Construct absolute path pointing straight to the Antigravity-AI Agents root
        # ROOT = Meta_App_Factory/Phantom_QA_Elite
        manifest_path = ROOT.parent.parent / "sync_manifest.json"
        
        if manifest_path.exists():
            with open(manifest_path, "r", encoding="utf-8") as f:
                manifest_array = json.load(f)
                
            if isinstance(manifest_array, list):
                for app in manifest_array:
                    app_name = app.get("name")
                    port = app.get("port")
                    if app_name and port and isinstance(port, int):
                        ports[app_name] = {
                            "url": f"http://localhost:{port}", 
                            "health": "/api/health",
                            "manifest_status": app.get("status", "ACTIVE"),
                            "cwd": app.get("cwd", ""),
                            "command": app.get("command", "")
                        }
        else:
            logger.error(f"Manifest not found at {manifest_path}")
            
    except Exception as e:
        logger.error(f"Failed parsing array manifest direct-read: {e}")
    
    return ports


def auto_heal_agent(name: str, info: dict):
    """Spawns an offline agent autonomously via background subprocess."""
    import subprocess
    import time
    import sys
    
    cwd = info.get("cwd")
    cmd_raw = info.get("command")
    if not cwd or not cmd_raw:
        logger.warning(f"Cannot auto-heal {name}: missing cwd or command.")
        return
        
    # 1. Virtual Environment Sync: Replace generic 'python' with sys.executable
    if cmd_raw.startswith("python "):
        cmd = [sys.executable] + cmd_raw.split(" ")[1:]
    else:
        cmd = cmd_raw.split(" ")
        
    last_heal = healing_locks.get(name, 0)
    now = time.time()
    
    # 30-second cooldown lock to prevent overlapping popen shells
    if now - last_heal < 30:
        return
        
    healing_locks[name] = now
    
    logger.warning(f"🚑 AUTO-HEAL TRIGGERED: Rebooting {name}...")
    
    port = info.get("url", "").split(":")[-1]
    
    # Fire an SSE broadcast to the UI
    qa_event_broadcast({
        "agent": "Phantom_QA_Elite",
        "status": "HEAL_ATTEMPT",
        "message": f"[Auto-Healing] Dispatching restart for {name} on Port {port}"
    })
    
    try:
        # 2. Terminal Visibility: Pop open a new console for the revived agent
        subprocess.Popen(
            cmd,
            cwd=cwd,
            creationflags=subprocess.CREATE_NEW_CONSOLE if os.name == 'nt' else 0
        )
        logger.info(f"✅ Auto-heal subprocess spawned for {name}.")
    except Exception as e:
        logger.error(f"Failed to auto-heal {name}: {e}")

@app.get("/api/pulse")
async def pulse_scan():
    """Scan all dynamically discovered C-Suite ports and report health using native sockets."""
    import socket
    results = {}
    known_ports = get_dynamic_ports()

    for name, info in known_ports.items():
        try:
            port = int(info["url"].split(":")[-1])
            with socket.create_connection(('127.0.0.1', port), timeout=1):
                results[name] = {
                    "status": "online",
                    "url": info["url"],
                    "details": {"status": "socket_connected"},
                    "manifest_status": info.get("manifest_status", "UNKNOWN")
                }
                # Clear healing lock if it successfully came online
                if name in healing_locks:
                    del healing_locks[name]
        except Exception:
            results[name] = {
                "status": "offline", 
                "url": info["url"], 
                "manifest_status": info.get("manifest_status", "UNKNOWN")
            }
            
            # Subsystem Auto-Healer Dispatch
            if name != "Phantom_QA_Elite" and str(info.get("manifest_status")).lower() == "active":
                auto_heal_agent(name, info)

    online = sum(1 for r in results.values() if r["status"] == "online")
    return {
        "timestamp": datetime.now().isoformat(),
        "total_apps": len(known_ports),
        "online": online,
        "offline": len(known_ports) - online,
        "apps": results,
    }


# ═══════════════════════════════════════════════════════════
#  ROUTES — Discovery
# ═══════════════════════════════════════════════════════════

@app.post("/api/discover")
async def discover_app(request: Request):
    data = await safe_parse_body(request)
    target_url = data.get("target_url", "").strip()
    if not target_url:
        return JSONResponse({"error": "target_url is required"}, status_code=400)

    from agents.architect import discover_endpoints
    discovery = await discover_endpoints(target_url)
    return {"discovery": discovery}


# ═══════════════════════════════════════════════════════════
#  ROUTES — Individual Agents
# ═══════════════════════════════════════════════════════════

@app.post("/api/test/architect")
async def test_architect(request: Request):
    data = await safe_parse_body(request)
    target_url = data.get("target_url", "").strip()
    if not target_url:
        return JSONResponse({"error": "target_url is required"}, status_code=400)

    description = data.get("description", "")
    plan = await run_architect(target_url, description)
    return {"agent": "architect", "test_plan": plan}


@app.post("/api/test/ghost")
async def test_ghost(request: Request):
    data = await safe_parse_body(request)
    target_url = data.get("target_url", "").strip()
    if not target_url:
        return JSONResponse({"error": "target_url is required"}, status_code=400)

    headed = data.get("headed", False)
    result = await run_ghost_user(target_url, test_plan=None, headed=headed,
                                   event_callback=ghost_event_callback)
    return {"agent": "ghost_user", "result": result}


@app.post("/api/test/skeptic")
async def test_skeptic(request: Request):
    data = await safe_parse_body(request)
    target_url = data.get("target_url", "").strip()
    if not target_url:
        return JSONResponse({"error": "target_url is required"}, status_code=400)

    result = await run_skeptic(target_url)
    return {"agent": "skeptic", "result": result}


# ═══════════════════════════════════════════════════════════
#  ROUTES — Full Test Bench
# ═══════════════════════════════════════════════════════════

@app.post("/api/test/full")
async def test_full(request: Request):
    """
    Run the complete 3-agent test bench:
    1. Architect → generates test plan
    2. Ghost User → Playwright UI tests
    3. Skeptic → API stress/break tests

    Returns composite verdict with Repair Payloads.
    """
    data = await safe_parse_body(request)
    target_url = data.get("target_url", "").strip()
    if not target_url:
        return JSONResponse({"error": "target_url is required"}, status_code=400)

    app_name = data.get("app_name", target_url.split("://")[-1].split("/")[0])
    description = data.get("description", "")
    skip_ghost = data.get("skip_ghost", False)
    start = time.time()

    logger.info(f"{'='*60}")
    logger.info(f"  PHANTOM QA ELITE — Full Test Bench")
    logger.info(f"  Target: {target_url}")
    logger.info(f"  App: {app_name}")
    logger.info(f"{'='*60}")

    # ── Phase 1: Architect ────────────────────────────────
    logger.info("▶ Phase 1/3: The Architect — Planning test strategy...")
    architect_result = await run_architect(target_url, description)

    # ── Phase 2: Ghost User ───────────────────────────────
    ghost_result = None
    if not skip_ghost and architect_result.get("_discovery", {}).get("has_frontend"):
        logger.info("▶ Phase 2/3: The Ghost User — Playwright UI testing...")
        try:
            ghost_result = await run_ghost_user(target_url, test_plan=architect_result,
                                                event_callback=ghost_event_callback)
        except Exception as e:
            logger.error(f"Ghost User failed: {e}")
            ghost_result = {
                "agent": "ghost_user", "score": 0, "total_tests": 0,
                "passed": 0, "failed": 0, "results": [],
                "error": str(e)[:200],
            }
    else:
        reason = "skipped by user" if skip_ghost else "no frontend detected"
        logger.info(f"▶ Phase 2/3: Ghost User — SKIPPED ({reason})")
        ghost_result = {
            "agent": "ghost_user", "score": 0, "skipped": True,
            "reason": reason, "results": [],
        }

    # ── Phase 3: Skeptic ──────────────────────────────────
    logger.info("▶ Phase 3/3: The Skeptic — API stress testing...")
    skeptic_result = await run_skeptic(target_url, test_plan=architect_result)

    # ── Composite Verdict ─────────────────────────────────
    scores = []
    weights = []

    if not ghost_result.get("skipped"):
        scores.append(ghost_result.get("score", 0))
        weights.append(0.5)  # UI weight

    scores.append(skeptic_result.get("score", 0))
    weights.append(0.5 if ghost_result.get("skipped") else 0.5)

    total_weight = sum(weights)
    composite = sum(s * w for s, w in zip(scores, weights)) / total_weight if total_weight > 0 else 0
    composite = round(composite)

    if composite >= 80:
        verdict = "PASS"
    elif composite >= 50:
        verdict = "WARN"
    else:
        verdict = "FAIL"

    elapsed = time.time() - start

    # ── Collect Repair Payloads ───────────────────────────
    repairs = skeptic_result.get("repair_payloads", [])

    # ── Save to Memory ────────────────────────────────────
    run_id = save_test_run(
        app_name=app_name,
        app_url=target_url,
        verdict=verdict,
        score=composite,
        duration=round(elapsed, 1),
        report_data={
            "verdict": verdict, "score": composite,
            "ghost_score": ghost_result.get("score"),
            "skeptic_score": skeptic_result.get("score"),
        },
        architect_plan=architect_result,
        ghost_summary=ghost_result,
        skeptic_summary=skeptic_result,
        fix_required=repairs if repairs else None,
    )

    # Save individual results
    for r in ghost_result.get("results", []):
        save_test_result(run_id, "ghost_user", r["test_name"], r["passed"],
                          r.get("details", ""), r.get("duration_ms", 0),
                          r.get("screenshot"))

    for r in skeptic_result.get("results", []):
        save_test_result(run_id, "skeptic", r["test_name"], r["passed"],
                          r.get("details", ""), r.get("duration_ms", 0))

    logger.info(f"{'='*60}")
    logger.info(f"  VERDICT: {verdict} (Score: {composite}/100)")
    logger.info(f"  Duration: {elapsed:.1f}s | Repairs: {len(repairs)}")
    logger.info(f"  Run saved: #{run_id}")
    logger.info(f"{'='*60}")

    return {
        "run_id": run_id,
        "verdict": verdict,
        "score": composite,
        "duration_seconds": round(elapsed, 1),
        "app_name": app_name,
        "target_url": target_url,
        "phases": {
            "architect": {
                "status": "complete",
                "endpoints_found": len(architect_result.get("_discovery", {}).get("endpoints", [])),
                "ui_tests_planned": len(architect_result.get("ui_tests", [])),
                "api_tests_planned": len(architect_result.get("api_tests", [])),
                "edge_cases_planned": len(architect_result.get("edge_cases", [])),
                "app_profile": architect_result.get("app_profile", {}),
                "persona": architect_result.get("persona_recommendation", {}),
            },
            "ghost_user": {
                "status": "skipped" if ghost_result.get("skipped") else "complete",
                "score": ghost_result.get("score", 0),
                "total": ghost_result.get("total_tests", 0),
                "passed": ghost_result.get("passed", 0),
                "failed": ghost_result.get("failed", 0),
                "results": ghost_result.get("results", []),
                "console_errors": ghost_result.get("console_errors", []),
            },
            "skeptic": {
                "status": "complete",
                "score": skeptic_result.get("score", 0),
                "total": skeptic_result.get("total_tests", 0),
                "passed": skeptic_result.get("passed", 0),
                "failed": skeptic_result.get("failed", 0),
                "results": skeptic_result.get("results", []),
                "vulnerabilities": skeptic_result.get("vulnerabilities_found", 0),
            },
        },
        "fix_required": repairs,
        "timestamp": datetime.now().isoformat(),
    }


# ═══════════════════════════════════════════════════════════
#  ROUTES — Reports & History
# ═══════════════════════════════════════════════════════════

@app.get("/api/reports")
async def list_reports():
    runs = get_test_runs(limit=20)
    return {"reports": runs, "total": len(runs)}


@app.get("/api/reports/{run_id}")
async def get_report(run_id: int):
    report = get_test_run(run_id)
    if not report:
        return JSONResponse({"error": "Report not found"}, status_code=404)
    return {"report": report}


# ═══════════════════════════════════════════════════════════
#  ROUTES — War Room
# ═══════════════════════════════════════════════════════════

@app.post("/api/warroom/respond")
async def warroom_handler(request: Request):
    data = await safe_parse_body(request)
    question = data.get("question", data.get("input", "")).strip()
    if not question:
        return JSONResponse({"error": "question is required"}, status_code=400)

    context = data.get("context", {})
    response = await warroom_respond(question, context)
    return response


# ═══════════════════════════════════════════════════════════
#  ROUTES — Ghost Stream (SSE)
# ═══════════════════════════════════════════════════════════

@app.get("/api/ghost-stream")
async def ghost_stream():
    """Server-Sent Events stream for real-time Ghost User activity."""

    async def event_generator() -> AsyncGenerator[str, None]:
        q: asyncio.Queue = asyncio.Queue(maxsize=200)
        _ghost_stream_clients.append(q)
        try:
            # Send initial connected event
            yield f"data: {json.dumps({'type': 'CONNECTED', 'timestamp': datetime.now().isoformat()})}\n\n"
            while True:
                try:
                    event = await asyncio.wait_for(q.get(), timeout=30.0)
                    yield f"data: {json.dumps(event)}\n\n"
                except asyncio.TimeoutError:
                    # Keep-alive heartbeat
                    yield f"data: {json.dumps({'type': 'HEARTBEAT', 'timestamp': datetime.now().isoformat()})}\n\n"
        except asyncio.CancelledError:
            pass
        finally:
            if q in _ghost_stream_clients:
                _ghost_stream_clients.remove(q)

    return StreamingResponse(event_generator(), media_type="text/event-stream",
                              headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ═══════════════════════════════════════════════════════════
#  ROUTES — QA Telemetry Stream (Phase 4)
# ═══════════════════════════════════════════════════════════

def qa_event_broadcast(event: dict):
    """Push a structured QA/Auto-Heal event to all connected /api/qa/stream clients
    and append to the replay buffer so fresh connections catch up immediately."""
    if "timestamp" not in event:
        event["timestamp"] = datetime.now().isoformat()
    # Maintain replay ring buffer
    _qa_event_buffer.append(event)
    if len(_qa_event_buffer) > _QA_BUFFER_MAX:
        del _qa_event_buffer[0]
    # Push to live clients
    for q in _qa_stream_clients[:]:
        try:
            q.put_nowait(event)
        except asyncio.QueueFull:
            pass


@app.get("/api/qa/stream")
async def qa_stream():
    """
    SSE endpoint: live telemetry feed for Phantom QA Architect + Auto-Heal loop.
    Connect from the Ghost Stream tab in the frontend:
      const src = new EventSource('http://localhost:5030/api/qa/stream');

    Payload schema:
      { timestamp, agent, message, status, filename?, score?, attempt? }

    status values: RUNNING | PASS | FAIL | SECURITY_BLOCK | TIMEOUT | HEAL_ATTEMPT | HEAL_PASS | HEAL_FAIL | CONNECTED | HEARTBEAT
    """
    async def event_generator() -> AsyncGenerator[str, None]:
        q: asyncio.Queue = asyncio.Queue(maxsize=500)
        _qa_stream_clients.append(q)
        try:
            yield f"data: {json.dumps({'type': 'CONNECTED', 'agent': 'QA_STREAM', 'message': 'Ghost Stream Telemetry online.', 'timestamp': datetime.now().isoformat()})}\n\n"
            # ── Replay buffer: catch up new connections immediately ──
            for buffered_event in list(_qa_event_buffer):
                yield f"data: {json.dumps(buffered_event)}\n\n"
            while True:
                try:
                    event = await asyncio.wait_for(q.get(), timeout=25.0)
                    yield f"data: {json.dumps(event)}\n\n"
                except asyncio.TimeoutError:
                    yield f"data: {json.dumps({'type': 'HEARTBEAT', 'timestamp': datetime.now().isoformat()})}\n\n"
        except asyncio.CancelledError:
            pass
        finally:
            if q in _qa_stream_clients:
                _qa_stream_clients.remove(q)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/qa/ingest")
async def qa_ingest(request: Request):
    """
    External push endpoint: ForgeOrchestrator / QAArchitect / Auto-Heal CTO
    POST structured event payloads here. They are immediately forwarded to
    all connected SSE clients on /api/qa/stream.

    Expected body: { agent, message, status, filename?, score?, attempt? }
    """
    data = await safe_parse_body(request)
    if not data.get("message"):
        return JSONResponse({"error": "message field required"}, status_code=400)

    event = {
        "timestamp": datetime.now().isoformat(),
        "agent":     data.get("agent", "UNKNOWN"),
        "message":   data.get("message", ""),
        "status":    data.get("status", "INFO"),
        "filename":  data.get("filename"),
        "score":     data.get("score"),
        "attempt":   data.get("attempt"),
    }
    # Strip None values so the frontend doesn't have to guard everything
    event = {k: v for k, v in event.items() if v is not None}

    qa_event_broadcast(event)
    logger.info(f"[QA/ingest] {event['agent']} → {event['status']}: {event['message'][:80]}")
    return {"status": "broadcast", "clients": len(_qa_stream_clients), "event": event}



# ═══════════════════════════════════════════════════════════
#  ROUTES — Repair Loop (Loop of Perfection)
# ═══════════════════════════════════════════════════════════

ATOMIZER_WEBHOOK = os.environ.get("ATOMIZER_WEBHOOK_URL", "http://localhost:8000/api/atomizer/repair")


@app.post("/api/repair/dispatch")
async def dispatch_repairs(request: Request):
    """Dispatch repair payloads from a failed test run to Atomizer V2."""
    data = await safe_parse_body(request)
    run_id = data.get("run_id")
    if not run_id:
        return JSONResponse({"error": "run_id is required"}, status_code=400)

    # Get the test run
    run = get_test_run(run_id)
    if not run:
        return JSONResponse({"error": "Test run not found"}, status_code=404)

    repairs = run.get("fix_required") or []
    if not repairs:
        return {"status": "no_repairs_needed", "run_id": run_id}

    # Save dispatch record
    dispatch_id = save_repair_dispatch(run_id, repairs, ATOMIZER_WEBHOOK)

    # Attempt to send to Atomizer
    import aiohttp
    dispatch_status = "dispatched"
    atomizer_response = None
    try:
        async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=10)) as session:
            payload = {
                "source": "Phantom_QA_Elite",
                "run_id": run_id,
                "dispatch_id": dispatch_id,
                "repairs": repairs,
                "callback_url": f"http://localhost:5030/api/repair/complete",
            }
            async with session.post(ATOMIZER_WEBHOOK, json=payload) as r:
                atomizer_response = await r.text()
                if r.status != 200:
                    dispatch_status = "atomizer_error"
    except Exception as e:
        dispatch_status = "atomizer_unreachable"
        atomizer_response = str(e)[:200]
        logger.warning(f"Atomizer unreachable: {e}")

    logger.info(f"🔧 Repair dispatch #{dispatch_id} for run #{run_id}: {dispatch_status} ({len(repairs)} payloads)")

    return {
        "dispatch_id": dispatch_id,
        "run_id": run_id,
        "status": dispatch_status,
        "repair_count": len(repairs),
        "atomizer_response": atomizer_response,
    }


@app.post("/api/repair/complete")
async def repair_complete(request: Request):
    """Webhook: called by Atomizer V2 when repair is complete."""
    data = await safe_parse_body(request)
    dispatch_id = data.get("dispatch_id")
    if not dispatch_id:
        return JSONResponse({"error": "dispatch_id is required"}, status_code=400)

    success = data.get("success", True)

    if dispatch_id:
        mark_repair_complete(dispatch_id, success)
        logger.info(f"🔧 Repair #{dispatch_id} marked {'complete' if success else 'failed'}")

        # Auto-retest if requested
        if data.get("auto_retest") and success:
            logger.info(f"🔧 Auto-retesting after repair #{dispatch_id}...")
            # This would trigger a re-test of the original run's target

    return {"status": "acknowledged", "dispatch_id": dispatch_id}


@app.post("/api/repair/retest/{run_id}")
async def retest_run(run_id: int):
    """Re-run a previous test configuration (e.g., after a repair)."""
    original = get_test_run(run_id)
    if not original:
        return JSONResponse({"error": "Original run not found"}, status_code=404)

    target_url = original.get("app_url", "")
    app_name = original.get("app_name", "")
    if not target_url:
        return JSONResponse({"error": "No target URL in original run"}, status_code=400)

    logger.info(f"🔁 Re-testing run #{run_id}: {app_name} @ {target_url}")

    # Re-run the full test bench
    architect_result = await run_architect(target_url)

    ghost_result = None
    if architect_result.get("_discovery", {}).get("has_frontend"):
        try:
            ghost_result = await run_ghost_user(target_url, test_plan=architect_result,
                                                 event_callback=ghost_event_callback)
        except Exception as e:
            ghost_result = {"agent": "ghost_user", "score": 0, "results": [], "error": str(e)[:200]}
    else:
        ghost_result = {"agent": "ghost_user", "score": 0, "skipped": True, "results": []}

    skeptic_result = await run_skeptic(target_url, test_plan=architect_result)

    # Score
    scores = []
    if not ghost_result.get("skipped"):
        scores.append(ghost_result.get("score", 0))
    scores.append(skeptic_result.get("score", 0))
    composite = round(sum(scores) / len(scores)) if scores else 0
    verdict = "PASS" if composite >= 80 else ("WARN" if composite >= 50 else "FAIL")

    new_run_id = save_test_run(
        app_name=app_name, app_url=target_url, verdict=verdict,
        score=composite, duration=0,
        report_data={"verdict": verdict, "score": composite, "retest_of": run_id},
        architect_plan=architect_result, ghost_summary=ghost_result,
        skeptic_summary=skeptic_result,
    )

    return {
        "retest_of": run_id,
        "new_run_id": new_run_id,
        "verdict": verdict,
        "score": composite,
    }


@app.get("/api/repairs")
async def list_repairs():
    """List all repair dispatches."""
    repairs = get_repair_dispatches()
    pending = get_pending_repairs()
    return {"repairs": repairs, "pending_count": len(pending)}


# ═══════════════════════════════════════════════════════════
#  ROUTES — Auditor's Desk (Manual + Automated Audit)
# ═══════════════════════════════════════════════════════════

from fastapi import UploadFile, File, Form

AUDIT_UPLOADS_DIR = ROOT / "audit_uploads"
AUDIT_UPLOADS_DIR.mkdir(exist_ok=True)


@app.post("/api/audit")
async def manual_audit(
    target_url: str = Form(...),
    audit_mode: str = Form("structural"),
    file_link: str = Form(""),
    file: UploadFile | None = File(None),
):
    """
    Manual audit from the Auditor's Desk UI.
    Routes to the appropriate agent based on audit_mode.
    """
    start = time.time()

    # Handle optional file upload
    file_path = None
    if file and file.filename:
        safe_name = f"audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        file_path = AUDIT_UPLOADS_DIR / safe_name
        content = await file.read()
        file_path.write_bytes(content)
        logger.info(f"[AUDIT] File uploaded: {file.filename} ({len(content)} bytes)")

    logger.info(f"[AUDIT] Manual {audit_mode} audit of {target_url}")
    result = await _run_audit(target_url, audit_mode, file_link, file_path)
    result["duration_seconds"] = round(time.time() - start, 1)
    result["source"] = "manual"
    return result


@app.post("/api/audit/auto")
async def automated_audit(request: Request):
    """
    Automated quality gate — called by CFO or other agents after report generation.
    Accepts: { source, target_url, file_link, file_name, report_data }
    Returns: verdict + blocks on FAIL + sends correction request.
    """
    data = await safe_parse_body(request)
    source = data.get("source", "unknown")
    target_url = data.get("target_url", "http://localhost:5041")
    file_link = data.get("file_link", "")
    file_name = data.get("file_name", "")
    report_data = data.get("report_data", {})

    start = time.time()
    logger.info(f"[AUDIT-AUTO] Automated audit from {source}: {file_name or target_url}")

    # Auto-audits always use mathematical mode for CFO reports
    audit_mode = data.get("audit_mode", "mathematical")
    result = await _run_audit(target_url, audit_mode, file_link, None, report_data)
    result["duration_seconds"] = round(time.time() - start, 1)
    result["source"] = source
    result["file_name"] = file_name

    # If FAIL → block file and send correction request back to source
    if result["verdict"] == "FAIL":
        result["blocked"] = True
        correction = (
            f"[Phantom QA] Quality gate FAILED for '{file_name}'. "
            f"Score: {result['score']}/100. "
            f"Failed checks: {result.get('failed', 0)}/{result.get('total_tests', 0)}. "
            f"Please review and correct the following issues before resubmission."
        )
        result["correction_request"] = correction

        # Try to send correction back to the source agent
        try:
            import aiohttp
            callback_url = data.get("callback_url", f"{target_url}/api/audit/correction")
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=5)) as session:
                payload = {
                    "source": "Phantom_QA_Elite",
                    "verdict": "FAIL",
                    "score": result["score"],
                    "correction_request": correction,
                    "findings": result.get("findings", []),
                    "file_name": file_name,
                }
                async with session.post(callback_url, json=payload) as r:
                    logger.info(f"[AUDIT-AUTO] Correction sent to {callback_url}: HTTP {r.status}")
        except Exception as e:
            logger.warning(f"[AUDIT-AUTO] Failed to send correction to source: {e}")
    else:
        result["blocked"] = False

    logger.info(f"[AUDIT-AUTO] Verdict: {result['verdict']} (Score: {result['score']}/100)")
    return result


# ── Mathematical Audit Hardening ─────────────────────────
import re

EXCEL_ERROR_CODES = ["#DIV/0!", "#REF!", "#VALUE!", "#NAME?", "#NULL!", "#N/A", "#NUM!"]
SPEND_BUDGET_KEYWORDS = ["spend", "budget", "cost", "expense", "investment", "outlay"]


def _run_mathematical_hardening(file_path=None, file_link: str = "",
                                 report_data: dict = None) -> list:
    """
    Strict financial integrity checks. Returns a list of critical findings.
    Any finding = automatic FAIL (score < 40).

    Checks:
      1. Excel Error Codes: #DIV/0!, #REF!, #VALUE!, #NAME?, etc.
      2. Financial Impossibilities: Negative spend/budget values.
      3. Formula Ghosting: Text cells that look like formulas (start with =).
    """
    critical_findings = []

    # ── Check uploaded file (Excel) ──────────────────────
    if file_path and Path(file_path).exists():
        ext = Path(file_path).suffix.lower()
        if ext in (".xlsx", ".xls"):
            critical_findings.extend(_scan_excel_file(file_path))
        elif ext in (".csv",):
            critical_findings.extend(_scan_csv_file(file_path))

    # ── Check report_data dict ───────────────────────────
    if report_data and isinstance(report_data, dict):
        critical_findings.extend(_scan_report_data(report_data))

    return critical_findings


def _scan_excel_file(file_path) -> list:
    """Scan an Excel file for error codes, negative financials, and formula ghosts."""
    findings = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Get header row to identify financial columns
            headers = {}
            for col_idx, cell in enumerate(ws[1] if ws.max_row >= 1 else [], 1):
                if cell.value and isinstance(cell.value, str):
                    headers[col_idx] = cell.value.strip().lower()

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=False), 1):
                for col_idx, cell in enumerate(row, 1):
                    val = cell.value
                    if val is None:
                        continue
                    cell_ref = f"{sheet_name}!{cell.coordinate}"

                    # Check 1: Excel Error Codes
                    if isinstance(val, str):
                        for err_code in EXCEL_ERROR_CODES:
                            if err_code in val.upper():
                                findings.append({
                                    "test_name": f"Excel Error Code: {err_code}",
                                    "passed": False,
                                    "details": f"Cell {cell_ref} contains error '{val}'",
                                })

                        # Check 3: Formula Ghosting
                        if val.strip().startswith("=") and cell.data_type == "s":
                            findings.append({
                                "test_name": "Formula Ghosting",
                                "passed": False,
                                "details": (
                                    f"Cell {cell_ref} contains text '{val[:60]}' "
                                    f"that looks like a formula but is stored as static text"
                                ),
                            })

                    # Check 2: Financial Impossibilities
                    if isinstance(val, (int, float)) and val < 0:
                        col_header = headers.get(col_idx, "")
                        if any(kw in col_header for kw in SPEND_BUDGET_KEYWORDS):
                            findings.append({
                                "test_name": "Financial Impossibility: Negative Spend",
                                "passed": False,
                                "details": (
                                    f"Cell {cell_ref} in column '{headers.get(col_idx, '?')}' "
                                    f"has value {val} — spend/budget cannot be negative"
                                ),
                            })

        wb.close()
    except ImportError:
        logger.warning("[AUDIT] openpyxl not installed — skipping Excel file scan")
    except Exception as e:
        logger.error(f"[AUDIT] Excel scan error: {e}")
        findings.append({
            "test_name": "Excel File Scan",
            "passed": False,
            "details": f"Failed to scan file: {str(e)[:200]}",
        })
    return findings


def _scan_csv_file(file_path) -> list:
    """Scan a CSV file for error codes, negative financials, and formula ghosts."""
    findings = []
    try:
        import csv
        with open(str(file_path), "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            headers = []
            for row_idx, row in enumerate(reader, 1):
                if row_idx == 1:
                    headers = [h.strip().lower() for h in row]
                    continue
                for col_idx, val in enumerate(row):
                    val = val.strip()
                    if not val:
                        continue
                    cell_ref = f"Row {row_idx}, Col {col_idx + 1}"

                    # Check 1: Excel Error Codes
                    for err_code in EXCEL_ERROR_CODES:
                        if err_code in val.upper():
                            findings.append({
                                "test_name": f"Excel Error Code: {err_code}",
                                "passed": False,
                                "details": f"{cell_ref} contains error '{val}'",
                            })

                    # Check 3: Formula Ghosting
                    if val.startswith("="):
                        findings.append({
                            "test_name": "Formula Ghosting",
                            "passed": False,
                            "details": f"{cell_ref} contains text '{val[:60]}' — looks like a formula stored as text",
                        })

                    # Check 2: Financial Impossibilities
                    col_header = headers[col_idx] if col_idx < len(headers) else ""
                    if any(kw in col_header for kw in SPEND_BUDGET_KEYWORDS):
                        try:
                            numeric = float(val.replace(",", "").replace("$", ""))
                            if numeric < 0:
                                findings.append({
                                    "test_name": "Financial Impossibility: Negative Spend",
                                    "passed": False,
                                    "details": f"{cell_ref} in '{col_header}' = {numeric} — cannot be negative",
                                })
                        except ValueError:
                            pass
    except Exception as e:
        logger.error(f"[AUDIT] CSV scan error: {e}")
    return findings


def _scan_report_data(report_data: dict, path: str = "") -> list:
    """Recursively scan report_data dict for error codes and negative financials."""
    findings = []

    def _walk(obj, key_path="root"):
        if isinstance(obj, dict):
            for k, v in obj.items():
                _walk(v, f"{key_path}.{k}")
        elif isinstance(obj, list):
            for i, v in enumerate(obj):
                _walk(v, f"{key_path}[{i}]")
        elif isinstance(obj, str):
            # Check for Excel error codes
            for err_code in EXCEL_ERROR_CODES:
                if err_code in obj.upper():
                    findings.append({
                        "test_name": f"Excel Error Code: {err_code}",
                        "passed": False,
                        "details": f"Field '{key_path}' contains error code '{obj[:80]}'",
                    })
            # Check for formula ghosting
            if obj.strip().startswith("="):
                findings.append({
                    "test_name": "Formula Ghosting",
                    "passed": False,
                    "details": f"Field '{key_path}' contains formula-like text: '{obj[:60]}'",
                })
        elif isinstance(obj, (int, float)):
            # Check for negative spend/budget
            key_lower = key_path.lower()
            if any(kw in key_lower for kw in SPEND_BUDGET_KEYWORDS) and obj < 0:
                findings.append({
                    "test_name": "Financial Impossibility: Negative Spend",
                    "passed": False,
                    "details": f"Field '{key_path}' = {obj} — spend/budget cannot be negative",
                })

    _walk(report_data)
    return findings


async def _run_audit(target_url: str, audit_mode: str, file_link: str = "",
                     file_path=None, report_data: dict = None) -> dict:
    """
    Core audit logic — runs the appropriate agent(s) based on audit mode
    and returns a unified verdict response.
    """
    findings = []
    total_tests = 0
    passed = 0
    failed = 0

    try:
        if audit_mode in ("structural", "mathematical"):
            # Use Skeptic for structural/mathematical audits
            skeptic_result = await run_skeptic(target_url)
            results = skeptic_result.get("results", [])
            for r in results:
                findings.append({
                    "test_name": r.get("test_name", ""),
                    "passed": r.get("passed", False),
                    "details": r.get("details", ""),
                })
            total_tests = skeptic_result.get("total_tests", len(results))
            passed = skeptic_result.get("passed", 0)
            failed = skeptic_result.get("failed", 0)
            score = skeptic_result.get("score", 0)

            # ── HARDENED MATHEMATICAL AUDIT ───────────────
            # If mode is "mathematical", apply strict financial checks
            # on file contents and report_data.
            if audit_mode == "mathematical":
                math_findings = _run_mathematical_hardening(
                    file_path=file_path,
                    file_link=file_link,
                    report_data=report_data,
                )
                if math_findings:
                    # Critical failures found — force score < 40
                    for mf in math_findings:
                        findings.append(mf)
                        total_tests += 1
                        failed += 1
                    # Recalculate score — cap at 35 if any critical check fails
                    critical_count = len(math_findings)
                    score = min(score, max(0, 35 - (critical_count * 8)))
                    logger.warning(
                        f"[AUDIT] Mathematical hardening: {critical_count} "
                        f"critical issues found — score capped at {score}"
                    )

        elif audit_mode == "stress-test":
            # Full Skeptic battery
            skeptic_result = await run_skeptic(target_url)
            results = skeptic_result.get("results", [])
            for r in results:
                findings.append({
                    "test_name": r.get("test_name", ""),
                    "passed": r.get("passed", False),
                    "details": r.get("details", ""),
                })
            total_tests = skeptic_result.get("total_tests", len(results))
            passed = skeptic_result.get("passed", 0)
            failed = skeptic_result.get("failed", 0)
            score = skeptic_result.get("score", 0)

        elif audit_mode == "uiux":
            # Ghost User UI tests
            try:
                ghost_result = await run_ghost_user(target_url, test_plan=None,
                                                     event_callback=ghost_event_callback)
                results = ghost_result.get("results", [])
                for r in results:
                    findings.append({
                        "test_name": r.get("test_name", ""),
                        "passed": r.get("passed", False),
                        "details": r.get("details", ""),
                    })
                total_tests = ghost_result.get("total_tests", len(results))
                passed = ghost_result.get("passed", 0)
                failed = ghost_result.get("failed", 0)
                score = ghost_result.get("score", 0)
            except Exception as e:
                logger.error(f"[AUDIT] Ghost User failed: {e}")
                findings.append({"test_name": "Ghost User Launch", "passed": False,
                                  "details": str(e)[:200]})
                total_tests = 1
                failed = 1
                score = 0

        else:
            # Default: Architect discovery
            architect_result = await run_architect(target_url)
            endpoints = architect_result.get("_discovery", {}).get("endpoints", [])
            total_tests = len(endpoints) + 1
            has_frontend = architect_result.get("_discovery", {}).get("has_frontend", False)

            findings.append({
                "test_name": "Endpoint Discovery",
                "passed": len(endpoints) > 0,
                "details": f"Found {len(endpoints)} endpoints",
            })
            if len(endpoints) > 0:
                passed += 1
            else:
                failed += 1

            for ep in endpoints[:10]:
                findings.append({
                    "test_name": f"Endpoint: {ep.get('path', '?')}",
                    "passed": True,
                    "details": f"{ep.get('method', '?')} — {ep.get('description', 'discovered')}",
                })
                passed += 1

            score = round((passed / max(total_tests, 1)) * 100)

    except Exception as e:
        logger.error(f"[AUDIT] Error during {audit_mode} audit: {e}")
        findings.append({"test_name": "Audit Engine", "passed": False,
                          "details": f"Internal error: {str(e)[:200]}"})
        total_tests = 1
        failed = 1
        score = 0

    # Determine verdict
    if score >= 80:
        verdict = "PASS"
    elif score >= 50:
        verdict = "WARN"
    else:
        verdict = "FAIL"

    cfo_analysis = None
    if audit_mode in ("mathematical", "structural") and verdict == "FAIL":
        try:
            import aiohttp
            # Forward to CFO Fragility Engine natively
            qa_event_broadcast({
                "timestamp": datetime.now().isoformat(),
                "agent": "Phantom QA Elite",
                "message": f"Mathematical failure detected on {target_url}. Forwarding exact traces to the CFO Fragility Engine for deeper forensic tear-down...",
                "status": "RUNNING",
                "score": score
            })
            
            trace_data = "\n".join([f"- {f.get('test_name', '')}: {f.get('details', '')}" for f in findings if not f.get('passed')])
            fail_prompt = f"Mathematical / Structural Audit Failed across test harness for {target_url}. Here are the exact trace logs. Please perform a deep structural and financial reasoning teardown on why this failed and how to reconstruct it:\n{trace_data}"
            
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=45)) as session:
                data_fd = aiohttp.FormData()
                data_fd.add_field('instruction', fail_prompt)
                
                async with session.post("http://localhost:5041/api/consult", data=data_fd) as cfo_r:
                    if cfo_r.status == 200:
                        cfo_data = await cfo_r.json()
                        raw_cfo = cfo_data.get("message", "") + "\n\n" + str(cfo_data.get("report", ""))
                        if cfo_data.get("status") == "received":
                            cfo_analysis = "CFO Engine response: " + raw_cfo
                        else:
                            cfo_analysis = raw_cfo
                        
                        qa_event_broadcast({
                            "timestamp": datetime.now().isoformat(),
                            "agent": "CFO Fragility Engine",
                            "message": f"Forensic analysis returned ({len(cfo_analysis)} chars).",
                            "status": "FAIL", 
                            "score": score
                        })
                    else:
                        cfo_analysis = f"CFO returned error code {cfo_r.status}: {await cfo_r.text()}"
        except Exception as e:
            logger.warning(f"Failed to bridge trace to CFO Engine: {e}")
            cfo_analysis = f"CFO Fragility link disrupted. Trace could not be forwarded. Error: {e}"

    # Save to memory
    run_id = save_test_run(
        app_name=f"AUDIT:{audit_mode}",
        app_url=target_url,
        verdict=verdict,
        score=score,
        duration=0,
        report_data={"verdict": verdict, "score": score, "audit_mode": audit_mode, "cfo_analysis": cfo_analysis},
        architect_plan=None,
        ghost_summary=None,
        skeptic_summary=None,
    )

    return {
        "verdict": verdict,
        "score": score,
        "audit_mode": audit_mode,
        "target_url": target_url,
        "file_link": file_link,
        "run_id": run_id,
        "total_tests": total_tests,
        "passed": passed,
        "failed": failed,
        "findings": findings,
        "cfo_analysis": cfo_analysis,
        "timestamp": datetime.now().isoformat(),
    }


# ═══════════════════════════════════════════════════════════
#  STARTUP
# ═══════════════════════════════════════════════════════════

PORT = 5030

if __name__ == "__main__":
    init_db()

    print("")
    print("=" * 60)
    print("")
    print("   PHANTOM QA ELITE -- Autonomous Quality Assurance")
    print("")
    print("   Port: %d" % PORT)
    print("   Dashboard: http://localhost:%d" % PORT)
    print("   API Docs:  http://localhost:%d/docs" % PORT)
    print("")
    print("   War Room:  /api/warroom/respond")
    print("   Health:    /api/health")
    print("   Pulse:     /api/pulse")
    print("")
    print("   Agents:")
    print("     [A]  The Architect (Test Planner)")
    print("     [G]  The Ghost User (Playwright UI)")
    print("     [S]  The Skeptic (API Bug Hunter)")
    print("")
    print("   Antigravity-AI | Phantom QA Elite v1.0.0")
    print("")
    print("=" * 60)
    print("")

    uvicorn.run(app, host="0.0.0.0", port=PORT, log_level="info")
