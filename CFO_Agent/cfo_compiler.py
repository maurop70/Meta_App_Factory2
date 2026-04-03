"""
CFO Model Compiler (Institutional Grade)
Transforms JSON blueprint specifications into dynamic Excel models based on complexity tiers.
"""
import os
import sys
import json
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from blocks import RevenueBlock, DebtBlock, ReturnsBlock

def _load_patterns():
    pattern_path = os.path.join(os.path.dirname(__file__), "cfo_patterns.json")
    if os.path.exists(pattern_path):
        with open(pattern_path, "r") as f:
            return json.load(f)
    return {"system_memory": []}

# Load the skills library audit
def _get_audit_module():
    _skills_lib = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', '..', '_ANTIGRAVITY_SKILLS_LIBRARY'))
    if _skills_lib not in sys.path:
        sys.path.insert(0, _skills_lib)
    try:
        import recursive_xml_audit
        return recursive_xml_audit
    except ImportError:
        return None

class CFOCompiler:
    def __init__(self, output_dir=None):
        self.output_dir = output_dir or os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "reports"
        )
        os.makedirs(self.output_dir, exist_ok=True)
        self.patterns = _load_patterns()
        self.audit_engine = _get_audit_module()

    def compile_model(self, spec: dict) -> dict:
        """
        Takes a JSON specification and compiles it into an Excel workbook.
        
        spec format:
        {
          "model_name": "Lemonade_Budget",
          "complexity": "elementary", # elementary, mid_market, institutional
          "blocks": ["revenue", "debt", "returns"],
          "parameters": { ... }
        }
        """
        if not EXCEL_AVAILABLE:
            return {"status": "FAIL", "message": "openpyxl not available."}

        model_name = spec.get('model_name', 'Dynamic_CFO_Model')
        complexity = spec.get('complexity', 'mid_market')
        requested_blocks = spec.get('blocks', ['revenue', 'returns'])
        params = spec.get('parameters', {})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Model_{complexity.title()}"
        ws.column_dimensions['A'].width = 30
        for c in range(2, 20):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 18

        # --- Layer 1 Settings ---
        # Toggle circular computation for institutional models
        if complexity == 'institutional':
            wb.calculation.iterate = True
            wb.calculation.maxIterations = 100

        # Title
        ws.cell(row=1, column=1, value=model_name.replace("_", " ").upper()).font = Font(size=14, bold=True, color='E94560')
        ws.cell(row=2, column=1, value=f"Complexity Level: {complexity.upper()} | Autonomously compiled by CFO Agent").font = Font(italic=True, color='888888')
        
        current_row = 4
        
        # --- Layer 2: Block Assembly ---
        block_instances = []
        for b_name in requested_blocks:
            if b_name == 'revenue':
                block_instances.append(RevenueBlock("Revenue", complexity, params))
            elif b_name == 'debt':
                block_instances.append(DebtBlock("Debt", complexity, params))
            elif b_name == 'returns':
                block_instances.append(ReturnsBlock("Returns", complexity, params))

        global_cell_map = {}
        for block in block_instances:
            # Pass global knowledge so far to next block
            block.cell_map.update(global_cell_map)
            current_row = block.render(ws, start_row=current_row, current_col=1)
            global_cell_map.update(block.get_references())

        # --- Verification Trace (VC Shadow Sheet) ---
        ws_audit = wb.create_sheet('Verification Trace')
        ws_audit.cell(row=1, column=1, value="AUDIT TRAIL & SYSTEM MEMORY").font = Font(bold=True)
        ws_audit.cell(row=2, column=1, value="Traceability of standard logic applied by CFO.")
        
        r = 4
        for pm in self.patterns.get("system_memory", []):
            ws_audit.cell(row=r, column=1, value=f"Rule: {pm.get('rule')}")
            ws_audit.cell(row=r, column=2, value=pm.get("rationale"))
            r += 1

        r += 2
        ws_audit.cell(row=r, column=1, value="Audit Signature")

        # --- Save Phase ---
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{model_name}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)

        # --- Layer 3: XML Audit ---
        if self.audit_engine:
            audit_res = self.audit_engine.audit_workbook(filepath)
            if audit_res.get('status') == 'FAIL':
                os.remove(filepath)
                return {
                    "status": "FAIL", 
                    "message": f"XML Recursive Audit Failed: {audit_res.get('errors')}"
                }
            
            # Reopen and inject ghost note
            ghost_note = self.audit_engine.generate_ghost_note(audit_res)
            wb_reopen = openpyxl.load_workbook(filepath)
            ws_audit_reopen = wb_reopen['Verification Trace']
            for row in ws_audit_reopen.iter_rows(min_row=1, max_row=20, min_col=1, max_col=2):
                for cell in row:
                    if cell.value == 'Audit Signature':
                        ws_audit_reopen.cell(row=cell.row, column=2, value=ghost_note)
                        break
            wb_reopen.save(filepath)

        return {
            "status": "PASS",
            "file_name": filename,
            "filepath": filepath,
            "message": f"Model successfully compiled with '{complexity}' tier complexity."
        }
