from .base_block import BaseBlock
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

class DebtBlock(BaseBlock):
    def render(self, ws, start_row: int, current_col: int = 1) -> int:
        r = start_row
        
        ws.cell(row=r, column=current_col, value="DEBT SCHEDULE").font = Font(bold=True, color='E94560')
        r += 1

        if self.complexity == 'elementary':
            # Elementary doesn't usually have a debt schedule, just straight costs
            ws.cell(row=r, column=current_col, value="Loans / Debt").font = Font(italic=True)
            ws.cell(row=r, column=current_col+1, value="N/A for elementary models")
            r += 2
            
        elif self.complexity == 'mid_market':
            # Linearized debt to prevent circular Google Sheets errors
            # (Matches rule: linearized_debt_fallback)
            beg_debt = self.parameters.get('beginning_debt', 1000000)
            rate = self.parameters.get('interest_rate', 0.08)

            ws.cell(row=r, column=current_col, value="Beginning Debt").font = Font(bold=True)
            ws.cell(row=r, column=current_col+1, value=beg_debt)
            self.cell_map['beg_debt'] = f"{get_column_letter(current_col+1)}{r}"
            r += 1

            ws.cell(row=r, column=current_col, value="Interest Rate").font = Font(bold=True)
            ws.cell(row=r, column=current_col+1, value=rate)
            self.cell_map['interest_rate'] = f"{get_column_letter(current_col+1)}{r}"
            r += 1

            ws.cell(row=r, column=current_col, value="Interest Expense")
            # Linearized
            formula = f"={self.cell_map['beg_debt']}*{self.cell_map['interest_rate']}"
            ws.cell(row=r, column=current_col+1, value=self._apply_pattern(formula))
            self.cell_map['interest_expense'] = f"{get_column_letter(current_col+1)}{r}"
            r += 2

        elif self.complexity == 'institutional':
            # Requires circular debt sculpting
            # Evaluated securely because Layer 1 (Compiler) will enable `wb.calculation.iterate = True`
            periods = self.parameters.get('periods', 5)
            rate = self.parameters.get('interest_rate', 0.08)
            initial_debt = self.parameters.get('beginning_debt', 50000000)
            
            headers = ["Metric"] + [f"Year {i}" for i in range(1, periods + 1)]
            for col_idx, h in enumerate(headers, start=current_col):
                ws.cell(row=r, column=col_idx, value=h).font = Font(bold=True)
            
            # Setup rows
            row_rate = r + 1
            row_beg = r + 2
            row_int = r + 3
            row_cfads = r + 4 # Cash flow available for debt service - stub parameter for now
            row_paydown = r + 5
            row_end = r + 6
            
            ws.cell(row=row_rate, column=current_col, value="Interest Rate")
            ws.cell(row=row_beg, column=current_col, value="Beginning Debt")
            ws.cell(row=row_int, column=current_col, value="Interest Expense (Avg Debt)")
            ws.cell(row=row_cfads, column=current_col, value="Operating Cash Flow (CFADS)")
            ws.cell(row=row_paydown, column=current_col, value="Mandatory Paydown")
            ws.cell(row=row_end, column=current_col, value="Ending Debt").font = Font(bold=True)

            for p in range(1, periods + 1):
                c_idx = current_col + p
                col_letter = get_column_letter(c_idx)
                
                # Rate
                ws.cell(row=row_rate, column=c_idx, value=rate)
                
                # Beginning Debt
                if p == 1:
                    ws.cell(row=row_beg, column=c_idx, value=initial_debt)
                else:
                    prev_col = get_column_letter(c_idx - 1)
                    ws.cell(row=row_beg, column=c_idx, value=self._apply_pattern(f"={prev_col}{row_end}"))
                
                # CFADS (Stub)
                cfads = self.parameters.get('base_cfads', 10000000) * (1.1 ** (p-1))
                ws.cell(row=row_cfads, column=c_idx, value=cfads)
                
                # Circular Interest Expense: Uses Average Debt!
                # Beg + End / 2 * Rate
                int_formula = f"=({col_letter}{row_beg}+{col_letter}{row_end})/2*{col_letter}{row_rate}"
                ws.cell(row=row_int, column=c_idx, value=self._apply_pattern(int_formula))
                
                # Paydown (CFADS - Interest ensures circularity)
                paydown_formula = f"=MIN({col_letter}{row_beg}, {col_letter}{row_cfads}-{col_letter}{row_int})"
                ws.cell(row=row_paydown, column=c_idx, value=self._apply_pattern(paydown_formula))
                
                # Ending Debt
                end_formula = f"={col_letter}{row_beg}-{col_letter}{row_paydown}"
                ws.cell(row=row_end, column=c_idx, value=self._apply_pattern(end_formula))
                
                self.cell_map[f'debt_y{p}'] = f"{col_letter}{row_end}"
                self.cell_map[f'interest_y{p}'] = f"{col_letter}{row_int}"
            
            r = row_end + 2
            
        return r
