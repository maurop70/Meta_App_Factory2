from .base_block import BaseBlock
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

class ReturnsBlock(BaseBlock):
    def render(self, ws, start_row: int, current_col: int = 1) -> int:
        r = start_row
        
        ws.cell(row=r, column=current_col, value="RETURNS ANALYSIS").font = Font(bold=True, color='E94560')
        r += 1

        if self.complexity == 'elementary':
            # Simple ROI or profit display
            savings = self.parameters.get('expected_savings', 50)
            ws.cell(row=r, column=current_col, value="Expected Profit/Savings").font = Font(bold=True)
            ws.cell(row=r, column=current_col+1, value=savings)
            r += 2
            
        elif self.complexity == 'mid_market':
            cost = self.parameters.get('initial_cost', 50000)
            revenue_ref = self.cell_map.get('total_revenue') or self.cell_map.get('revenue_y1', "100000")
            
            ws.cell(row=r, column=current_col, value="Initial Investment")
            ws.cell(row=r, column=current_col+1, value=cost)
            self.cell_map['initial_investment'] = f"{get_column_letter(current_col+1)}{r}"
            r += 1

            ws.cell(row=r, column=current_col, value="ROI %").font = Font(bold=True)
            if revenue_ref and not str(revenue_ref).isnumeric():
                # We have a cell reference
                formula = f"=({revenue_ref}-{self.cell_map['initial_investment']})/{self.cell_map['initial_investment']}*100"
            else:
                formula = f"=({revenue_ref}-{cost})/{cost}*100"
            
            ws.cell(row=r, column=current_col+1, value=self._apply_pattern(formula))
            r += 2

        elif self.complexity == 'institutional':
            # Complex IRR and Cash on Cash
            periods = self.parameters.get('periods', 5)
            initial_equity = self.parameters.get('initial_equity', 20000000)
            exit_multiple = self.parameters.get('exit_ebitda_multiple', 10.0)
            discount_rate = self.parameters.get('discount_rate', 0.1)
            
            headers = ["Metric", "Year 0"] + [f"Year {i}" for i in range(1, periods + 1)]
            for col_idx, h in enumerate(headers, start=current_col):
                ws.cell(row=r, column=col_idx, value=h).font = Font(bold=True)
            r += 1

            row_cf = r
            ws.cell(row=row_cf, column=current_col, value="Sponsor Cash Flows")
            ws.cell(row=row_cf, column=current_col+1, value=-initial_equity)
            
            for p in range(1, periods + 1):
                c_idx = current_col + 1 + p
                col_letter = get_column_letter(c_idx)
                
                if p == periods:
                    # Final year = Exit EBITDA * Multiple - Ending Debt
                    # Stub calculation
                    exit_ebitda = self.parameters.get('final_ebitda', 15000000)
                    debt_ref = self.cell_map.get(f'debt_y{p}', "0")
                    if str(debt_ref).isnumeric():
                        exit_val = (exit_ebitda * exit_multiple) - float(debt_ref)
                        ws.cell(row=row_cf, column=c_idx, value=exit_val)
                    else:
                        exit_formula = f"=({exit_ebitda}*{exit_multiple})-{debt_ref}"
                        ws.cell(row=row_cf, column=c_idx, value=self._apply_pattern(exit_formula))
                else:
                    ws.cell(row=row_cf, column=c_idx, value=0) # No sponsor dividends during hold typically
            
            r += 2
            ws.cell(row=r, column=current_col, value="IRR %").font = Font(bold=True)
            end_col_letter = get_column_letter(current_col + 1 + periods)
            irr_range = f"{get_column_letter(current_col+1)}{row_cf}:{end_col_letter}{row_cf}"
            ws.cell(row=r, column=current_col+1, value=self._apply_pattern(f"=IRR({irr_range})"))
            
            ws.cell(row=r+1, column=current_col, value="MOIC").font = Font(bold=True)
            # Sum positive CFs / Absolute Initial
            moic_form = f"=SUMIFS({irr_range}, {irr_range}, \">0\")/ABS({get_column_letter(current_col+1)}{row_cf})"
            ws.cell(row=r+1, column=current_col+1, value=self._apply_pattern(moic_form))
            r += 3
            
        return r
