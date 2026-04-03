from .base_block import BaseBlock
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

class RevenueBlock(BaseBlock):
    def render(self, ws, start_row: int, current_col: int = 1) -> int:
        r = start_row
        
        # Title
        ws.cell(row=r, column=current_col, value="REVENUE BUILD").font = Font(bold=True, color='E94560')
        r += 1

        if self.complexity == 'elementary':
            price = self.parameters.get('price', 5)
            volume = self.parameters.get('volume', 100)
            
            ws.cell(row=r, column=current_col, value="Price").font = Font(bold=True)
            ws.cell(row=r, column=current_col+1, value=price)
            self.cell_map['price'] = f"{get_column_letter(current_col+1)}{r}"
            r += 1
            
            ws.cell(row=r, column=current_col, value="Volume").font = Font(bold=True)
            ws.cell(row=r, column=current_col+1, value=volume)
            self.cell_map['volume'] = f"{get_column_letter(current_col+1)}{r}"
            r += 1

            ws.cell(row=r, column=current_col, value="Total Revenue").font = Font(bold=True)
            formula = f"={self.cell_map['price']}*{self.cell_map['volume']}"
            ws.cell(row=r, column=current_col+1, value=self._apply_pattern(formula))
            self.cell_map['total_revenue'] = f"{get_column_letter(current_col+1)}{r}"
            r += 2

        elif self.complexity == 'mid_market':
            # Simple projected MRR or multi-product mix
            base_mrr = self.parameters.get('base_mrr', 50000)
            growth = self.parameters.get('growth_rate', 0.05)
            periods = self.parameters.get('periods', 5)

            ws.cell(row=r, column=current_col, value="Base MRR").font = Font(bold=True)
            ws.cell(row=r, column=current_col+1, value=base_mrr)
            self.cell_map['base_mrr'] = f"{get_column_letter(current_col+1)}{r}"
            r += 1
            
            ws.cell(row=r, column=current_col, value="Growth Rate").font = Font(bold=True)
            ws.cell(row=r, column=current_col+1, value=growth)
            self.cell_map['growth_rate'] = f"{get_column_letter(current_col+1)}{r}"
            r += 1

            for p in range(1, periods + 1):
                ws.cell(row=r, column=current_col, value=f"Year {p} Revenue")
                if p == 1:
                    formula = f"={self.cell_map['base_mrr']}*12"
                else:
                    formula = f"={self.cell_map[f'revenue_y{p-1}']}*(1+{self.cell_map['growth_rate']})"
                
                ws.cell(row=r, column=current_col+1, value=self._apply_pattern(formula))
                self.cell_map[f'revenue_y{p}'] = f"{get_column_letter(current_col+1)}{r}"
                r += 1
            r += 1

        elif self.complexity == 'institutional':
            # Needs to handle cohort bridging, churn, net retention
            # Simulating institutional logic mapping 
            arr = self.parameters.get('starting_arr', 10000000)
            net_retention = self.parameters.get('net_retention', 1.05)
            new_acv = self.parameters.get('new_acv', 2000000)
            periods = self.parameters.get('periods', 5)
            
            # Setup columns for periods
            headers = ["Metric"] + [f"Year {i}" for i in range(1, periods + 1)]
            for col_idx, h in enumerate(headers, start=current_col):
                ws.cell(row=r, column=col_idx, value=h).font = Font(bold=True)
            r += 1

            # Starting ARR
            row_arr = r
            ws.cell(row=row_arr, column=current_col, value="Starting ARR")
            ws.cell(row=row_arr+1, column=current_col, value="Net Retention (+)")
            ws.cell(row=row_arr+2, column=current_col, value="New Business ACV (+)")
            ws.cell(row=row_arr+3, column=current_col, value="Ending ARR (=)").font = Font(bold=True)
            
            for p in range(1, periods + 1):
                c_idx = current_col + p
                col_letter = get_column_letter(c_idx)
                
                if p == 1:
                    ws.cell(row=row_arr, column=c_idx, value=arr)
                else:
                    prev_col = get_column_letter(c_idx - 1)
                    ws.cell(row=row_arr, column=c_idx, value=self._apply_pattern(f"={prev_col}{row_arr+3}"))

                # Retention
                ws.cell(row=row_arr+1, column=c_idx, value=self._apply_pattern(f"={col_letter}{row_arr}*({net_retention}-1)"))
                # New ACV
                ws.cell(row=row_arr+2, column=c_idx, value=new_acv * (1.1 ** (p-1))) # Example growth on new business
                # Ending
                ws.cell(row=row_arr+3, column=c_idx, value=self._apply_pattern(f"=SUM({col_letter}{row_arr}:{col_letter}{row_arr+2})"))
                self.cell_map[f'revenue_y{p}'] = f"{col_letter}{row_arr+3}"
            
            r += 5

        return r
