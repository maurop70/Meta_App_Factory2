import pandas as pd
import numpy as np

class FinancialPlanner:
    """
    Deterministic Financial Model.
    Calculates P&L, Break-even, and ROI based on input assumptions.
    Bypasses LLM 'Advice' filters by performing pure arithmetic.
    """
    def __init__(self, output_dir=None):
        import os
        if output_dir:
            self.output_dir = output_dir
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            self.output_dir = os.path.join(base_dir, "generated_files")
        os.makedirs(self.output_dir, exist_ok=True)

    def create_live_excel_model(self, assumptions):
        """
        Generates a .xlsx file with active formulas linked to an Assumptions sheet.
        Returns the path to the generated file.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Border, Side
            import os

            wb = Workbook()
            
            # --- SHEET 1: ASSUMPTIONS ---
            ws_assumptions = wb.active
            ws_assumptions.title = "Assumptions"
            
            # Headers
            ws_assumptions["A1"] = "Parameter"
            ws_assumptions["B1"] = "Value"
            ws_assumptions["C1"] = "Description"
            ws_assumptions["A1"].font = Font(bold=True)
            ws_assumptions["B1"].font = Font(bold=True)
            
            # Data Mapping (Key -> Cell)
            # We'll map assumption keys to specific cells for formula reference
            # e.g. "price_product_a" -> Assumption!B2
            
            row = 2
            key_map = {} # stores "price_product_a": "Assumptions!B2"
            
            # Define specific order for readability
            ordered_keys = [
                "unit_cost_product_a", "price_product_a", 
                "unit_cost_product_b", "price_product_b",
                "fixed_costs_monthly", "startup_investment",
                "sales_growth_rate", "initial_sales_volume"
            ]
            
            for key in ordered_keys:
                val = float(assumptions.get(key, 0))
                cell_ref = f"B{row}"
                ws_assumptions[f"A{row}"] = key
                ws_assumptions[cell_ref] = val
                key_map[key] = f"Assumptions!{cell_ref}"
                row += 1
                
            # Formatting
            ws_assumptions.column_dimensions["A"].width = 30
            ws_assumptions.column_dimensions["B"].width = 15

            # --- SHEET 2: P&L (Formulas) ---
            ws_pnl = wb.create_sheet("P&L Year 1")
            
            # Headers
            headers = ["Month", "Volume", "Revenue", "COGS", "Gross Profit", "Fixed Costs", "Net Profit"]
            for col_num, header in enumerate(headers, 1):
                cell = ws_pnl.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)

            # Cells for Formula References
            ref_vol = key_map["initial_sales_volume"]
            ref_growth = key_map["sales_growth_rate"]
            ref_price_a = key_map["price_product_a"]
            ref_price_b = key_map["price_product_b"]
            ref_cost_a = key_map["unit_cost_product_a"]
            ref_cost_b = key_map["unit_cost_product_b"]
            ref_fixed = key_map["fixed_costs_monthly"]

            # Generate Rows for 12 Months
            for month in range(1, 13):
                r = month + 1
                growth_factor = f"((1+{ref_growth})^{month-1})" # Exponential growth formula
                
                # Month
                ws_pnl[f"A{r}"] = month
                
                # Volume (Formula: Initial * Growth^(m-1))
                # Excel: =Assumptions!B9 * (1+Assumptions!B8)^(A2-1)
                ws_pnl[f"B{r}"] = f"={ref_vol}*(1+{ref_growth})^(A{r}-1)"
                
                # Revenue (Formula: Volume * 0.5 * PriceA + Volume * 0.5 * PriceB)
                # B{r} is Volume
                ws_pnl[f"C{r}"] = f"=(B{r}*0.5*{ref_price_a})+(B{r}*0.5*{ref_price_b})"
                
                # COGS (Formula: Volume * 0.5 * CostA + Volume * 0.5 * CostB)
                ws_pnl[f"D{r}"] = f"=(B{r}*0.5*{ref_cost_a})+(B{r}*0.5*{ref_cost_b})"
                
                # Gross Profit (Revenue - COGS)
                ws_pnl[f"E{r}"] = f"=C{r}-D{r}"
                
                # Fixed Costs
                ws_pnl[f"F{r}"] = f"={ref_fixed}"
                
                # Net Profit (GP - Fixed)
                ws_pnl[f"G{r}"] = f"=E{r}-F{r}"

            # Totals Row
            last_row = 14
            ws_pnl[f"A{last_row}"] = "TOTAL"
            ws_pnl[f"A{last_row}"].font = Font(bold=True)
            for col in ["B", "C", "D", "E", "F", "G"]:
                ws_pnl[f"{col}{last_row}"] = f"=SUM({col}2:{col}13)"
                ws_pnl[f"{col}{last_row}"].font = Font(bold=True)

            # Formatting Columns
            for col in ["C", "D", "E", "F", "G"]:
                 for row in range(2, 16):
                      ws_pnl[f"{col}{row}"].number_format = '$#,##0.00'
            ws_pnl.column_dimensions["C"].width = 15
            ws_pnl.column_dimensions["E"].width = 15
            ws_pnl.column_dimensions["G"].width = 15

            # Save
            base_dir = os.path.dirname(os.path.abspath(__file__))
            # output_dir = os.path.join(base_dir, "generated_files")
            # os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(self.output_dir, "Financial_Model_Live.xlsx")
            wb.save(output_path)
            
            return output_path

        except Exception as e:
            return f"Error creating Excel model: {str(e)}"

    def generate_financial_plan(self, assumptions):
        """
        Legacy markdown creation + triggers new Excel creation.
        """
        # ... logic to generate markdown ...
        # (Keeping existing markdown logic but calling the new method too)
        
        excel_path = self.create_live_excel_model(assumptions)
        
        # ... (Rest of original markdown logic) ...
        # For brevity, I will merge this within the existing method or just update the tool handler to call this.
        # But per instruction, I will REPLACE the current method content and incorporate both.
        
        markdown_report = self._generate_markdown_report(assumptions)
        
        return f"{markdown_report}\n\n[SYSTEM] A Live Excel Model has also been generated at: {excel_path}"

    def _generate_markdown_report(self, assumptions):
        # ... (Original Logic Moved Here) ...
        try:
             # 1. Parse Assumptions
            params = {
                "unit_cost_a": float(assumptions.get("unit_cost_product_a", 0)),
                "price_a": float(assumptions.get("price_product_a", 0)),
                "unit_cost_b": float(assumptions.get("unit_cost_product_b", 0)),
                "price_b": float(assumptions.get("price_product_b", 0)),
                "fixed_costs": float(assumptions.get("fixed_costs_monthly", 0)),
                "investment": float(assumptions.get("startup_investment", 0)),
                "growth": float(assumptions.get("sales_growth_rate", 0.05)),
                "volume": int(assumptions.get("initial_sales_volume", 1000))
            }

            # 2. Monthly P&L (Year 1)
            months = list(range(1, 13))
            volume = [params["volume"] * ((1 + params["growth"]) ** i) for i in range(12)]
            
            # Simple mix: 50/50 split for A and B for estimation
            revenue = [(v * 0.5 * params["price_a"]) + (v * 0.5 * params["price_b"]) for v in volume]
            cogs = [(v * 0.5 * params["unit_cost_a"]) + (v * 0.5 * params["unit_cost_b"]) for v in volume]
            gross_profit = [r - c for r, c in zip(revenue, cogs)]
            net_profit = [gp - params["fixed_costs"] for gp in gross_profit]

            df_pnl = pd.DataFrame({
                "Month": months,
                "Volume": [int(v) for v in volume],
                "Revenue": revenue,
                "COGS": cogs,
                "Gross Profit": gross_profit,
                "Fixed Costs": [params["fixed_costs"]] * 12,
                "Net Profit": net_profit
            })
            
            # Formatting
            df_display = df_pnl.copy()
            for col in ["Revenue", "COGS", "Gross Profit", "Fixed Costs", "Net Profit"]:
                df_display[col] = df_display[col].apply(lambda x: f"${x:,.2f}")

            pnl_table = df_display.to_markdown(index=False)

            # 3. Break-Even Analysis
            avg_price = (params["price_a"] + params["price_b"]) / 2
            avg_cost = (params["unit_cost_a"] + params["unit_cost_b"]) / 2
            contribution_margin = avg_price - avg_cost
            
            if contribution_margin > 0:
                break_even_units = params["fixed_costs"] / contribution_margin
                break_even_text = f"**Break-Even Point:** {int(break_even_units)} units/month"
            else:
                break_even_text = "**Break-Even Point:** N/A (Negative Contribution Margin)"

            # 4. ROI (Simple Payback)
            total_profit_y1 = sum(net_profit)
            payback_period = "Not in Year 1"
            cumulative_cash = -params["investment"]
            for i, profit in enumerate(net_profit):
                cumulative_cash += profit
                if cumulative_cash >= 0:
                    payback_period = f"Month {i+1}"
                    break
            
            roi_text = f"**Projected ROI (Year 1):** {((total_profit_y1 / params['investment']) * 100):.1f}%" if params["investment"] > 0 else "N/A"

            report = f"""
### Financial Projection (Deterministic Model)

**Assumptions used:**
- Average Price: ${(params['price_a'] + params['price_b'])/2:.2f}
- Fixed Monthly Costs: ${params['fixed_costs']:,.2f}
- Startup Investment: ${params['investment']:,.2f}

{break_even_text}
{roi_text}
**Payback Period:** {payback_period}

#### Monthly P&L (Year 1)
{pnl_table}
            """
            return report
        except Exception as e:
            return f"Error in Markdown Report: {e}"

if __name__ == "__main__":
    fp = FinancialPlanner()
    assumptions = {
        "unit_cost_product_a": 5.00, # Gelato
        "price_product_a": 15.00,
        "unit_cost_product_b": 2.00, # Macaron
        "price_product_b": 6.00,
        "fixed_costs_monthly": 5000,
        "startup_investment": 50000,
        "sales_growth_rate": 0.05,
        "initial_sales_volume": 500
    }
    print(fp.generate_financial_plan(assumptions))
