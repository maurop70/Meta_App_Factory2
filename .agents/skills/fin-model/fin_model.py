"""
fin_model.py — Live, formula-driven Excel financial models (MAF skill: fin-model)
════════════════════════════════════════════════════════════════════════════════
Meta App Factory | Native Python Ecosystem (manifest §1: no Node.js generation)

Builds a driver-based, single-source-of-truth Excel workbook where EVERY number
downstream of the `Inputs` tab is a live Excel formula (=...), never a Python-
computed constant. Change an input and the whole model — unit economics, P&L,
cash flow, break-even, sensitivity grid, tornado, and Bear/Base/Bull scenarios —
recalculates in Excel.

Design contract (enforced):
  • One `Inputs` tab holds every assumption. Nothing hardcoded downstream.
  • Cell references, not literals, inside formulas (=Inputs!$B$5*(1+Inputs!$B$6)).
  • Color coding by role: blue=input, black=formula, green=cross-sheet link.
  • Special-character sheet names ('P&L', 'Cash & Break-Even') are ALWAYS quoted
    in cross-sheet formulas; an explicit audit verifies this.
  • Peak funding requirement = the deepest point of cumulative free cash flow
    (EBITDA − capex − ΔNWC − cash tax), with NOL-adjusted cash tax.
  • Sensitivity is built as compact LIVE closed-form formulas (not fragile Excel
    Data Tables); the centre cell ties out to the P&L EBITDA for the target year.

Verification (hard ship gate, see verify.py):
  • Recalculate with the `formulas` engine; ZERO #REF!/#DIV/0!/#VALUE!/#NAME?/#N/A.
  • Sensitivity centre / Base scenario tie out to the P&L base model.

Usage:
    from fin_model import FinModel
    result = FinModel().build(assumptions, out_path="model.xlsx")
    # result["xlsx_path"], result["summary"], result["verification"]
"""

from __future__ import annotations

import os
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Formatting constants (industry-standard) ─────────────────────────────────
FMT_CURRENCY = '$#,##0;($#,##0);"-"'      # negatives in parens, zeros as dashes
FMT_CURRENCY2 = '$#,##0.00;($#,##0.00);"-"'
FMT_PCT = '0.0%'
FMT_MULT = '0.0"x"'
FMT_INT = '#,##0;(#,##0);"-"'

BASE_FONT = "Arial"
COL_INPUT = "0000FF"     # blue  — hardcoded input
COL_CALC = "000000"      # black — formula / calc on same sheet
COL_LINK = "008000"      # green — link to another sheet
COL_EXT = "FF0000"       # red   — external link (unused here, reserved)

# Brand-neutral default banner colours (overridable via brand tokens in Phase 2)
BANNER_FILL = "1F3864"   # deep navy
BANNER_TEXT = "FFFFFF"
SECTION_FILL = "D9E1F2"  # pale blue section header
INPUT_TAB = "2E75B6"
PNL_TAB = "1F3864"
CASH_TAB = "548235"
SENS_TAB = "BF8F00"
README_TAB = "808080"


def _font(color=COL_CALC, bold=False, size=10, italic=False, color_white=False):
    return Font(name=BASE_FONT, size=size, bold=bold, italic=italic,
                color=(BANNER_TEXT if color_white else color))


# ═════════════════════════════════════════════════════════════════════════════
#  SHADOW CALCULATOR  (pure-Python mirror of the Excel formulas)
#  Produces the executive summary AND the expected values used to verify that
#  the live workbook recalculates to the same numbers ("tie out to hand est.").
# ═════════════════════════════════════════════════════════════════════════════

def _cogs_per_unit(a: dict) -> float:
    rm = sum(float(r.get("qty_per_unit", 0)) * float(r.get("unit_cost", 0))
             for r in a.get("raw_materials", []))
    line = float(a.get("line_time_minutes_per_unit", 0)) * float(a.get("line_cost_per_minute", 0))
    return (rm + line + float(a.get("labor_cost_per_unit", 0))
            + float(a.get("packaging_cost_per_unit", 0))
            + float(a.get("logistics_cost_per_unit", 0)))


def shadow_model(a: dict) -> dict:
    """Deterministic Python mirror of the workbook. Single source for the summary."""
    years = int(a.get("years", 5))
    price = float(a.get("retail_price_per_unit", 0))
    g = float(a.get("annual_growth_rate", 0))
    upm = float(a.get("units_per_month", 0))
    cogs_pu = _cogs_per_unit(a)
    opex = float(a.get("monthly_fixed_opex", 0)) * 12.0
    tax_rate = float(a.get("tax_rate", 0))
    nwc_days = float(a.get("nwc_days", 0))
    capex_items = a.get("capex_items", [])
    total_capex = sum(float(c.get("cost", 0)) for c in capex_items)
    annual_dep = sum(float(c.get("cost", 0)) / max(float(c.get("useful_life_years", 1)), 1e-9)
                     for c in capex_items)

    rows = []
    nol_bf = 0.0
    prev_nwc = 0.0
    cum_fcf = 0.0
    cum_ebitda = 0.0
    for y in range(1, years + 1):
        units = upm * 12.0 * ((1.0 + g) ** (y - 1))
        revenue = units * price
        cogs = units * cogs_pu
        gross = revenue - cogs
        ebitda = gross - opex
        dep = annual_dep
        ebit = ebitda - dep
        taxable = max(0.0, ebit - nol_bf)
        nol_used = min(nol_bf, max(0.0, ebit))
        nol_cf = nol_bf - nol_used + max(0.0, -ebit)
        cash_tax = tax_rate * taxable
        nwc = nwc_days / 365.0 * revenue
        d_nwc = nwc - prev_nwc
        capex_out = total_capex if y == 1 else 0.0
        fcf = ebitda - capex_out - cash_tax - d_nwc
        cum_fcf += fcf
        cum_ebitda += ebitda
        rows.append(dict(year=y, units=units, revenue=revenue, cogs=cogs, gross=gross,
                         opex=opex, ebitda=ebitda, dep=dep, ebit=ebit, taxable=taxable,
                         nol_bf=nol_bf, nol_cf=nol_cf, cash_tax=cash_tax, nwc=nwc,
                         d_nwc=d_nwc, capex_out=capex_out, fcf=fcf, cum_fcf=cum_fcf,
                         cum_ebitda=cum_ebitda))
        nol_bf = nol_cf
        prev_nwc = nwc

    contrib = price - cogs_pu
    be_units_annual = (opex / contrib) if contrib > 0 else float("inf")
    peak_funding = -min([0.0] + [r["cum_fcf"] for r in rows])
    ebitda_inflection = next((r["year"] for r in rows if r["ebitda"] > 0), None)
    cum_ebitda_crossover = next((r["year"] for r in rows if r["cum_ebitda"] > 0), None)

    return dict(rows=rows, cogs_pu=cogs_pu, contrib=contrib, price=price, opex=opex,
                be_units_annual=be_units_annual, peak_funding=peak_funding,
                gross_margin_pct=(contrib / price if price else 0.0),
                ebitda_inflection_year=ebitda_inflection,
                cum_ebitda_crossover_year=cum_ebitda_crossover,
                total_capex=total_capex, annual_dep=annual_dep)


def scenario_ebitda(a: dict, vol_m: float, price_m: float, cogs_m: float, year: int) -> float:
    """Closed-form EBITDA for a target year under volume/price/cogs multipliers."""
    g = float(a.get("annual_growth_rate", 0))
    upm = float(a.get("units_per_month", 0)) * vol_m
    units = upm * 12.0 * ((1.0 + g) ** (year - 1))
    price = float(a.get("retail_price_per_unit", 0)) * price_m
    cogs_pu = _cogs_per_unit(a) * cogs_m
    opex = float(a.get("monthly_fixed_opex", 0)) * 12.0
    return units * (price - cogs_pu) - opex


# ═════════════════════════════════════════════════════════════════════════════
#  WORKBOOK BUILDER
# ═════════════════════════════════════════════════════════════════════════════

PNL = "P&L"                      # ampersand → must be quoted in refs
CASH = "Cash & Break-Even"       # ampersand + spaces → must be quoted in refs
SENS = "Sensitivity"
INPUTS = "Inputs"
DRIVERS = "Unit Economics"
README = "README"


class FinModel:
    def __init__(self):
        self.ref = {}   # input name -> "$B$<row>"  (absolute, within Inputs)

    # ---- low-level helpers -------------------------------------------------
    def _inp(self, ws, row, label, value, fmt=None, note="", indent=0, key=None):
        """Write a single hardcoded input row (blue) on the Inputs sheet."""
        c_label = ws.cell(row=row, column=1, value=("   " * indent) + label)
        c_label.font = _font(COL_CALC)
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = _font(COL_INPUT)          # blue = input
        if fmt:
            cell.number_format = fmt
        if note:
            n = ws.cell(row=row, column=3, value=note)
            n.font = _font("808080", italic=True, size=9)
        addr = f"$B${row}"
        if key:
            self.ref[key] = addr
        return addr

    def _section(self, ws, row, title, span=3):
        c = ws.cell(row=row, column=1, value=title)
        c.font = _font(BANNER_TEXT, bold=True, color_white=True)
        for col in range(1, span + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=BANNER_FILL)
        ws.cell(row=row, column=1).font = _font(bold=True, color_white=True)

    def _banner(self, ws, title, subtitle, span=7):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
        c = ws.cell(row=1, column=1, value=title)
        c.font = Font(name=BASE_FONT, size=14, bold=True, color=BANNER_TEXT)
        c.fill = PatternFill("solid", fgColor=BANNER_FILL)
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.row_dimensions[1].height = 26
        for col in range(1, span + 1):
            ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor=BANNER_FILL)
        if subtitle:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
            s = ws.cell(row=2, column=1, value=subtitle)
            s.font = _font("808080", italic=True, size=9)

    def I(self, key):
        """Absolute cross-sheet reference into Inputs (green link colour applied at use)."""
        return f"{INPUTS}!{self.ref[key]}"

    # ---- public API (module-level build_and_verify is the recommended entry) -
    def build(self, assumptions: dict, out_path: str) -> dict:
        a = self._defaults(assumptions)
        shadow = shadow_model(a)
        years = int(a["years"])

        wb = Workbook()
        wb.remove(wb.active)
        ws_readme = wb.create_sheet(README)
        ws_in = wb.create_sheet(INPUTS)
        ws_ue = wb.create_sheet(DRIVERS)
        ws_pnl = wb.create_sheet(PNL)
        ws_cash = wb.create_sheet(CASH)
        ws_sens = wb.create_sheet(SENS)

        for ws, color in ((ws_readme, README_TAB), (ws_in, INPUT_TAB), (ws_ue, INPUT_TAB),
                          (ws_pnl, PNL_TAB), (ws_cash, CASH_TAB), (ws_sens, SENS_TAB)):
            ws.sheet_properties.tabColor = color
            ws.sheet_view.showGridLines = False

        self._build_inputs(ws_in, a)
        ue_rows = self._build_unit_economics(ws_ue, a)
        pnl_rows = self._build_pnl(ws_pnl, a, ue_rows, years)
        self._build_cash(ws_cash, a, pnl_rows, years)
        self._build_sensitivity(ws_sens, a, shadow, pnl_rows, years)
        self._build_readme(ws_readme, a, shadow)

        os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
        wb.save(out_path)

        summary = self._summary(a, shadow)
        audit = self.audit_sheet_refs(out_path)
        verify_targets = {
            "pnl_sheet": PNL,
            "pnl_ebitda_row": pnl_rows["ebitda"],
            "pnl_revenue_row": pnl_rows["revenue"],
            "years": years,
            "tieout_cell": self._tieout_addr,                 # "Sensitivity!B<row>"
            "scenario_tieout_cell": self._scenario_tieout_addr,
            "expected_ebitda": [round(r["ebitda"], 2) for r in shadow["rows"]],
            "expected_revenue": [round(r["revenue"], 2) for r in shadow["rows"]],
        }
        chart_targets = {"pnl": self.chart_targets_pnl, "cash": self.chart_targets_cash,
                         "sens": self.chart_targets_sens}
        return {"xlsx_path": os.path.abspath(out_path), "summary": summary,
                "shadow": shadow, "special_ref_audit": audit,
                "verify_targets": verify_targets, "chart_targets": chart_targets,
                "assumptions": a}

    # ---- defaults ----------------------------------------------------------
    def _defaults(self, a: dict) -> dict:
        a = dict(a or {})
        a.setdefault("product_name", "Venture Product")
        a.setdefault("years", 5)
        a.setdefault("units_per_month", 0)
        a.setdefault("annual_growth_rate", 0.0)
        a.setdefault("retail_price_per_unit", 0.0)
        a.setdefault("raw_materials", [])
        a.setdefault("line_time_minutes_per_unit", 0.0)
        a.setdefault("line_cost_per_minute", 0.0)
        a.setdefault("labor_cost_per_unit", 0.0)
        a.setdefault("packaging_cost_per_unit", 0.0)
        a.setdefault("logistics_cost_per_unit", 0.0)
        a.setdefault("monthly_fixed_opex", 0.0)
        a.setdefault("capex_items", [])
        a.setdefault("tax_rate", 0.21)
        a.setdefault("nwc_days", 30.0)
        a.setdefault("num_stores", 0)
        # Scenario multipliers (editable in the workbook)
        sc = a.setdefault("scenarios", {})
        sc.setdefault("bear", {"volume": 0.75, "price": 0.95, "cogs": 1.10})
        sc.setdefault("base", {"volume": 1.00, "price": 1.00, "cogs": 1.00})
        sc.setdefault("bull", {"volume": 1.30, "price": 1.05, "cogs": 0.92})
        a.setdefault("sensitivity", {})
        return a

    # ---- Inputs sheet ------------------------------------------------------
    def _build_inputs(self, ws, a):
        self._banner(ws, f"{a['product_name']} — Model Inputs",
                     "Single source of truth · BLUE cells are editable assumptions · "
                     "change any value and the whole model recalculates")
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 40
        r = 4
        self._section(ws, r, "VOLUME & PRICING"); r += 1
        self._inp(ws, r, "Starting volume (units / month)", a["units_per_month"], FMT_INT,
                  "Year-1 monthly run-rate", key="upm"); r += 1
        self._inp(ws, r, "Annual volume growth", a["annual_growth_rate"], FMT_PCT,
                  "Applied YoY; 0% = flat (no speculation)", key="growth"); r += 1
        self._inp(ws, r, "Retail price / unit", a["retail_price_per_unit"], FMT_CURRENCY2,
                  key="price"); r += 1
        self._inp(ws, r, "# Retail doors (operational driver)", a["num_stores"], FMT_INT,
                  "Used for units/store/week break-even", key="stores"); r += 2

        self._section(ws, r, "UNIT COGS  (per unit)"); r += 1
        # Raw materials sub-table
        rm_start = r
        self.ref["rm_rows"] = []
        if a["raw_materials"]:
            for m in a["raw_materials"]:
                # qty in col B, unit cost in col C, extended cost (formula) in col D
                ws.cell(row=r, column=1, value="   " + str(m.get("item", "Material"))).font = _font()
                qty = ws.cell(row=r, column=2, value=float(m.get("qty_per_unit", 0)))
                qty.font = _font(COL_INPUT); qty.number_format = "0.000"
                uc = ws.cell(row=r, column=3, value=float(m.get("unit_cost", 0)))
                uc.font = _font(COL_INPUT); uc.number_format = FMT_CURRENCY2
                ext = ws.cell(row=r, column=4, value=f"=B{r}*C{r}")
                ext.font = _font(COL_CALC); ext.number_format = FMT_CURRENCY2
                self.ref["rm_rows"].append(r)
                r += 1
        else:
            ws.cell(row=r, column=1, value="   (no raw materials provided)").font = _font("808080", italic=True)
            r += 1
        ws.column_dimensions["D"].width = 14
        self._inp(ws, r, "Line time (minutes / unit)", a["line_time_minutes_per_unit"], "0.0",
                  key="line_min"); r += 1
        self._inp(ws, r, "Line cost / minute", a["line_cost_per_minute"], FMT_CURRENCY2,
                  key="line_cpm"); r += 1
        self._inp(ws, r, "Direct labor / unit", a["labor_cost_per_unit"], FMT_CURRENCY2,
                  key="labor"); r += 1
        self._inp(ws, r, "Packaging / unit", a["packaging_cost_per_unit"], FMT_CURRENCY2,
                  key="pack"); r += 1
        self._inp(ws, r, "Logistics / unit", a["logistics_cost_per_unit"], FMT_CURRENCY2,
                  key="logi"); r += 2

        self._section(ws, r, "OPERATING & CAPITAL"); r += 1
        self._inp(ws, r, "Fixed OpEx / month", a["monthly_fixed_opex"], FMT_CURRENCY,
                  "Rent, salaries, software", key="opex_m"); r += 1
        self._inp(ws, r, "Cash tax rate", a["tax_rate"], FMT_PCT,
                  "Applied to NOL-adjusted EBIT", key="tax"); r += 1
        self._inp(ws, r, "Net working capital (days of revenue)", a["nwc_days"], "0",
                  "Drives ΔNWC cash drag", key="nwc"); r += 1
        # CapEx table: cost (B), life (C), annual dep (D formula)
        capex_rows = []
        self.ref["capex_rows"] = capex_rows
        ws.cell(row=r, column=1, value="CapEx items").font = _font(bold=True); r += 1
        if a["capex_items"]:
            for c in a["capex_items"]:
                ws.cell(row=r, column=1, value="   " + str(c.get("item", "CapEx"))).font = _font()
                cost = ws.cell(row=r, column=2, value=float(c.get("cost", 0)))
                cost.font = _font(COL_INPUT); cost.number_format = FMT_CURRENCY
                life = ws.cell(row=r, column=3, value=float(c.get("useful_life_years", 1)))
                life.font = _font(COL_INPUT); life.number_format = "0.0"
                dep = ws.cell(row=r, column=4, value=f"=B{r}/C{r}")
                dep.font = _font(COL_CALC); dep.number_format = FMT_CURRENCY
                capex_rows.append(r)
                r += 1
        else:
            ws.cell(row=r, column=1, value="   (no CapEx provided)").font = _font("808080", italic=True)
            r += 1
        r += 1

        self._section(ws, r, "SCENARIO MULTIPLIERS  (Bear / Base / Bull)"); r += 1
        ws.cell(row=r, column=2, value="Volume").font = _font(bold=True)
        ws.cell(row=r, column=3, value="Price").font = _font(bold=True)
        ws.cell(row=r, column=4, value="COGS").font = _font(bold=True)
        r += 1
        for name in ("bear", "base", "bull"):
            sc = a["scenarios"][name]
            ws.cell(row=r, column=1, value="   " + name.capitalize()).font = _font()
            for col, kk in ((2, "volume"), (3, "price"), (4, "cogs")):
                cell = ws.cell(row=r, column=col, value=float(sc[kk]))
                cell.font = _font(COL_INPUT); cell.number_format = FMT_MULT
            self.ref[f"sc_{name}"] = {"volume": f"$B${r}", "price": f"$C${r}", "cogs": f"$D${r}"}
            r += 1

    # ---- Unit Economics sheet ---------------------------------------------
    def _build_unit_economics(self, ws, a):
        self._banner(ws, "Unit Economics", "Per-unit contribution build · green = link to Inputs", span=4)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 16
        r = 4
        ws.cell(row=r, column=1, value="Per-unit COGS").font = _font(bold=True); r += 1
        rm_cells = []
        # Raw materials: pull extended cost (col D) from Inputs for each material row
        for src_row in self.ref.get("rm_rows", []):
            ws.cell(row=r, column=1, value="   Raw material").font = _font()
            c = ws.cell(row=r, column=2, value=f"={INPUTS}!$D${src_row}")
            c.font = _font(COL_LINK); c.number_format = FMT_CURRENCY2
            rm_cells.append(f"B{r}")
            r += 1
        rm_sum_row = r
        ws.cell(row=r, column=1, value="   Raw materials subtotal").font = _font()
        if rm_cells:
            c = ws.cell(row=r, column=2, value="=SUM(" + ",".join(rm_cells) + ")")
        else:
            c = ws.cell(row=r, column=2, value=0)
        c.font = _font(COL_CALC); c.number_format = FMT_CURRENCY2
        r += 1
        # Line time, labor, packaging, logistics (links to Inputs)
        def linkrow(label, expr):
            nonlocal r
            ws.cell(row=r, column=1, value="   " + label).font = _font()
            c = ws.cell(row=r, column=2, value=expr)
            c.font = _font(COL_LINK); c.number_format = FMT_CURRENCY2
            row_used = r
            r += 1
            return row_used
        lt = linkrow("Line time", f"={self.I('line_min')}*{self.I('line_cpm')}")
        lab = linkrow("Direct labor", f"={self.I('labor')}")
        pk = linkrow("Packaging", f"={self.I('pack')}")
        lo = linkrow("Logistics", f"={self.I('logi')}")
        cogs_row = r
        ws.cell(row=r, column=1, value="Total COGS / unit").font = _font(bold=True)
        c = ws.cell(row=r, column=2,
                    value=f"=B{rm_sum_row}+B{lt}+B{lab}+B{pk}+B{lo}")
        c.font = _font(COL_CALC, bold=True); c.number_format = FMT_CURRENCY2
        r += 2
        price_row = r
        ws.cell(row=r, column=1, value="Retail price / unit").font = _font()
        c = ws.cell(row=r, column=2, value=f"={self.I('price')}")
        c.font = _font(COL_LINK); c.number_format = FMT_CURRENCY2
        r += 1
        contrib_row = r
        ws.cell(row=r, column=1, value="Contribution / unit").font = _font(bold=True)
        c = ws.cell(row=r, column=2, value=f"=B{price_row}-B{cogs_row}")
        c.font = _font(COL_CALC, bold=True); c.number_format = FMT_CURRENCY2
        r += 1
        gm_row = r
        ws.cell(row=r, column=1, value="Gross margin %").font = _font(bold=True)
        c = ws.cell(row=r, column=2, value=f"=B{contrib_row}/B{price_row}")
        c.font = _font(COL_CALC, bold=True); c.number_format = FMT_PCT
        r += 1
        # stash addresses used by Cash & Sensitivity sheets
        self._ue_cogs = f"B{cogs_row}"
        self._ue_contrib_addr = f"B{contrib_row}"
        return {"cogs_pu": f"B{cogs_row}", "contrib": f"B{contrib_row}",
                "price": f"B{price_row}", "gm": f"B{gm_row}", "sheet": DRIVERS}

    # ---- P&L sheet ---------------------------------------------------------
    def _build_pnl(self, ws, a, ue, years):
        self._banner(ws, "Income Statement  (P&L)",
                     "All values link from Inputs & Unit Economics · green = cross-sheet link", span=years + 1)
        ws.column_dimensions["A"].width = 30
        for y in range(years):
            ws.column_dimensions[get_column_letter(2 + y)].width = 15
        # Year header row (years as TEXT strings)
        hdr = 4
        ws.cell(row=hdr, column=1, value="($, fiscal year)").font = _font(bold=True)
        for y in range(years):
            c = ws.cell(row=hdr, column=2 + y, value=f"Year {y + 1}")
            c.font = _font(bold=True); c.alignment = Alignment(horizontal="right")
        rows = {}

        def line(name, label, fmt, fn, bold=False, link=False):
            nonlocal_r = line.r
            rows[name] = nonlocal_r          # register row BEFORE cells (self-refs need it)
            ws.cell(row=nonlocal_r, column=1, value=label).font = _font(bold=bold)
            for y in range(years):
                col = get_column_letter(2 + y)
                cell = ws.cell(row=nonlocal_r, column=2 + y, value=fn(y, col))
                cell.font = _font(COL_LINK if link else COL_CALC, bold=bold)
                cell.number_format = fmt
            line.r += 1
        line.r = hdr + 1

        ue_cogs = f"'{DRIVERS}'!{ue['cogs_pu']}"
        ue_price = f"'{DRIVERS}'!{ue['price']}"

        # Units: Y1 = upm*12 ; Yn = prev*(1+growth)
        def units_fn(y, col):
            if y == 0:
                return f"={self.I('upm')}*12"
            prev = get_column_letter(2 + y - 1)
            return f"={prev}{rows['units']}*(1+{self.I('growth')})"
        line("units", "Units sold", FMT_INT, units_fn, link=True)
        line("revenue", "Revenue", FMT_CURRENCY,
             lambda y, col: f"={col}{rows['units']}*{ue_price}", link=True)
        line("cogs", "COGS", FMT_CURRENCY,
             lambda y, col: f"=-{col}{rows['units']}*{ue_cogs}", link=True)
        line("gross", "Gross profit", FMT_CURRENCY,
             lambda y, col: f"={col}{rows['revenue']}+{col}{rows['cogs']}", bold=True)
        line("opex", "Fixed OpEx", FMT_CURRENCY,
             lambda y, col: f"=-{self.I('opex_m')}*12", link=True)
        line("ebitda", "EBITDA", FMT_CURRENCY,
             lambda y, col: f"={col}{rows['gross']}+{col}{rows['opex']}", bold=True)
        # Depreciation = SUM of Inputs CapEx annual dep (col D)
        dep_cells = "+".join(f"{INPUTS}!$D${cr}" for cr in self.ref.get("capex_rows", [])) or "0"
        line("dep", "Depreciation", FMT_CURRENCY,
             lambda y, col: f"=-({dep_cells})", link=bool(self.ref.get("capex_rows")))
        line("ebit", "EBIT", FMT_CURRENCY,
             lambda y, col: f"={col}{rows['ebitda']}+{col}{rows['dep']}", bold=True)
        # NOL carryforward block. nol_bf (year n) pulls prior-year nol_cf, which is
        # placed two rows below → compute the row numbers up front for the forward ref.
        nol_bf_row = line.r
        nol_cf_row = line.r + 2
        line("nol_bf", "NOL b/f", FMT_CURRENCY,
             lambda y, col: ("=0" if y == 0
                             else f"={get_column_letter(2 + y - 1)}{nol_cf_row}"))
        line("taxable", "Taxable income", FMT_CURRENCY,
             lambda y, col: f"=MAX(0,{col}{rows['ebit']}-{col}{rows['nol_bf']})")
        line("nol_cf", "NOL c/f", FMT_CURRENCY,
             lambda y, col: (f"={col}{rows['nol_bf']}-MIN({col}{rows['nol_bf']},"
                             f"MAX(0,{col}{rows['ebit']}))+MAX(0,-{col}{rows['ebit']})"))
        line("tax", "Cash tax", FMT_CURRENCY,
             lambda y, col: f"=-{self.I('tax')}*{col}{rows['taxable']}", link=True)
        line("ni", "Net income", FMT_CURRENCY,
             lambda y, col: f"={col}{rows['ebit']}+{col}{rows['tax']}", bold=True)
        rows["_sheet"] = PNL
        self.chart_targets_pnl = {"sheet": PNL, "hdr": hdr, "units": rows["units"],
                                  "revenue": rows["revenue"], "ebitda": rows["ebitda"],
                                  "first_col": 2, "years": years}
        return rows

    # ---- Cash Flow & Break-Even sheet -------------------------------------
    def _build_cash(self, ws, a, pnl, years):
        self._banner(ws, "Cash Flow & Break-Even",
                     "Free cash flow, peak funding (deepest cumulative FCF) & break-even",
                     span=years + 1)
        ws.column_dimensions["A"].width = 32
        for y in range(years):
            ws.column_dimensions[get_column_letter(2 + y)].width = 15
        hdr = 4
        ws.cell(row=hdr, column=1, value="($, fiscal year)").font = _font(bold=True)
        for y in range(years):
            c = ws.cell(row=hdr, column=2 + y, value=f"Year {y + 1}")
            c.font = _font(bold=True); c.alignment = Alignment(horizontal="right")
        rows = {}

        def line(name, label, fmt, fn, bold=False, link=False):
            r = line.r
            rows[name] = r                   # register row BEFORE cells (self-refs need it)
            ws.cell(row=r, column=1, value=label).font = _font(bold=bold)
            for y in range(years):
                col = get_column_letter(2 + y)
                cell = ws.cell(row=r, column=2 + y, value=fn(y, col))
                cell.font = _font(COL_LINK if link else COL_CALC, bold=bold)
                cell.number_format = fmt
            line.r += 1
        line.r = hdr + 1

        line("ebitda", "EBITDA", FMT_CURRENCY,
             lambda y, col: f"='{PNL}'!{col}{pnl['ebitda']}", link=True)
        line("tax", "Cash tax", FMT_CURRENCY,
             lambda y, col: f"='{PNL}'!{col}{pnl['tax']}", link=True)
        # CapEx: total in Year 1 only
        capex_sum = "+".join(f"{INPUTS}!$B${cr}" for cr in self.ref.get("capex_rows", [])) or "0"
        line("capex", "CapEx", FMT_CURRENCY,
             lambda y, col: (f"=-({capex_sum})" if y == 0 else "=0"),
             link=bool(self.ref.get("capex_rows")))
        # NWC = nwc_days/365 * revenue ; ΔNWC = NWC_y - NWC_{y-1}
        line("nwc", "Net working capital", FMT_CURRENCY,
             lambda y, col: f"={self.I('nwc')}/365*'{PNL}'!{col}{pnl['revenue']}", link=True)
        line("dnwc", "Less: ΔNWC", FMT_CURRENCY,
             lambda y, col: (f"=-{col}{rows['nwc']}" if y == 0
                             else f"=-({col}{rows['nwc']}-{get_column_letter(2 + y - 1)}{rows['nwc']})"))
        line("fcf", "Free cash flow", FMT_CURRENCY,
             lambda y, col: (f"={col}{rows['ebitda']}+{col}{rows['tax']}"
                             f"+{col}{rows['capex']}+{col}{rows['dnwc']}"), bold=True)
        line("cum", "Cumulative FCF", FMT_CURRENCY,
             lambda y, col: (f"={col}{rows['fcf']}" if y == 0
                             else f"={get_column_letter(2 + y - 1)}{rows['cum']}+{col}{rows['fcf']}"),
             bold=True)
        line("cumebitda", "Cumulative EBITDA", FMT_CURRENCY,
             lambda y, col: (f"={col}{rows['ebitda']}" if y == 0
                             else f"={get_column_letter(2 + y - 1)}{rows['cumebitda']}+{col}{rows['ebitda']}"))

        # ---- Break-even & funding KPI block ----
        r = line.r + 1
        last = get_column_letter(2 + years - 1)
        cum_range = f"B{rows['cum']}:{last}{rows['cum']}"
        self._section(ws, r, "KEY METRICS", span=2); r += 1

        def kpi(label, formula, fmt, note=""):
            nonlocal r
            ws.cell(row=r, column=1, value=label).font = _font(bold=True)
            c = ws.cell(row=r, column=2, value=formula)
            c.font = _font(COL_CALC, bold=True); c.number_format = fmt
            if note:
                ws.cell(row=r, column=3, value=note).font = _font("808080", italic=True, size=9)
            r += 1
        # contribution & break-even reference Unit Economics contribution cell
        contrib_ref = f"'{DRIVERS}'!{self._ue_contrib}"
        kpi("Contribution / unit", f"={contrib_ref}", FMT_CURRENCY2)
        be_units = f"=({self.I('opex_m')}*12)/{contrib_ref}"
        kpi("Break-even units / year", be_units, FMT_INT, "Annual fixed OpEx ÷ contribution")
        be_row = r - 1
        kpi("Break-even units / month", f"=B{be_row}/12", FMT_INT)
        # operational driver: units/store/week (guard divide-by-zero with stores)
        kpi("Break-even units / door / week",
            f"=IF({self.I('stores')}=0,0,B{be_row}/{self.I('stores')}/52)", "0.0",
            "Operational driver")
        kpi("Peak funding requirement", f"=-MIN(0,MIN({cum_range}))", FMT_CURRENCY,
            "Deepest point of cumulative FCF")
        kpi("Cumulative FCF, Year " + str(years),
            f"={last}{rows['cum']}", FMT_CURRENCY)
        rows["_sheet"] = CASH
        self._cash_rows = rows
        self.chart_targets_cash = {"sheet": CASH, "hdr": hdr, "fcf": rows["fcf"],
                                   "cum": rows["cum"], "first_col": 2, "years": years}
        return rows

    # ---- Sensitivity sheet -------------------------------------------------
    def _build_sensitivity(self, ws, a, shadow, pnl, years):
        self._banner(ws, "Sensitivity & Scenarios",
                     "Live closed-form grid · centre cell ties out to P&L EBITDA · "
                     "tornado & Bear/Base/Bull", span=8)
        ws.column_dimensions["A"].width = 26
        for col in "BCDEFGH":
            ws.column_dimensions[col].width = 14
        ty = int(a["sensitivity"].get("target_year", years))   # target year for metrics
        base_upm = float(a["units_per_month"])
        base_cogs = _cogs_per_unit(a)

        # sweep values centred on base (centre column/row = base, guarantees tie-out)
        row_vals = a["sensitivity"].get("row_values") or \
            [round(base_upm * m) for m in (0.8, 0.9, 1.0, 1.1, 1.2)]
        col_vals = a["sensitivity"].get("col_values") or \
            [round(base_cogs * m, 4) for m in (0.9, 0.95, 1.0, 1.05, 1.1)]
        # force exact base into centre so the tie-out cell is exactly 0
        row_vals[len(row_vals) // 2] = base_upm
        col_vals[len(col_vals) // 2] = round(base_cogs, 6)

        r = 4
        ws.cell(row=r, column=1,
                value=f"2-Way: EBITDA in Year {ty}  (rows = volume/mo · cols = COGS/unit)").font = _font(bold=True)
        r += 1
        top = r
        ws.cell(row=r, column=1, value="Units/mo ↓ \\ COGS →").font = _font(italic=True, size=9)
        for j, cv in enumerate(col_vals):
            c = ws.cell(row=r, column=2 + j, value=cv)
            c.font = _font(COL_INPUT); c.number_format = FMT_CURRENCY2
        r += 1
        grid_top = r
        price_ref = self.I("price")
        growth_ref = self.I("growth")
        opex_ref = self.I("opex_m")
        centre_addr = None
        for i, rv in enumerate(row_vals):
            rc = ws.cell(row=r, column=1, value=rv)
            rc.font = _font(COL_INPUT); rc.number_format = FMT_INT
            for j, cv in enumerate(col_vals):
                col = get_column_letter(2 + j)
                # units_ty = rowunits*12*(1+growth)^(ty-1) ; EBITDA = units*(price-cogs)-opex*12
                units_expr = f"$A{r}*12*(1+{growth_ref})^({ty}-1)"
                formula = f"={units_expr}*({price_ref}-{col}${top})-{opex_ref}*12"
                cell = ws.cell(row=r, column=2 + j, value=formula)
                cell.font = _font(COL_CALC); cell.number_format = FMT_CURRENCY
                if i == len(row_vals) // 2 and j == len(col_vals) // 2:
                    centre_addr = f"{col}{r}"
            r += 1
        # tie-out check: centre cell − P&L EBITDA(target year) must be 0
        pnl_ebitda_ty = f"'{PNL}'!{get_column_letter(2 + ty - 1)}{pnl['ebitda']}"
        r += 1
        ws.cell(row=r, column=1, value="Tie-out check (must be 0)").font = _font(bold=True)
        tc = ws.cell(row=r, column=2, value=f"={centre_addr}-{pnl_ebitda_ty}")
        tc.font = _font(COL_CALC, bold=True); tc.number_format = FMT_CURRENCY2
        self._tieout_addr = f"{SENS}!B{r}"
        r += 2

        # ---- Tornado (±15% one driver at a time, sorted by swing → funnel) ----
        ws.cell(row=r, column=1, value=f"Tornado — EBITDA Year {ty} sensitivity (±15%)").font = _font(bold=True)
        r += 1
        ws.cell(row=r, column=1, value="Driver").font = _font(bold=True)
        ws.cell(row=r, column=2, value="Low (−15%)").font = _font(bold=True)
        ws.cell(row=r, column=3, value="High (+15%)").font = _font(bold=True)
        ws.cell(row=r, column=4, value="Swing").font = _font(bold=True)
        r += 1
        # closed-form EBITDA with a multiplier applied to one driver
        def ebitda_expr(vol_m="1", price_m="1", cogs_m="1"):
            units = f"{self.I('upm')}*{vol_m}*12*(1+{growth_ref})^({ty}-1)"
            price = f"{self.I('price')}*{price_m}"
            cogs = f"({base_cogs_ref})*{cogs_m}"
            return f"{units}*({price}-{cogs})-{opex_ref}*12"
        base_cogs_ref = f"'{DRIVERS}'!{self._ue_cogs}"
        drivers = [
            ("Volume", dict(vol_m="0.85"), dict(vol_m="1.15")),
            ("Price", dict(price_m="0.85"), dict(price_m="1.15")),
            ("COGS / unit", dict(cogs_m="1.15"), dict(cogs_m="0.85")),
            ("OpEx", None, None),  # handled specially (opex affects subtractor)
        ]
        # estimate swing magnitude in Python to order rows as a funnel
        def py_eb(vol=1, price=1, cogs=1, opex=1):
            units = base_upm * vol * 12 * ((1 + float(a["annual_growth_rate"])) ** (ty - 1))
            return units * (float(a["retail_price_per_unit"]) * price - base_cogs * cogs) \
                - float(a["monthly_fixed_opex"]) * 12 * opex
        swing_est = {
            "Volume": abs(py_eb(vol=1.15) - py_eb(vol=0.85)),
            "Price": abs(py_eb(price=1.15) - py_eb(price=0.85)),
            "COGS / unit": abs(py_eb(cogs=0.85) - py_eb(cogs=1.15)),
            "OpEx": abs(py_eb(opex=0.85) - py_eb(opex=1.15)),
        }
        order = sorted(swing_est, key=lambda k: swing_est[k], reverse=True)
        tornado_first = r
        for name in order:
            ws.cell(row=r, column=1, value=name).font = _font()
            if name == "OpEx":
                low = f"={ebitda_expr()}+{opex_ref}*12*0.15"   # opex −15% → +EBITDA
                high = f"={ebitda_expr()}-{opex_ref}*12*0.15"
            elif name == "Volume":
                low = f"={ebitda_expr(vol_m='0.85')}"; high = f"={ebitda_expr(vol_m='1.15')}"
            elif name == "Price":
                low = f"={ebitda_expr(price_m='0.85')}"; high = f"={ebitda_expr(price_m='1.15')}"
            else:  # COGS
                low = f"={ebitda_expr(cogs_m='1.15')}"; high = f"={ebitda_expr(cogs_m='0.85')}"
            cl = ws.cell(row=r, column=2, value=low); cl.font = _font(COL_CALC); cl.number_format = FMT_CURRENCY
            ch = ws.cell(row=r, column=3, value=high); ch.font = _font(COL_CALC); ch.number_format = FMT_CURRENCY
            cs = ws.cell(row=r, column=4, value=f"=ABS(C{r}-B{r})")
            cs.font = _font(COL_CALC, bold=True); cs.number_format = FMT_CURRENCY
            r += 1
        tornado_last = r - 1
        r += 1

        # ---- Bear / Base / Bull (driven by editable Inputs multipliers) ----
        ws.cell(row=r, column=1, value=f"Scenarios — EBITDA Year {ty}").font = _font(bold=True)
        r += 1
        ws.cell(row=r, column=1, value="Scenario").font = _font(bold=True)
        ws.cell(row=r, column=2, value="EBITDA").font = _font(bold=True)
        r += 1
        scen_first = r
        for name in ("bear", "base", "bull"):
            sc = self.ref[f"sc_{name}"]
            ws.cell(row=r, column=1, value=name.capitalize()).font = _font()
            vol = f"{INPUTS}!{sc['volume']}"; pr = f"{INPUTS}!{sc['price']}"; cg = f"{INPUTS}!{sc['cogs']}"
            formula = (f"={self.I('upm')}*{vol}*12*(1+{growth_ref})^({ty}-1)"
                       f"*({self.I('price')}*{pr}-('{DRIVERS}'!{self._ue_cogs})*{cg})-{opex_ref}*12")
            c = ws.cell(row=r, column=2, value=formula)
            c.font = _font(COL_LINK); c.number_format = FMT_CURRENCY
            r += 1
        scen_last = r - 1
        # Base scenario tie-out check
        ws.cell(row=r, column=1, value="Base ties to P&L (must be 0)").font = _font(bold=True)
        bc = ws.cell(row=r, column=2, value=f"=B{scen_first + 1}-{pnl_ebitda_ty}")
        bc.font = _font(COL_CALC, bold=True); bc.number_format = FMT_CURRENCY2
        self._scenario_tieout_addr = f"{SENS}!B{r}"
        self.chart_targets_sens = {
            "sheet": SENS, "tornado_label_col": "A", "tornado_swing_col": "D",
            "tornado_first_row": tornado_first, "tornado_last_row": tornado_last,
            "scen_label_col": "A", "scen_val_col": "B",
            "scen_first_row": scen_first, "scen_last_row": scen_last,
        }

    # ---- README ------------------------------------------------------------
    def _build_readme(self, ws, a, shadow):
        self._banner(ws, f"{a['product_name']} — Financial Model", "Driver-based · fully live · "
                     "edit only the BLUE cells on Inputs", span=2)
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 80
        r = 4
        notes = [
            ("Purpose", "Five-year, formula-driven operating & cash model with break-even, "
                        "peak-funding and sensitivity analysis."),
            ("How to use", "Every assumption lives on the Inputs tab (blue cells). Change any "
                           "value; all downstream tabs recalculate automatically. Do not hardcode "
                           "numbers in calc tabs."),
            ("Colour legend", "Blue = input · Black = formula · Green = link to another sheet."),
            ("Tab flow", "README → Inputs → Unit Economics → P&L → Cash Flow & Break-Even → Sensitivity."),
            ("Conservatism", "Fixed OpEx is held flat (not scaled with volume); growth applies only "
                             "if a non-zero rate is entered; cash tax is NOL-adjusted. These are "
                             "deliberately conservative assumptions."),
            ("Peak funding", "Defined as the deepest point of cumulative free cash flow "
                             "(EBITDA − CapEx − ΔNWC − cash tax) — the maximum capital you must "
                             "raise before the business self-funds."),
        ]
        for label, txt in notes:
            ws.cell(row=r, column=1, value=label).font = _font(bold=True)
            c = ws.cell(row=r, column=2, value=txt)
            c.font = _font(); c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 42
            r += 1

    # convenience addresses set during unit-economics build
    @property
    def _ue_contrib(self):
        return self.__dict__.get("_ue_contrib_addr", "B0")

    # ---- special-character sheet-reference audit ---------------------------
    def audit_sheet_refs(self, xlsx_path: str) -> dict:
        """Scan every formula for UNQUOTED references to special-character sheets.
        Unquoted 'P&L'/'Cash & Break-Even' refs silently become #NAME? — fail loud."""
        from openpyxl import load_workbook
        import re
        wb = load_workbook(xlsx_path)
        specials = [s for s in wb.sheetnames if not re.match(r"^[A-Za-z0-9_]+$", s)]
        violations = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        for sp in specials:
                            # bad: bare SheetName! without surrounding quotes
                            if re.search(r"(?<!')" + re.escape(sp) + r"!", v) and f"'{sp}'!" not in v:
                                violations.append({"sheet": ws.title, "cell": cell.coordinate,
                                                   "formula": v, "missing_quote_for": sp})
        return {"special_sheets": specials, "violations": violations,
                "clean": len(violations) == 0}

    # ---- executive summary -------------------------------------------------
    def _summary(self, a, shadow) -> dict:
        rows = shadow["rows"]
        scen = {n: scenario_ebitda(a, a["scenarios"][n]["volume"], a["scenarios"][n]["price"],
                                   a["scenarios"][n]["cogs"], int(a["sensitivity"].get("target_year", a["years"])))
                for n in ("bear", "base", "bull")}
        return {
            "product_name": a["product_name"],
            "currency": "USD",
            "gross_margin_pct": round(shadow["gross_margin_pct"] * 100, 1),
            "ebitda_by_year": [round(r["ebitda"]) for r in rows],
            "ebitda_inflection_year": shadow["ebitda_inflection_year"],
            "cumulative_ebitda_crossover_year": shadow["cum_ebitda_crossover_year"],
            "peak_funding_requirement": round(shadow["peak_funding"]),
            "break_even_units_per_month": (None if shadow["be_units_annual"] == float("inf")
                                           else round(shadow["be_units_annual"] / 12)),
            "scenario_ebitda_target_year": {k: round(v) for k, v in scen.items()},
            "scenario_range": [round(min(scen.values())), round(max(scen.values()))],
        }


def build_and_verify(assumptions: dict, out_path: str) -> dict:
    """Recommended entrypoint: build the live workbook AND run the mandatory
    recalc / tie-out gate. The skill never reports 'done' without verification.
    Returns the build result with a `verification` block; `verification.passed`
    is the hard ship gate (zero formula errors + sensitivity/scenario tie-outs).
    """
    res = FinModel().build(assumptions, out_path)
    try:
        from .verify import run_full_check
    except ImportError:
        from verify import run_full_check
    res["verification"] = run_full_check(res)
    res.pop("shadow", None)   # large internal mirror; not part of the contract
    return res
