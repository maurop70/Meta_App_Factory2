"""
fin_model_presentation.py — make a fin-model boardroom-ready (MAF skill)
════════════════════════════════════════════════════════════════════════
Adds, on top of a CORRECT fin-model workbook:
  • A README/cover that states purpose, how-to-use, distribution assumptions
    and an honest conservatism note (enriched in place).
  • NATIVE, EDITABLE in-workbook charts (openpyxl) coloured from brand tokens:
    revenue-bars + EBITDA-line combo · cumulative-FCF line · sorted tornado ·
    scenario columns. Series colours come from the palette, not library defaults.
  • Clean print/PDF setup: landscape, fit-to-width, sensible margins, and print
    areas extended to include the charts (so they are never silently clipped).
  • Brand-consistent section banners + tab colours so the workbook reads as ONE
    document.
  • A board PDF (reportlab) that visually mirrors the spreadsheet's palette and
    grid — same colour-coding legend, same number formats, same structure.

QA: render every PDF page to PNG (pdfplumber, no Poppler needed) and inspect for
overflow / clipping / wrong colours; loop until clean (see qa.py).
"""

from __future__ import annotations

import os
import sys

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

try:
    from .brand_tokens import load_tokens, BrandTokens
except ImportError:
    from brand_tokens import load_tokens, BrandTokens

# Reuse fin-model's Python shadow to recompute exact (verified) values for the PDF.
_FM_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "fin-model")
if _FM_DIR not in sys.path:
    sys.path.insert(0, _FM_DIR)
from fin_model import shadow_model, scenario_ebitda, _cogs_per_unit  # noqa: E402

DASH = "Dashboard"
DEFAULT_BANNER = "1F3864"   # fin-model default banner fill to detect & recolor
DEFAULT_SECTION = "D9E1F2"


# ═══════════════════════════════════════════════════════════════════════════
#  NATIVE WORKBOOK CHARTS + PRINT SETUP + BRAND RECOLOR
# ═══════════════════════════════════════════════════════════════════════════

def _style_series_fill(series, hex_rgb):
    series.graphicalProperties = GraphicalProperties(solidFill=hex_rgb)


def _style_series_line(series, hex_rgb, width_emu=28000):
    gp = GraphicalProperties()
    gp.line = LineProperties(solidFill=hex_rgb, w=width_emu)
    gp.noFill = True
    series.graphicalProperties = gp
    series.smooth = False


def add_native_charts(wb, chart_targets, tokens: BrandTokens):
    """Insert a Dashboard sheet with four editable charts referencing live cells."""
    if DASH in wb.sheetnames:
        del wb[DASH]
    ws = wb.create_sheet(DASH, 1)   # right after README
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = BrandTokens.rgb(tokens.primary)
    # banner
    ws.merge_cells("A1:N1")
    c = ws.cell(row=1, column=1, value="Board Dashboard")
    c.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BrandTokens.rgb(tokens.primary))
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    pnl = chart_targets["pnl"]; cash = chart_targets["cash"]; sens = chart_targets["sens"]
    ws_pnl = wb[pnl["sheet"]]; ws_cash = wb[cash["sheet"]]; ws_sens = wb[sens["sheet"]]
    years = pnl["years"]
    last_col = 1 + years

    # 1) Revenue bars + EBITDA line combo
    bar = BarChart(); bar.type = "col"; bar.title = "Revenue & EBITDA"
    bar.height = 7.5; bar.width = 15
    rev = Reference(ws_pnl, min_col=1, max_col=last_col, min_row=pnl["revenue"], max_row=pnl["revenue"])
    bar.add_data(rev, from_rows=True, titles_from_data=True)
    cats = Reference(ws_pnl, min_col=2, max_col=last_col, min_row=pnl["hdr"], max_row=pnl["hdr"])
    bar.set_categories(cats)
    _style_series_fill(bar.series[0], BrandTokens.rgb(tokens.secondary))
    line = LineChart()
    eb = Reference(ws_pnl, min_col=1, max_col=last_col, min_row=pnl["ebitda"], max_row=pnl["ebitda"])
    line.add_data(eb, from_rows=True, titles_from_data=True)
    line.set_categories(cats)
    line.y_axis.axId = 200
    line.y_axis.crosses = "max"
    _style_series_line(line.series[0], BrandTokens.rgb(tokens.primary))
    bar += line
    bar.y_axis.title = "Revenue ($)"
    line.y_axis.title = "EBITDA ($)"
    ws.add_chart(bar, "A3")

    # 2) Cumulative FCF line
    fcfc = LineChart(); fcfc.title = "Cumulative Free Cash Flow"; fcfc.height = 7.5; fcfc.width = 15
    cum = Reference(ws_cash, min_col=1, max_col=last_col, min_row=cash["cum"], max_row=cash["cum"])
    fcfc.add_data(cum, from_rows=True, titles_from_data=True)
    catsc = Reference(ws_cash, min_col=2, max_col=last_col, min_row=cash["hdr"], max_row=cash["hdr"])
    fcfc.set_categories(catsc)
    _style_series_line(fcfc.series[0], BrandTokens.rgb(tokens.primary))
    ws.add_chart(fcfc, "J3")

    # 3) Tornado (horizontal bars; rows already sorted into a funnel)
    tor = BarChart(); tor.type = "bar"; tor.title = "Tornado — EBITDA sensitivity"
    tor.height = 7.5; tor.width = 15
    swing = Reference(ws_sens, min_col=4, max_col=4,
                      min_row=sens["tornado_first_row"], max_row=sens["tornado_last_row"])
    tor.add_data(swing, titles_from_data=False)
    tlabels = Reference(ws_sens, min_col=1, max_col=1,
                        min_row=sens["tornado_first_row"], max_row=sens["tornado_last_row"])
    tor.set_categories(tlabels)
    _style_series_fill(tor.series[0], BrandTokens.rgb(tokens.secondary))
    tor.legend = None
    ws.add_chart(tor, "A18")

    # 4) Scenario columns
    sc = BarChart(); sc.type = "col"; sc.title = "Scenarios — EBITDA (target year)"
    sc.height = 7.5; sc.width = 15
    sval = Reference(ws_sens, min_col=2, max_col=2,
                     min_row=sens["scen_first_row"], max_row=sens["scen_last_row"])
    sc.add_data(sval, titles_from_data=False)
    slabels = Reference(ws_sens, min_col=1, max_col=1,
                        min_row=sens["scen_first_row"], max_row=sens["scen_last_row"])
    sc.set_categories(slabels)
    _style_series_fill(sc.series[0], BrandTokens.rgb(tokens.accents[0] if tokens.accents else tokens.primary))
    sc.legend = None
    ws.add_chart(sc, "J18")

    # print area MUST include the charts (anchored down to ~row 33, col N)
    ws.print_area = "A1:N34"
    return ws


def _recolor_workbook(wb, tokens: BrandTokens):
    """Repaint default fin-model banners/sections/tabs to the brand palette so the
    whole workbook reads as one branded document."""
    primary = BrandTokens.rgb(tokens.primary)
    section = BrandTokens.rgb(tokens.secondary)
    tabmap = {"README": tokens.muted, "Inputs": tokens.secondary, "Unit Economics": tokens.secondary,
              "P&L": tokens.primary, "Cash & Break-Even": tokens.accents[0] if tokens.accents else tokens.primary,
              "Sensitivity": tokens.accents[1] if len(tokens.accents) > 1 else tokens.secondary,
              "Dashboard": tokens.primary}
    for ws in wb.worksheets:
        if ws.title in tabmap:
            ws.sheet_properties.tabColor = BrandTokens.rgb(tabmap[ws.title])
        for row in ws.iter_rows():
            for cell in row:
                f = cell.fill
                if f is not None and f.fgColor is not None and f.fgColor.rgb:
                    rgb = str(f.fgColor.rgb)[-6:].upper()
                    if rgb == DEFAULT_BANNER:
                        cell.fill = PatternFill("solid", fgColor=primary)
                    elif rgb == DEFAULT_SECTION:
                        cell.fill = PatternFill("solid", fgColor=section)


def apply_print_setup(wb):
    """Landscape, fit-to-width one page, sane margins, repeat header row."""
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.55, bottom=0.5,
                                      header=0.2, footer=0.2)
        ws.print_options.horizontalCentered = True
        if ws.title not in (DASH,) and ws.print_area is None:
            # extend print area to the used range so nothing is clipped
            ws.print_area = f"A1:{get_column_letter(max(ws.max_column, 2))}{ws.max_row}"


# ═══════════════════════════════════════════════════════════════════════════
#  BOARD PDF  (reportlab — mirrors the spreadsheet palette & grid)
# ═══════════════════════════════════════════════════════════════════════════

def _money(v):
    if v is None:
        return ""
    v = round(v)
    if v == 0:
        return "–"           # en-dash for zero (matches "-" spreadsheet format)
    return f"(${abs(v):,.0f})" if v < 0 else f"${v:,.0f}"


def _pct(v):
    return f"{v*100:.1f}%" if v is not None else ""


_STD_FONTS = {"Helvetica", "Helvetica-Bold", "Helvetica-Oblique", "Times-Roman",
              "Times-Bold", "Courier", "Courier-Bold", "Symbol", "ZapfDingbats"}


def _safe_font(name, bold=False):
    """Font-safety: only emit fonts reportlab actually has registered, else fall
    back — so a brand font we can't embed never breaks the PDF layout."""
    if name in _STD_FONTS:
        return name
    return "Helvetica-Bold" if bold else "Helvetica"


def build_board_pdf(build_result, tokens, pdf_path):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import (BaseDocTemplate, PageTemplate, Frame, Paragraph,
                                    Spacer, Table, TableStyle, NextPageTemplate, PageBreak)
    from reportlab.lib.styles import ParagraphStyle
    import charts_pdf as ch

    a = build_result["assumptions"]
    sm = shadow_model(a)
    summary = build_result["summary"]
    rows = sm["rows"]
    years = len(rows)
    cats = [f"Y{r['year']}" for r in rows]

    PRIMARY = HexColor(tokens.primary); SECOND = HexColor(tokens.secondary)
    INK = HexColor(tokens.ink); PAPER = HexColor(tokens.paper); MUTED = HexColor(tokens.muted)
    LIGHT = HexColor(tokens.secondary);
    hfont = _safe_font(tokens.heading_font, bold=True)
    bfont = _safe_font(tokens.body_font)

    pw, phh = landscape(letter)
    styles = {
        "h1": ParagraphStyle("h1", fontName=hfont, fontSize=26, textColor=PRIMARY, leading=30),
        "h2": ParagraphStyle("h2", fontName=hfont, fontSize=14, textColor=PRIMARY, leading=18,
                              spaceBefore=6, spaceAfter=6),
        "body": ParagraphStyle("body", fontName=bfont, fontSize=9.5, textColor=INK, leading=14),
        "small": ParagraphStyle("small", fontName=bfont, fontSize=8, textColor=MUTED, leading=11),
        "kpi_v": ParagraphStyle("kpi_v", fontName=hfont, fontSize=18, textColor=PRIMARY, leading=20),
        "kpi_l": ParagraphStyle("kpi_l", fontName=bfont, fontSize=8, textColor=MUTED, leading=10),
        "white": ParagraphStyle("white", fontName=hfont, fontSize=11, textColor=PAPER, leading=13),
    }

    def header_footer(canvas, doc):
        canvas.saveState()
        # top brand rule (the recurring motif: a solid primary band)
        canvas.setFillColor(PRIMARY)
        canvas.rect(0, phh - 6, pw, 6, fill=1, stroke=0)
        canvas.setFillColor(MUTED)
        canvas.setFont(bfont, 7.5)
        canvas.drawString(0.5 * inch, 0.32 * inch, f"{a['product_name']} — Financial Model")
        canvas.drawRightString(pw - 0.5 * inch, 0.32 * inch, f"Page {doc.page}")
        canvas.setStrokeColor(HexColor(tokens.muted))
        canvas.setLineWidth(0.3)
        canvas.line(0.5 * inch, 0.45 * inch, pw - 0.5 * inch, 0.45 * inch)
        canvas.restoreState()

    frame = Frame(0.5 * inch, 0.55 * inch, pw - 1.0 * inch, phh - 1.25 * inch, id="main")
    doc = BaseDocTemplate(pdf_path, pagesize=landscape(letter),
                          leftMargin=0.5 * inch, rightMargin=0.5 * inch,
                          topMargin=0.7 * inch, bottomMargin=0.55 * inch)
    doc.addPageTemplates([PageTemplate(id="all", frames=[frame], onPage=header_footer)])

    story = []

    # ── helper: branded table (mirrors spreadsheet grid + colour-coding) ──
    def grid_table(data, col_widths, num_cols, total_rows=()):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        st = [
            ("FONTNAME", (0, 0), (-1, 0), hfont),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("TEXTCOLOR", (0, 0), (-1, 0), PAPER),
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
            ("FONTNAME", (0, 1), (-1, -1), bfont),
            ("FONTSIZE", (0, 1), (-1, -1), 8.5),
            ("TEXTCOLOR", (0, 1), (-1, -1), INK),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("LINEBELOW", (0, 0), (-1, 0), 0.6, PRIMARY),
            ("LINEBELOW", (0, 1), (-1, -2), 0.25, HexColor(tokens.muted)),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [PAPER, ch._tint(tokens.secondary, 0.90)]),
        ]
        for rr in total_rows:
            st.append(("FONTNAME", (0, rr), (-1, rr), hfont))
            st.append(("LINEABOVE", (0, rr), (-1, rr), 0.6, PRIMARY))
            st.append(("TEXTCOLOR", (0, rr), (-1, rr), PRIMARY))
        t.setStyle(TableStyle(st))
        return t

    # ════════ PAGE 1 — COVER ════════
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(a["product_name"], styles["h1"]))
    story.append(Paragraph("Five-Year Financial Model &amp; Board Pack", styles["h2"]))
    story.append(Spacer(1, 0.12 * inch))

    total_w = pw - 1.0 * inch           # usable frame width
    n_cards = 5
    gap = 0.12 * inch
    card_w = (total_w - gap * (n_cards - 1)) / n_cards

    def kpi_card(label, value):
        inner = Table([[Paragraph(value, styles["kpi_v"])], [Paragraph(label, styles["kpi_l"])]],
                      colWidths=[card_w])
        inner.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), ch._tint(tokens.secondary, 0.88)),
            ("LINEABOVE", (0, 0), (-1, 0), 2.2, PRIMARY),
            ("TOPPADDING", (0, 0), (-1, -1), 7), ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LEFTPADDING", (0, 0), (-1, -1), 9), ("RIGHTPADDING", (0, 0), (-1, -1), 9),
        ]))
        return inner

    inflection = summary["ebitda_inflection_year"]
    cards = [
        kpi_card("Gross margin", f"{summary['gross_margin_pct']:.0f}%"),
        kpi_card(f"EBITDA, Year {years}", _money(rows[-1]["ebitda"])),
        kpi_card("EBITDA inflection", f"Year {inflection}" if inflection else "—"),
        kpi_card("Peak funding need", _money(summary["peak_funding_requirement"])),
        kpi_card("Break-even / mo", f"{summary['break_even_units_per_month']:,} u" if summary["break_even_units_per_month"] else "—"),
    ]
    # interleave gap spacer columns so cards fit the frame exactly
    cells, widths = [], []
    for i, cdef in enumerate(cards):
        cells.append(cdef); widths.append(card_w)
        if i < len(cards) - 1:
            cells.append(""); widths.append(gap)
    card_row = Table([cells], colWidths=widths)
    card_row.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0),
                                  ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                                  ("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.append(card_row)
    story.append(Spacer(1, 0.18 * inch))
    sr = summary["scenario_range"]
    story.append(Paragraph(
        f"<b>Scenario range (Year {years} EBITDA):</b> {_money(sr[0])} (Bear) to {_money(sr[1])} (Bull). "
        f"Base case ties out to the live P&amp;L. All figures recalculate from the model's "
        f"<font color='{tokens.secondary}'><b>Inputs</b></font> tab.", styles["body"]))
    story.append(Spacer(1, 0.10 * inch))
    story.append(Paragraph(
        "<b>How to read this pack:</b> every assumption lives on the Inputs tab (blue cells). "
        "Fixed OpEx is held flat (not scaled with volume), growth applies only at the stated rate, "
        "and cash tax is NOL-adjusted — deliberately conservative. Peak funding is the deepest point "
        "of cumulative free cash flow (EBITDA − CapEx − ΔNWC − cash tax).", styles["small"]))
    # colour-coding legend (mirrors the workbook)
    story.append(Spacer(1, 0.10 * inch))
    legend = Table([[
        Paragraph("<font color='#0000FF'>■</font> Input", styles["small"]),
        Paragraph("<font color='#000000'>■</font> Formula", styles["small"]),
        Paragraph("<font color='#008000'>■</font> Cross-sheet link", styles["small"]),
    ]], colWidths=[1.2 * inch, 1.2 * inch, 1.8 * inch])
    story.append(legend)
    # cover at-a-glance trajectory chart (fills the page, mirrors Dashboard combo)
    story.append(Spacer(1, 0.22 * inch))
    story.append(Paragraph("Revenue &amp; EBITDA trajectory", styles["h2"]))
    cover_combo = ch.bar_line_combo(total_w, 2.7 * inch, cats,
                                    [r["revenue"] for r in rows], [r["ebitda"] for r in rows], tokens)
    story.append(cover_combo)

    # ════════ PAGE 2 — P&L ════════
    story.append(PageBreak())
    story.append(Paragraph("Income Statement (P&amp;L)", styles["h2"]))
    hdr = ["($, fiscal year)"] + [f"Year {y+1}" for y in range(years)]
    def line_row(label, key, fn=_money):
        return [label] + [fn(r[key]) for r in rows]
    pnl_data = [hdr,
                line_row("Units sold", "units", lambda v: f"{v:,.0f}"),
                line_row("Revenue", "revenue"),
                line_row("COGS", "cogs", lambda v: _money(-v)),
                line_row("Gross profit", "gross"),
                line_row("Fixed OpEx", "opex", lambda v: _money(-v)),
                line_row("EBITDA", "ebitda"),
                line_row("Depreciation", "dep", lambda v: _money(-v)),
                line_row("EBIT", "ebit"),
                line_row("Cash tax", "cash_tax", lambda v: _money(-v)),
                ]
    ni = ["Net income"] + [_money(r["ebit"] - r["cash_tax"]) for r in rows]
    pnl_data.append(ni)
    cw = [1.7 * inch] + [(pw - 1.0 * inch - 1.7 * inch) / years] * years
    story.append(grid_table(pnl_data, cw, years, total_rows=(4, 6, 8, len(pnl_data) - 1)))
    story.append(Spacer(1, 0.12 * inch))
    story.append(Paragraph("Gross profit, EBITDA, EBIT and Net income rows are highlighted; "
                           "negatives show in parentheses and zeros as dashes, matching the workbook.",
                           styles["small"]))

    # ════════ PAGE 3 — CASH & BREAK-EVEN + charts ════════
    story.append(PageBreak())
    story.append(Paragraph("Cash Flow &amp; Break-Even", styles["h2"]))
    cf_data = [hdr,
               line_row("EBITDA", "ebitda"),
               line_row("Cash tax", "cash_tax", lambda v: _money(-v)),
               line_row("CapEx", "capex_out", lambda v: _money(-v)),
               line_row("Less: ΔNWC", "d_nwc", lambda v: _money(-v)),
               line_row("Free cash flow", "fcf"),
               line_row("Cumulative FCF", "cum_fcf"),
               ]
    story.append(grid_table(cf_data, cw, years, total_rows=(5, 6)))
    story.append(Spacer(1, 0.14 * inch))
    fcf = ch.line_chart(pw - 1.0 * inch, 2.1 * inch, cats, [r["cum_fcf"] for r in rows], tokens,
                        label="Cumulative Free Cash Flow  —  trough marks the peak funding requirement")
    story.append(fcf)
    story.append(Spacer(1, 0.10 * inch))
    be = summary["break_even_units_per_month"]
    kpis = Table([[
        Paragraph(f"<b>Peak funding:</b> {_money(summary['peak_funding_requirement'])}", styles["body"]),
        Paragraph(f"<b>Break-even:</b> {be:,} units/mo" if be else "Break-even: —", styles["body"]),
        Paragraph(f"<b>EBITDA inflection:</b> Year {inflection}" if inflection else "EBITDA inflection: —", styles["body"]),
    ]], colWidths=[(pw - 1.0 * inch) / 3] * 3)
    story.append(kpis)

    # ════════ PAGE 4 — SENSITIVITY ════════
    story.append(PageBreak())
    story.append(Paragraph("Sensitivity &amp; Scenarios", styles["h2"]))
    ty = int(a["sensitivity"].get("target_year", years))
    base_upm = float(a["units_per_month"]); base_cogs = _cogs_per_unit(a)
    row_vals = [round(base_upm * m) for m in (0.8, 0.9, 1.0, 1.1, 1.2)]
    col_vals = [round(base_cogs * m, 2) for m in (0.9, 0.95, 1.0, 1.05, 1.1)]
    grid = [["Units/mo \\ COGS"] + [f"${c:.2f}" for c in col_vals]]
    for rv in row_vals:
        line = [f"{rv:,}"]
        for cv in col_vals:
            units = rv * 12 * ((1 + float(a["annual_growth_rate"])) ** (ty - 1))
            line.append(_money(units * (float(a["retail_price_per_unit"]) - cv) - float(a["monthly_fixed_opex"]) * 12))
        grid.append(line)
    gcw = [1.3 * inch] + [(pw - 1.0 * inch - 1.3 * inch) / len(col_vals)] * len(col_vals)
    gt = Table(grid, colWidths=gcw)
    # heatmap-ish: centre cell emphasised, header band primary
    gstyle = [
        ("FONTNAME", (0, 0), (-1, 0), hfont), ("FONTNAME", (0, 0), (0, -1), hfont),
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY), ("TEXTCOLOR", (0, 0), (-1, 0), PAPER),
        ("BACKGROUND", (0, 0), (0, -1), SECOND), ("TEXTCOLOR", (0, 1), (0, -1), PAPER),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5), ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.25, HexColor(tokens.muted)),
        ("BACKGROUND", (3, 3), (3, 3), ch._tint(tokens.primary, 0.6)),
        ("FONTNAME", (3, 3), (3, 3), hfont),
    ]
    gt.setStyle(TableStyle(gstyle))
    story.append(gt)
    story.append(Paragraph("Centre cell (base volume × base COGS) ties out to the P&amp;L EBITDA for "
                           f"Year {ty}.", styles["small"]))
    story.append(Spacer(1, 0.12 * inch))
    # tornado + scenarios
    # build tornado data from shadow
    def py_eb(vol=1, price=1, cogs=1, opex=1):
        units = base_upm * vol * 12 * ((1 + float(a["annual_growth_rate"])) ** (ty - 1))
        return units * (float(a["retail_price_per_unit"]) * price - base_cogs * cogs) \
            - float(a["monthly_fixed_opex"]) * 12 * opex
    swings = {"Volume": abs(py_eb(vol=1.15) - py_eb(vol=0.85)),
              "Price": abs(py_eb(price=1.15) - py_eb(price=0.85)),
              "COGS / unit": abs(py_eb(cogs=0.85) - py_eb(cogs=1.15)),
              "OpEx": abs(py_eb(opex=0.85) - py_eb(opex=1.15))}
    order = sorted(swings, key=lambda k: swings[k], reverse=True)
    tor = ch.hbar_tornado((pw - 1.0 * inch) / 2 - 6, 1.7 * inch, order, [swings[k] for k in order], tokens)
    scen = {n: scenario_ebitda(a, a["scenarios"][n]["volume"], a["scenarios"][n]["price"],
                               a["scenarios"][n]["cogs"], ty) for n in ("bear", "base", "bull")}
    sv = ch.vbar_scenarios((pw - 1.0 * inch) / 2 - 6, 1.7 * inch,
                           ["Bear", "Base", "Bull"], [scen["bear"], scen["base"], scen["bull"]], tokens)
    charts_row = Table([[tor, sv]], colWidths=[(pw - 1.0 * inch) / 2] * 2)
    charts_row.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                                    ("LEFTPADDING", (0, 0), (-1, -1), 0)]))
    story.append(charts_row)

    doc.build(story)
    return os.path.abspath(pdf_path)


# ═══════════════════════════════════════════════════════════════════════════
#  ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════════════════

def present(build_result: dict, brand_tokens=None, pdf_path: str = None,
            qa: bool = True) -> dict:
    """Make a built fin-model boardroom-ready.
      build_result : the dict returned by fin-model's build_and_verify().
      brand_tokens : dict | path-to-json | None (falls back to navy default).
    Returns paths to the enriched workbook and the board PDF (+ QA report)."""
    tokens = load_tokens(brand_tokens)
    xlsx_path = build_result["xlsx_path"]

    wb = load_workbook(xlsx_path)
    add_native_charts(wb, build_result["chart_targets"], tokens)
    _recolor_workbook(wb, tokens)
    apply_print_setup(wb)
    wb.save(xlsx_path)

    if pdf_path is None:
        pdf_path = os.path.splitext(xlsx_path)[0] + "_board.pdf"
    # charts_pdf is imported inside build_board_pdf via sys.path of this dir
    if os.path.dirname(os.path.abspath(__file__)) not in sys.path:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    build_board_pdf(build_result, tokens, pdf_path)

    out = {"xlsx_path": xlsx_path, "pdf_path": os.path.abspath(pdf_path),
           "charts": ["revenue+EBITDA combo", "cumulative FCF", "tornado", "scenarios"]}
    if qa:
        try:
            from .qa import qa_pdf
        except ImportError:
            from qa import qa_pdf
        out["qa"] = qa_pdf(pdf_path)
    return out
