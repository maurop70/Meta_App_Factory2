import os
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import logging
import sqlite3
import datetime
import time
import csv
import json
import io
import aiofiles
from typing import Optional, Dict, Any
import jwt
import uuid
from datetime import datetime, timedelta, timezone
import concurrent.futures
from itertools import islice
from fastapi import FastAPI, APIRouter, HTTPException, Query, UploadFile, File, BackgroundTasks, Header, Body, Depends, Security, Response, Path
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse


from dotenv import load_dotenv
import base64
from fastapi import Request

# Initialize Environment
_here = os.path.dirname(os.path.abspath(__file__))
_erp  = os.path.dirname(_here)
_factory = os.path.dirname(_erp)

# I/O Leak Patch: Only hit the disk for .env if variables aren't inherited via spawn
if not os.environ.get("JWT_PRIVATE_KEY_B64"):
    env_path = os.path.join(_factory, '.env')
    load_dotenv(env_path)

# Initialize Standard Output Logging
import sys
logger = logging.getLogger("MaintenanceBackend")
logger.setLevel(logging.INFO)

formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

if logger.hasHandlers():
    logger.handlers.clear()

# Strictly delegate logs to stdout for Docker daemon rotation
stream_handler = logging.StreamHandler(sys.stdout)
stream_handler.setFormatter(formatter)
logger.addHandler(stream_handler)

if not os.environ.get("JWT_PRIVATE_KEY_B64"):
    logger.info(f"Loading .env from: {env_path}")


# --- CRYPTOGRAPHIC CONFIGURATION ---
import httpx
import asyncio

from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field, field_validator
from contextlib import asynccontextmanager

from local_db import get_db_connection
import os
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from fpdf import FPDF

ALGORITHM = "RS256"
PUBLIC_KEY = None
security = HTTPBearer()

# Ingestion Logic for Orchestration Boundary
GLOBAL_AI_DIRECTIVE_CONTEXT = ""

# STRUCTURAL PATCH: Asynchronous boot context to prevent thread-locking
@asynccontextmanager
async def lifespan(app: FastAPI):
    global PUBLIC_KEY
    global GLOBAL_AI_DIRECTIVE_CONTEXT
    
    # Load Orchestration Boundary
    directive_path = os.path.join(_here, "GLOBAL_AI_DIRECTIVE.md")
    try:
        if os.path.exists(directive_path):
            with open(directive_path, "r", encoding="utf-8") as f:
                GLOBAL_AI_DIRECTIVE_CONTEXT = f.read()
            logger.info("Orchestration Boundary Engaged: GLOBAL_AI_DIRECTIVE.md fully loaded into active system memory.")
        else:
            logger.warning("Orchestration Boundary Warning: GLOBAL_AI_DIRECTIVE.md not found at /app/GLOBAL_AI_DIRECTIVE.md.")
    except Exception as e:
        logger.error(f"Failed to ingest GLOBAL_AI_DIRECTIVE.md: {e}")

    # Guard against schema drift (e.g. Phase 46 DDL applied but code stale)
    _assert_work_orders_schema()
    logger.info("[SCHEMA] work_orders projection verified against live database.")

    # NGINX DOCTRINE: Trailing slash strictly enforced
    url = "http://127.0.0.1:9000/api/v1/auth/public-key" 
    
    async with httpx.AsyncClient() as client:
        max_retries = 5
        for attempt in range(max_retries):
            try:
                response = await client.get(url, timeout=5.0, follow_redirects=True)
                response.raise_for_status()
                PUBLIC_KEY = response.json()["public_key"]
                logger.info("[SECURITY] Successfully fetched RS256 Public Key from Module 0 Gateway.")
                break
            except Exception as e:
                logger.warning(f"[SECURITY] Gateway fetch failed (attempt {attempt+1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    await asyncio.sleep(2)
                else:
                    raise RuntimeError("FATAL: Could not retrieve PUBLIC_KEY from Module 0 Gateway.")
    yield
    # Application teardown logic executes here

app = FastAPI(title="Maintenance Work Order API - Global ERP Connected", lifespan=lifespan)
api_router = APIRouter(prefix="/api")

def is_jti_revoked(jti: str) -> bool:
    conn = get_db_connection()
    try:
        row = conn.execute("SELECT 1 FROM revoked_tokens WHERE jti = ?", (jti,)).fetchone()
        return row is not None
    except Exception as e:
        logger.error(f"Failed to check JTI status: {e}")
        return False
    finally:
        conn.close()

def verify_jwt_token(credentials: HTTPAuthorizationCredentials = Security(security)) -> dict:
    if not PUBLIC_KEY:
        raise HTTPException(status_code=503, detail="Cryptographic Gateway Unavailable.")
        
    token = credentials.credentials
    try:
        payload = jwt.decode(token, PUBLIC_KEY, algorithms=[ALGORITHM])
        
        # JTI check routed to persistent storage
        if is_jti_revoked(payload.get("jti")):
            raise HTTPException(status_code=401, detail="Token Revoked (JTI Blacklisted).")
            
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token Expired.")
    except jwt.InvalidTokenError as e:
        logger.error(f"JWT Verification Error: {str(e)}")
        raise HTTPException(status_code=403, detail="Cryptographic Verification Failed.")

# Serve frontend is deprecated due to NGINX Reverse-Proxy Decoupling Doctrine

# CORS: Decoupled frontend access
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173", "http://127.0.0.1:5173",
        "http://localhost:5174", "http://127.0.0.1:5174",
        "http://localhost:5175", "http://127.0.0.1:5175",
        "http://localhost:5176", "http://127.0.0.1:5176"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- GLOBAL ERP DATA MODELS ---

class WorkOrderSubmit(BaseModel):
    reported_by: str
    asset_id: str
    issue_description: str

class StatusUpdate(BaseModel):
    order_id: str
    new_status: str # PENDING, IN_PROGRESS, COMPLETE
    user_id: str 
    resolution_notes: str = ""

class UserSubmit(BaseModel):
    name: str
    role_id: str           # FK to erp_roles.id — carries full permissions
    authorization_level: str # TECH, MANAGER, ADMIN
    phone_number: str
    pin_code: str
    department_id: str

class EquipmentSubmit(BaseModel):
    machine_name: str
    location: str
    department_id: str
    operational_status: str = "OPERATIONAL"

class EquipmentIngestionRecord(BaseModel):
    nomenclature: str = Field(..., min_length=2, description="Human-readable equipment name")
    category_id: str = Field(..., description="FK reference to erp_categories.id")
    status: str = Field(default="ACTIVE", description="Strictly enforced operational state")
    department_id: str = Field(..., description="FK reference to erp_departments.id")
    assigned_tech_id: Optional[str] = None
    location_id: str = Field(..., description="Physical location tag")
    assigned_hm_id: str = Field(..., description="Hub Manager assigned to this equipment")
    
    @field_validator('status', mode='before')
    @classmethod
    def validate_status(cls, v):
        allowed_states = {"ACTIVE", "DEGRADED", "OFFLINE"}
        if str(v).upper() not in allowed_states:
            raise ValueError(f"Structural Violation: Status must be one of {allowed_states}")
        return str(v).upper()

    @field_validator('assigned_tech_id', mode='before')
    @classmethod
    def empty_string_to_none(cls, v):
        if isinstance(v, str) and not v.strip():
            return None
        return v

class EquipmentActuationPayload(BaseModel):
    status: str = Field(..., description="ACTIVE, DEGRADED, OFFLINE, RETIRED")
    assigned_tech_id: Optional[str] = None

    @field_validator('status', mode='before')
    @classmethod
    def validate_status(cls, v):
        allowed_states = {"ACTIVE", "DEGRADED", "OFFLINE", "RETIRED"}
        if str(v).upper() not in allowed_states:
            raise ValueError(f"Structural Violation: Status must be one of {allowed_states}")
        return str(v).upper()

    @field_validator('assigned_tech_id', mode='before')
    @classmethod
    def empty_string_to_none(cls, v):
        if isinstance(v, str) and not v.strip():
            return None
        return v

class EmployeeIngestionRecord(BaseModel):
    name: str = Field(..., min_length=2, max_length=100)
    role: str = Field(..., description="Mapped Role (e.g., ADMIN, TECH)")
    pin_code: str = Field(..., min_length=4, max_length=12, description="Raw PIN to be hashed")
    is_active: int = Field(default=1, ge=0, le=1)
    
    # [PHASE 35.3 RELATIONAL INJECTIONS]
    department_id: str = Field(..., min_length=2, description="Physical operational department FK")
    reports_to_hm_id: Optional[str] = None
    
    @field_validator('role', mode='before')
    @classmethod
    def validate_role(cls, v):
        allowed = {"ADMINISTRATOR", "ADMIN", "HM", "TECH"}
        if str(v).upper() not in allowed:
            raise ValueError(f"Structural Violation: Role must be one of {allowed}")
        return str(v).upper()
    
    @field_validator('reports_to_hm_id', mode='after')
    @classmethod
    def require_hm_for_tech(cls, v, info):
        role = info.data.get('role', '').upper()
        if role in ["TECH", "TECHNICIAN"] and not v:
            raise ValueError("Structural Violation: A TECHNICIAN must be assigned a reporting HM.")
        return v

# [PHASE 34.9] Parts Catalog Ingestion Schema
class PartIngestionRecord(BaseModel):
    nomenclature: str = Field(..., min_length=2, description="Human-readable part name")
    category_id: str = Field(..., description="Foreign Key to erp_categories")
    quantity_on_hand: int = Field(default=0, ge=0, description="Initial physical stock count")
    reorder_threshold: int = Field(default=5, ge=0, description="Minimum threshold before alert")
    unit_cost: float = Field(default=0.0, ge=0.0, description="Financial cost per unit")

from typing import List

class MwoConsumptionPayload(BaseModel):
    part_ids: List[str] = Field(..., min_items=1, description="Array of exact physical part identifiers")

class PartConsumptionPayload(BaseModel):
    part_id: str = Field(..., description="The physical part identifier from erp_parts")
    quantity_consumed: int = Field(..., gt=0, description="Strictly positive integer")

class SupplierCreate(BaseModel):
    supplier_id: str = Field(..., min_length=2, max_length=50)
    name: str = Field(..., min_length=2, max_length=200)
    email: str = Field(..., max_length=254)
    phone: Optional[str] = Field(None, max_length=50)
    address: Optional[str] = Field(None, max_length=500)
    default_lead_time_days: int = Field(default=7, ge=0, le=365)

    @field_validator('email', mode='before')
    @classmethod
    def validate_email(cls, v):
        import re
        v = str(v).strip()
        if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', v):
            raise ValueError("Structural Violation: email must be a valid address (user@domain.tld)")
        return v

    @field_validator('supplier_id', 'name', mode='before')
    @classmethod
    def strip_compulsory(cls, v):
        v = str(v).strip()
        if not v:
            raise ValueError("Structural Violation: compulsory field cannot be blank")
        return v

    @field_validator('phone', 'address', mode='before')
    @classmethod
    def blank_to_none(cls, v):
        if isinstance(v, str) and not v.strip():
            return None
        return v

class SKUCreate(BaseModel):
    sku_id: str
    nomenclature: str
    unit_cost: float
    reorder_threshold: int
    supplier_id: Optional[str] = None
    # Minimum Order Quantity: auto-drafted reorders are rounded UP to this
    min_order_qty: Optional[int] = Field(default=1, ge=1)
    # Atomic inline registration: supplier + SKU land in ONE transaction,
    # eliminating the orphan-supplier window of the old two-call chain.
    new_supplier: Optional[SupplierCreate] = None

class SKUUpdate(BaseModel):
    nomenclature: Optional[str] = None
    unit_cost: Optional[float] = Field(None, ge=0)
    reorder_threshold: Optional[int] = Field(None, ge=0)
    min_order_qty: Optional[int] = Field(None, ge=1)
    supplier_id: Optional[str] = None

class PartCreate(BaseModel):
    sku_id: str
    serial_number: Optional[str] = None

class PartResponse(BaseModel):
    part_id: str
    sku_id: str
    serial_number: Optional[str]
    status: str

class InventoryMutation(BaseModel):
    sku_id: str
    mutation_type: str = Field(..., description="IN or OUT")
    quantity: int

    @field_validator('mutation_type')
    @classmethod
    def validate_mutation(cls, v):
        if str(v).upper() not in ["IN", "OUT"]:
            raise ValueError("mutation_type must be IN or OUT")
        return str(v).upper()

class ProcurementActuation(BaseModel):
    status: str = Field(..., description="Target state: APPROVED, REJECTED, or FULFILLED")
    authorized_quantity: Optional[int] = Field(default=None, ge=1, description="Required only when status is APPROVED")

class PinVerify(BaseModel):
    employee_id: str = Field(..., description="Unique ERP Employee Identifier")
    pin: str = Field(..., min_length=4, max_length=12, description="First-party ERP Employee PIN")

class MWOIngestionRecord(BaseModel):
    mwo_id: str = Field(..., description="Unique Work Order Identifier")
    equipment_id: str = Field(..., description="Target machine or asset identifier")
    description: str = Field(..., min_length=5, description="Contextual issue description")
    
    status: str = Field(default="UNASSIGNED", description="Strictly enforced operational state")
    dm_urgency: str = Field(default="Normal")
    hm_priority: str = Field(default="Normal")
    
    assigned_tech: Optional[str] = None
    consumed_sku: Optional[str] = None
    manual_log: Optional[str] = None
    archival_pdf_path: Optional[str] = None
    # STRICT TYPING DEPLOYED. Pydantic handles ISO 8601 validation natively.
    start_date: Optional[datetime] = None 
    
    @field_validator('status', mode='before')
    @classmethod
    def validate_status(cls, v):
        allowed_states = {"UNASSIGNED", "ASSIGNED", "IN_PROGRESS", "PAUSED", "PENDING_REVIEW", "COMPLETED"}
        if str(v).upper() not in allowed_states:
            raise ValueError(f"Structural Violation: Status must be one of {allowed_states}")
        return str(v).upper()

    @field_validator('dm_urgency', 'hm_priority', mode='before')
    @classmethod
    def validate_urgency(cls, v):
        if v is None or (isinstance(v, str) and not v.strip()): 
            return "Normal"
        allowed = {"Low", "Normal", "High", "Critical"}
        # Case-insensitive match, strict formatting
        formatted = str(v).capitalize()
        if formatted not in allowed:
            raise ValueError(f"Structural Violation: Urgency/Priority must be one of {allowed}")
        return formatted

    @field_validator('assigned_tech', 'consumed_sku', 'manual_log', 'start_date', mode='before')
    @classmethod
    def empty_string_to_none(cls, v):
        if isinstance(v, str) and not v.strip():
            return None
        return v

class MWOUpdate(BaseModel):
    status: Optional[str] = None
    manual_log: Optional[str] = None
    assigned_tech: Optional[str] = None
    hm_priority: Optional[str] = None

    @field_validator('manual_log', mode='before')
    @classmethod
    def empty_string_to_none(cls, v):
        if isinstance(v, str) and not v.strip():
            return None
        return v

class MWOExecutePayload(BaseModel):
    action: str = Field(..., description="START, PAUSE, COMPLETE")
    manual_log: Optional[str] = None

    @field_validator('action', mode='before')
    @classmethod
    def validate_action(cls, v):
        allowed = {"START", "PAUSE", "COMPLETE"}
        if str(v).upper() not in allowed:
            raise ValueError(f"Action must be one of {allowed}")
        return str(v).upper()

# --- SECURITY LAYER ---

def sanitize_user(user: dict) -> dict:
    """Strip pin_code from any user payload. Inject computed has_pin boolean."""
    has_pin = bool(user.get("pin_code"))
    sanitized = {k: v for k, v in user.items() if k != "pin_code"}
    sanitized["has_pin"] = has_pin
    return sanitized

# Canonical work_orders projection — single source of truth for all
# SELECTs. consumed_sku was dropped in Phase 46 (apply_phase46_ddl.py);
# part consumption now lives in mwo_consumed_parts. Verified against the
# live schema at startup by _assert_work_orders_schema().
WORK_ORDER_COLS = (
    "mwo_id, status, dm_urgency, hm_priority, description, assigned_tech, "
    "manual_log, created_at, triaged_at, execution_start, execution_end, "
    "completed_at, start_date, equipment_id, location_id, material_cost, "
    "archival_pdf_path"
)

def _assert_work_orders_schema():
    """Fail loudly at boot if code and DB schema have drifted."""
    conn = get_db_connection()
    try:
        live = {row[1] for row in conn.execute("PRAGMA table_info(work_orders)")}
        expected = {c.strip() for c in WORK_ORDER_COLS.split(",")}
        missing = expected - live
        if missing:
            raise RuntimeError(
                f"work_orders schema drift: code references missing column(s) "
                f"{sorted(missing)}. Re-run pending DDL or fix WORK_ORDER_COLS."
            )
    finally:
        conn.close()

def get_current_mwo(mwo_id: str):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(f"SELECT {WORK_ORDER_COLS} FROM work_orders WHERE mwo_id = ?", (mwo_id,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="MWO not found.")
        return dict(row)
    finally:
        conn.close()

def verify_rbac_pipeline(
    payload: MWOUpdate = Body(...), 
    current_mwo: dict = Depends(get_current_mwo),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    updates = payload.model_dump(exclude_unset=True)
    if not updates:
        return current_mwo
        
    role = jwt_payload.get("role")
    current_status = current_mwo.get("status")
    new_status = updates.get("status", current_status)
    
    if role == "ADMINISTRATOR":
        return current_mwo
        
    if role == "DM":
        if not set(updates.keys()).issubset({"dm_urgency"}):
            raise HTTPException(status_code=403, detail="RBAC Violation: DM unauthorized mutation.")
        return current_mwo
        
    if role == "TECHNICIAN":
        if not set(updates.keys()).issubset({"status", "manual_log"}):
            raise HTTPException(status_code=403, detail="RBAC Violation: Technician unauthorized mutation.")
            
        if new_status != current_status:
            if not ((current_status == "ASSIGNED" and new_status == "IN_PROGRESS") or
                    (current_status == "IN_PROGRESS" and new_status == "PENDING_REVIEW")):
                raise HTTPException(status_code=403, detail="RBAC Violation: Invalid pipeline transition.")
        return current_mwo
        
    if role == "HM":
        if not set(updates.keys()).issubset({"assigned_tech", "hm_priority", "status", "manual_log"}):
            raise HTTPException(status_code=403, detail="RBAC Violation: HM unauthorized mutation.")
            
        if new_status != current_status:
            if not ((current_status == "UNASSIGNED" and new_status == "ASSIGNED") or
                    (current_status == "PENDING_REVIEW" and new_status == "COMPLETED") or
                    (new_status == "UNASSIGNED")):
                raise HTTPException(status_code=403, detail="RBAC Violation: Invalid pipeline transition.")
        return current_mwo
        
    raise HTTPException(status_code=403, detail="RBAC / Pipeline Violation: Unrecognized role.")


# --- ENDPOINTS ---




# Authentication and Identity management (including login, refresh, and rate limiting)
# have been fully delegated to the Module 0 Gateway on port 9000.

# [PHASE 34.5 INJECTION] Provide HM routing matrix to the UI
@api_router.get("/admin/hms")
def get_hms(
    department_id: str = Query(..., description="Target department ID to filter HMs"),
    limit: int = Query(50, ge=1, le=100),
    offset: int = Query(0, ge=0),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        # Strictly bounded parameterized query scoped to the exact department
        cursor.execute(
            """
            SELECT id, name 
            FROM erp_employees 
            WHERE role = 'HM' AND is_active = 1 AND department_id = ? 
            LIMIT ? OFFSET ?
            """,
            (department_id, limit, offset)
        )
        rows = [dict(row) for row in cursor.fetchall()]
        return {"status": "success", "data": rows}
    except Exception as e:
        logger.error(f"Failed to fetch bounded HM matrix: {e}")
        raise HTTPException(status_code=500, detail="Matrix synchronization failed.")
    finally:
        conn.close()

# [PHASE 34.5 INJECTION]
@api_router.post("/admin/ingest/single-user")
async def ingest_single_user(payload: EmployeeIngestionRecord, jwt_payload: dict = Depends(verify_jwt_token)):
    # 1. RBAC Adherence
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Administrative clearance required.")
        
    # Mandatory Silent Patch: Explicit Route-Level Validation
    if payload.role == 'TECH' and not payload.reports_to_hm_id:
        raise HTTPException(status_code=400, detail="Structural Violation: TECH must be bound to a reporting HM.")
        
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 2. CPU-Bound Password Hashing
        import bcrypt
        import asyncio
        pin_hash = await asyncio.to_thread(
            lambda p: bcrypt.hashpw(p.encode('utf-8'), bcrypt.gensalt()).decode('utf-8'), 
            payload.pin_code
        )
        
        # 3. PK Synthesization
        import uuid
        new_id = f"U-{uuid.uuid4().hex[:6].upper()}"
        
        # 4. Atomic Database Insertion (Cryptographic isolation - NO plaintext pin_code)
        cursor.execute("PRAGMA foreign_keys = ON")
        cursor.execute(
            """
            INSERT INTO erp_employees (id, name, role, pin_hash, is_active, department_id, reports_to_hm_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (new_id, payload.name, payload.role, pin_hash, payload.is_active, payload.department_id, payload.reports_to_hm_id)
        )
        conn.commit()
        
        return {"status": "success", "message": f"Personnel {new_id} successfully ingested."}
    except sqlite3.IntegrityError:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=409, detail="Structural Violation: Employee ID already exists.")
    except Exception as e:
        logger.error(f"Single Ingestion Error: {e}")
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail="Internal Execution Error.")
    finally:
        if conn:
            conn.close()

# [PHASE 35.1 INJECTION — FK-NORMALIZED]
@api_router.post("/admin/ingest/equipment")
def ingest_equipment(payload: EquipmentIngestionRecord, jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Administrative clearance required.")
        
    # Autonomously generate PK — operators never dictate physical keys
    equipment_id = f"EQ-{uuid.uuid4().hex[:6].upper()}"

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("PRAGMA foreign_keys = ON")
        cursor.execute(
            """
            INSERT INTO erp_equipment (equipment_id, nomenclature, category_id, status, department_id, assigned_tech_id, location_id, assigned_hm_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (equipment_id, payload.nomenclature, payload.category_id, payload.status, payload.department_id, payload.assigned_tech_id, payload.location_id, payload.assigned_hm_id)
        )
        conn.commit()
        return {"status": "success", "message": f"Equipment {equipment_id} successfully ingested.", "equipment_id": equipment_id}
    except sqlite3.IntegrityError as e:
        raise HTTPException(status_code=409, detail=f"Structural Violation: FK constraint or duplicate key. {e}")
    except Exception as e:
        logger.error(f"Equipment Ingestion Error: {e}")
        raise HTTPException(status_code=500, detail="Internal Execution Error.")
    finally:
        conn.close()


def generate_xlsx_template(headers, template_name, categories=[], departments=[], locations=[], hms=[], techs=[]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ingestion_Template"
    
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
        
    lookup_ws = wb.create_sheet(title="_Lookups")
    lookup_ws.sheet_state = 'hidden'
    
    current_col = 1
    def add_lookup_col(name, items):
        nonlocal current_col
        lookup_ws.cell(row=1, column=current_col, value=name)
        for r_idx, item in enumerate(items, 2):
            lookup_ws.cell(row=r_idx, column=current_col, value=item)
        col_letter = openpyxl.utils.get_column_letter(current_col)
        formula = f"=_Lookups!${col_letter}$2:${col_letter}${len(items)+1 if len(items)>0 else 2}"
        current_col += 1
        return formula

    if "category_name" in headers and categories:
        form = add_lookup_col("Categories", categories)
        dv = DataValidation(type="list", formula1=form, allow_blank=False)
        col_letter = openpyxl.utils.get_column_letter(headers.index("category_name") + 1)
        dv.add(f"{col_letter}2:{col_letter}1048576")
        ws.add_data_validation(dv)
        
    if "department_name" in headers and departments:
        form = add_lookup_col("Departments", departments)
        dv = DataValidation(type="list", formula1=form, allow_blank=False)
        col_letter = openpyxl.utils.get_column_letter(headers.index("department_name") + 1)
        dv.add(f"{col_letter}2:{col_letter}1048576")
        ws.add_data_validation(dv)
        
    if "location" in headers and locations:
        form = add_lookup_col("Locations", locations)
        dv = DataValidation(type="list", formula1=form, allow_blank=False)
        col_letter = openpyxl.utils.get_column_letter(headers.index("location") + 1)
        dv.add(f"{col_letter}2:{col_letter}1048576")
        ws.add_data_validation(dv)
        
    if "hm_name" in headers and hms:
        form = add_lookup_col("HMs", hms)
        dv = DataValidation(type="list", formula1=form, allow_blank=False)
        col_letter = openpyxl.utils.get_column_letter(headers.index("hm_name") + 1)
        dv.add(f"{col_letter}2:{col_letter}1048576")
        ws.add_data_validation(dv)
        
    if "reports_to_hm_name" in headers and hms:
        form = add_lookup_col("HMs", hms)
        dv = DataValidation(type="list", formula1=form, allow_blank=True)
        col_letter = openpyxl.utils.get_column_letter(headers.index("reports_to_hm_name") + 1)
        dv.add(f"{col_letter}2:{col_letter}1048576")
        ws.add_data_validation(dv)
        
    if "tech_name" in headers and techs:
        form = add_lookup_col("Techs", techs)
        dv = DataValidation(type="list", formula1=form, allow_blank=True)
        col_letter = openpyxl.utils.get_column_letter(headers.index("tech_name") + 1)
        dv.add(f"{col_letter}2:{col_letter}1048576")
        ws.add_data_validation(dv)
        
    if "status" in headers:
        dv = DataValidation(type="list", formula1='"ACTIVE,DEGRADED,OFFLINE"', allow_blank=False)
        col_letter = openpyxl.utils.get_column_letter(headers.index("status") + 1)
        dv.add(f"{col_letter}2:{col_letter}1048576")
        ws.add_data_validation(dv)

    if "role" in headers:
        dv = DataValidation(type="list", formula1='"ADMINISTRATOR,ADMIN,DM,HM,TECH"', allow_blank=False)
        col_letter = openpyxl.utils.get_column_letter(headers.index("role") + 1)
        dv.add(f"{col_letter}2:{col_letter}1048576")
        ws.add_data_validation(dv)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={template_name}.xlsx"}
    )

# [PHASE 35.1.1] Bulk CSV Ingestion Endpoints
@api_router.get("/admin/ingest/equipment/template")
def get_equipment_ingestion_template(jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN", "HM"]:
        raise HTTPException(status_code=403, detail="RBAC Violation.")
    
    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("SELECT name FROM erp_categories")
        categories = [row['name'] for row in c.fetchall()]
        c.execute("SELECT name FROM erp_departments")
        departments = [row['name'] for row in c.fetchall()]
        c.execute("SELECT name FROM erp_locations")
        locations = [row['name'] for row in c.fetchall()]
        c.execute("SELECT name FROM erp_employees WHERE role='HM'")
        hms = [row['name'] for row in c.fetchall()]
    finally:
        conn.close()
        
    headers = ["nomenclature", "category_name", "status", "department_name", "location", "hm_name"]
    return generate_xlsx_template(headers, "equipment_ingestion_template", categories, departments, locations, hms)

@api_router.post("/admin/ingest/equipment/bulk")
async def bulk_ingest_equipment(file: UploadFile = File(...), jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Administrative clearance required.")
        
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an XLSX file.")
        
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if not header_row: raise ValueError("Empty file")
        expected_fields = {"nomenclature", "category_name", "status", "department_name", "location", "hm_name"}
        if not expected_fields.issubset(set(header_row)):
            raise HTTPException(status_code=400, detail=f"Invalid XLSX structure. Expected minimum headers: {', '.join(expected_fields)}")
        rows = []
        for row_values in ws.iter_rows(min_row=2, values_only=True):
            if not any(row_values): continue
            row_dict = dict(zip(header_row, row_values))
            rows.append(row_dict)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read XLSX: {e}")
    if not rows:
        raise HTTPException(status_code=400, detail="CSV file is empty.")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("PRAGMA foreign_keys = ON")
        
        # Pre-fetch lookup mappings to avoid N+1 queries
        cursor.execute("SELECT id, name FROM erp_categories")
        cat_map = {row['name'].lower().strip(): row['id'] for row in cursor.fetchall()}
        
        cursor.execute("SELECT id, name FROM erp_departments")
        dep_map = {row['name'].lower().strip(): row['id'] for row in cursor.fetchall()}
        
        cursor.execute("SELECT id, name FROM erp_locations")
        loc_map = {row['name'].lower().strip(): row['id'] for row in cursor.fetchall()}
        
        cursor.execute("SELECT id, name FROM erp_employees WHERE role = 'HM'")
        hm_map = {row['name'].lower().strip(): row['id'] for row in cursor.fetchall()}

        cursor.execute("BEGIN TRANSACTION")
        
        inserted_count = 0
        for i, row in enumerate(rows):
            # Clean up keys and values
            row_data = {k.strip(): str(v).strip() if v is not None else None for k, v in row.items()}
            
            # Resolve Names to IDs
            cat_name = row_data.get('category_name')
            if not cat_name:
                raise ValueError(f"Row {i+2}: Missing category_name")
            if cat_name.lower() not in cat_map:
                new_cat = f"CAT-{uuid.uuid4().hex[:6].upper()}"
                cursor.execute("INSERT INTO erp_categories (id, name) VALUES (?, ?)", (new_cat, cat_name.strip()))
                cat_map[cat_name.lower()] = new_cat
            cat_id = cat_map[cat_name.lower()]
                
            dep_name = row_data.get('department_name')
            if not dep_name:
                raise ValueError(f"Row {i+2}: Missing department_name")
            if dep_name.lower() not in dep_map:
                new_dep = f"DEP-{uuid.uuid4().hex[:6].upper()}"
                cursor.execute("INSERT INTO erp_departments (id, name) VALUES (?, ?)", (new_dep, dep_name.strip()))
                dep_map[dep_name.lower()] = new_dep
            dep_id = dep_map[dep_name.lower()]
            
            hm_name = row_data.get('hm_name')
            if not hm_name or hm_name.lower() not in hm_map:
                raise ValueError(f"Row {i+2}: Unknown HM Name '{hm_name}'")
            hm_id = hm_map[hm_name.lower()]
            
            loc_name = row_data.get('location')
            if not loc_name:
                raise ValueError(f"Row {i+2}: Missing location")
            if loc_name.lower() not in loc_map:
                new_loc = f"LOC-{uuid.uuid4().hex[:6].upper()}"
                cursor.execute("INSERT INTO erp_locations (id, name) VALUES (?, ?)", (new_loc, loc_name.strip()))
                loc_map[loc_name.lower()] = new_loc
            loc_id = loc_map[loc_name.lower()]
            
            # Autonomously generate PK
            equipment_id = f"EQ-{uuid.uuid4().hex[:6].upper()}"
            
            # Status validation
            status = row_data.get('status', 'ACTIVE') or 'ACTIVE'
            status = status.upper()
            if status not in {"ACTIVE", "DEGRADED", "OFFLINE"}:
                raise ValueError(f"Row {i+2}: Status must be ACTIVE, DEGRADED, or OFFLINE.")
                
            cursor.execute(
                """
                INSERT INTO erp_equipment (equipment_id, nomenclature, category_id, status, department_id, location_id, assigned_hm_id)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    equipment_id, 
                    row_data.get('nomenclature'), 
                    cat_id, 
                    status, 
                    dep_id, 
                    loc_id, 
                    hm_id
                )
            )
            inserted_count += 1
            
        conn.commit()
        return {"status": "success", "message": f"Successfully ingested {inserted_count} equipment records."}
        
    except sqlite3.IntegrityError as e:
        conn.rollback()
        raise HTTPException(status_code=409, detail=f"Structural Violation on bulk insert. {e}")
    except ValueError as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        conn.rollback()
        logger.error(f"Bulk Equipment Ingestion Error: {e}")
        raise HTTPException(status_code=500, detail="Internal Execution Error during bulk ingestion.")
    finally:
        conn.close()

@api_router.get("/admin/equipment")
def get_equipment(
    limit: int = Query(50, ge=1, le=100),
    offset: int = Query(0, ge=0),
    department_id: Optional[str] = Query(None, description="Optional filter by department FK"),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        base_query = """
            SELECT e.equipment_id, e.nomenclature, e.category_id, c.name AS category,
                   e.status, e.department_id, d.name AS department, e.assigned_tech_id,
                   e.location_id, l.name AS location_name
            FROM erp_equipment e
            LEFT JOIN erp_categories c ON e.category_id = c.id
            LEFT JOIN erp_departments d ON e.department_id = d.id
            LEFT JOIN erp_locations l ON e.location_id = l.id
        """
        if department_id:
            cursor.execute(
                f"{base_query} WHERE e.department_id = ? LIMIT ? OFFSET ?",
                (department_id, limit, offset)
            )
        else:
            cursor.execute(
                f"{base_query} LIMIT ? OFFSET ?",
                (limit, offset)
            )
        rows = [dict(row) for row in cursor.fetchall()]
        return {"status": "success", "data": rows}
    except Exception as e:
        logger.error(f"Failed to fetch equipment matrix: {e}")
        raise HTTPException(status_code=500, detail="Matrix synchronization failed.")
    finally:
        conn.close()

# [PHASE 35.1] Paginated Lookup Routes
@api_router.get("/admin/lookups/categories")
def get_categories(
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN", "HM"]:
        raise HTTPException(status_code=403, detail="RBAC Violation.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name, manager_id FROM erp_categories ORDER BY name LIMIT ? OFFSET ?", (limit, offset))
        return {"data": [dict(r) for r in cursor.fetchall()]}
    finally:
        conn.close()

@api_router.get("/admin/lookups/departments")
def get_departments(
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN", "HM"]:
        raise HTTPException(status_code=403, detail="RBAC Violation.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name FROM erp_departments ORDER BY name LIMIT ? OFFSET ?", (limit, offset))
        return {"data": [dict(r) for r in cursor.fetchall()]}
    finally:
        conn.close()

@api_router.get("/admin/lookups/locations")
def get_locations(
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name FROM erp_locations ORDER BY name LIMIT ? OFFSET ?", (limit, offset))
        return {"data": [dict(r) for r in cursor.fetchall()]}
    finally:
        conn.close()

@api_router.post("/admin/ingest/department")
def ingest_department(payload: dict = Body(...), jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation")
    name = payload.get("name")
    if not name or len(name) < 2:
        raise HTTPException(status_code=400, detail="Invalid department name")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        import uuid
        new_id = f"DEP-{uuid.uuid4().hex[:6].upper()}"
        cursor.execute("INSERT INTO erp_departments (id, name) VALUES (?, ?)", (new_id, name))
        conn.commit()
        return {"status": "success", "id": new_id, "name": name}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="Department already exists")
    finally:
        conn.close()

@api_router.post("/admin/ingest/location")
def ingest_location(payload: dict = Body(...), jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation")
    name = payload.get("name")
    if not name or len(name) < 2:
        raise HTTPException(status_code=400, detail="Invalid location name")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        import uuid
        new_id = f"LOC-{uuid.uuid4().hex[:6].upper()}"
        cursor.execute("INSERT INTO erp_locations (id, name) VALUES (?, ?)", (new_id, name))
        conn.commit()
        return {"status": "success", "id": new_id, "name": name}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="Location already exists")
    finally:
        conn.close()

@api_router.get("/admin/lookups/technicians")
def get_technicians(
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN", "HM"]:
        raise HTTPException(status_code=403, detail="RBAC Violation.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, name FROM erp_employees WHERE authorization_level = 'TECHNICIAN' AND is_active = 1 ORDER BY name LIMIT ? OFFSET ?",
            (limit, offset)
        )
        return {"data": [dict(r) for r in cursor.fetchall()]}
    finally:
        conn.close()

# [PHASE 35.2] Equipment Actuation Gateway
@api_router.put("/admin/equipment/{equipment_id}/actuate")
def actuate_equipment(
    equipment_id: str,
    payload: EquipmentActuationPayload,
    jwt_payload: dict = Depends(verify_jwt_token)
):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN", "HM"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: ADMIN or HM clearance required to actuate equipment.")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("PRAGMA foreign_keys = ON")
        cursor.execute("BEGIN TRANSACTION")
        
        # 1. Existence + Current State Check
        cursor.execute(
            "SELECT equipment_id, status, assigned_tech_id FROM erp_equipment WHERE equipment_id = ?",
            (equipment_id,)
        )
        record = cursor.fetchone()
        if not record:
            raise HTTPException(status_code=404, detail="Equipment not found.")
        
        # 2. RETIRED Immutability: Once retired, no further mutations
        if record['status'] == 'RETIRED':
            raise HTTPException(status_code=400, detail="Structural Violation: RETIRED equipment is permanently locked.")
        
        # 3. If retiring, clear tech assignment (equipment removed from service)
        tech_id = payload.assigned_tech_id
        if payload.status == 'RETIRED':
            tech_id = None
        
        # 4. Mutation
        cursor.execute(
            "UPDATE erp_equipment SET status = ?, assigned_tech_id = ? WHERE equipment_id = ?",
            (payload.status, tech_id, equipment_id)
        )
        conn.commit()
        
        return {
            "status": "success",
            "message": f"Equipment {equipment_id} actuated to {payload.status}.",
            "retired": payload.status == 'RETIRED'
        }
    except HTTPException:
        conn.rollback()
        raise
    except sqlite3.IntegrityError as e:
        conn.rollback()
        raise HTTPException(status_code=409, detail=f"FK Violation: {e}")
    except Exception as e:
        conn.rollback()
        logger.error(f"Equipment Actuation Error: {e}")
        raise HTTPException(status_code=500, detail="Internal server error.")
    finally:
        conn.close()

@api_router.post("/admin/ingest/part")
def ingest_part(payload: PartIngestionRecord, jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Administrative clearance required.")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        # PK Synthesization
        import uuid
        new_id = f"PRT-{uuid.uuid4().hex[:6].upper()}"
        
        # Enforce Relational Integrity
        cursor.execute("PRAGMA foreign_keys = ON")
        cursor.execute(
            """
            INSERT INTO erp_parts (part_id, nomenclature, category_id, quantity_on_hand, reorder_threshold, unit_cost)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (new_id, payload.nomenclature, payload.category_id, payload.quantity_on_hand, payload.reorder_threshold, payload.unit_cost)
        )
        conn.commit()
        return {"status": "success", "message": f"Part {new_id} successfully cataloged."}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="Structural Violation: Relational boundary mismatch or nomenclature uniqueness conflict.")
    except Exception as e:
        logger.error(f"Part Ingestion Error: {e}")
        raise HTTPException(status_code=500, detail="Internal Execution Error.")
    finally:
        conn.close()

@api_router.get("/admin/parts")
def get_parts(
    limit: int = Query(50, ge=1, le=100),
    offset: int = Query(0, ge=0),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT p.part_id, p.nomenclature, c.name as category, p.quantity_on_hand, p.reorder_threshold, p.unit_cost 
            FROM erp_parts p
            LEFT JOIN erp_categories c ON p.category_id = c.id
            ORDER BY p.part_id ASC
            LIMIT ? OFFSET ?
            """,
            (limit, offset)
        )
        rows = [dict(row) for row in cursor.fetchall()]
        return {"status": "success", "data": rows}
    except Exception as e:
        logger.error(f"Failed to fetch parts matrix: {e}")
        raise HTTPException(status_code=500, detail="Matrix synchronization failed.")
    finally:
        conn.close()

from datetime import datetime, timezone

def worker_evaluate_threshold(part_id: str):
    """
    Strictly isolated asynchronous threshold evaluation worker.
    Executes outside the HTTP response cycle.
    """
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "maintenance_erp.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        c = conn.cursor()
        
        # 1. Evaluate Threshold Boundary
        c.execute("""
            SELECT s.quantity_on_hand, s.reorder_threshold 
            FROM erp_parts p 
            JOIN erp_skus s ON p.sku_id = s.sku_id 
            WHERE p.part_id = ?
        """, (part_id,))
        part = c.fetchone()
        
        if not part:
            logger.error(f"[WORKER TERMINATION] part_id {part_id} not found in physical ledger.")
            return
            
        if part['quantity_on_hand'] > part['reorder_threshold']:
            # Threshold not breached. Silently terminate CPU cycle.
            return
            
        # 2. Logical Concurrency Check (Save UUID generation and INSERT overhead)
        c.execute("""
            SELECT procurement_id FROM erp_procurement_queue 
            WHERE part_id = ? AND status IN ('PENDING', 'APPROVED')
        """, (part_id,))
        
        if c.fetchone():
            # Active procurement cycle already exists. Prevent supply chain spam.
            logger.info(f"[WORKER TERMINATION] Active procurement already queued for {part_id}.")
            return
            
        # 3. Payload Synthesization
        procurement_id = f"PROC-{uuid.uuid4().hex[:6].upper()}"
        triggered_at = datetime.now(timezone.utc).isoformat()
        
        # 4. Atomic Insertion
        c.execute("""
            INSERT INTO erp_procurement_queue (procurement_id, part_id, triggered_at, status)
            VALUES (?, ?, ?, 'PENDING')
        """, (procurement_id, part_id, triggered_at))
        
        conn.commit()
        logger.warning(f"[ACTUATION ENGAGED] Low Stock Alert: {procurement_id} queued for {part_id}.")
        
    except sqlite3.IntegrityError as e:
        # The partial index intercepted a race condition spanning the logical check.
        conn.rollback()
        logger.info(f"[CONCURRENCY LOCK ENGAGED] Redundant procurement blocked for {part_id}: {e}")
    except Exception as e:
        conn.rollback()
        logger.error(f"[FATAL WORKER EXCEPTION] Evaluation failed for {part_id}: {e}")
    finally:
        conn.close()


@api_router.get("/inventory/available")
def get_available_inventory(
    limit: int = Query(50, ge=1, le=100),
    offset: int = Query(0, ge=0),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    role = jwt_payload.get("role")
    # Universal operational read access
    if role not in ["ADMINISTRATOR", "ADMIN", "HM", "TECH"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Invalid clearance.")

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        # Explicit Enumeration & Zero-Stock Exclusion Enforced
        cursor.execute(
            """
            SELECT p.part_id, s.nomenclature, s.quantity_on_hand 
            FROM erp_parts p
            JOIN erp_skus s ON p.sku_id = s.sku_id
            WHERE p.status = 'IN_STOCK' AND s.quantity_on_hand > 0 
            LIMIT ? OFFSET ?
            """,
            (limit, offset)
        )
        rows = [dict(row) for row in cursor.fetchall()]
        return {"status": "success", "data": rows}
    except Exception as e:
        logger.error(f"Inventory extraction failed: {e}")
        raise HTTPException(status_code=500, detail="Catalog extraction error.")
# [PHASE 37] Procurement Authorization Matrix

@api_router.put("/admin/procurement/{procurement_id}/actuate")
def actuate_procurement(
    procurement_id: str,
    payload: ProcurementActuation,
    jwt_payload: dict = Depends(verify_jwt_token)
):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN", "HM"]:
        raise HTTPException(status_code=403, detail="RBAC Violation.")
        
    if payload.status not in ["APPROVED", "REJECTED", "FULFILLED"]:
        raise HTTPException(status_code=400, detail="Invalid target status.")
        
    if payload.status == "APPROVED" and not payload.authorized_quantity:
        raise HTTPException(status_code=400, detail="Structural Violation: APPROVED state requires an authorized_quantity.")

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        # Escalate to immediate write-lock to prevent read-modify-write race conditions
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        
        cursor.execute("SELECT status, part_id, authorized_quantity FROM erp_procurement_queue WHERE procurement_id = ?", (procurement_id,))
        record = cursor.fetchone()
        
        if not record:
            raise HTTPException(status_code=404, detail="Procurement record not found.")
            
        current_status = record["status"]
        part_id = record["part_id"]
        
        cursor.execute("SELECT sku_id FROM erp_parts WHERE part_id = ?", (part_id,))
        sku_record = cursor.fetchone()
        if not sku_record:
            raise HTTPException(status_code=500, detail="Fatal execution error: part_id in procurement queue does not map to a valid sku_id.")
        sku_id = sku_record["sku_id"]
        
        db_authorized_qty = record["authorized_quantity"]
        
        # State Machine Enforcement
        if current_status in ["REJECTED", "FULFILLED"]:
            raise HTTPException(status_code=400, detail=f"Terminal State Violation: Cannot transition from {current_status}.")
        if current_status == "PENDING" and payload.status not in ["APPROVED", "REJECTED"]:
            raise HTTPException(status_code=400, detail=f"State Violation: PENDING shifts to APPROVED or REJECTED.")
        if current_status == "APPROVED" and payload.status != "FULFILLED":
            raise HTTPException(status_code=400, detail=f"State Violation: APPROVED shifts to FULFILLED.")

        # Actuation & Payload Binding
        if payload.status == "APPROVED":
            cursor.execute(
                "UPDATE erp_procurement_queue SET status = ?, authorized_quantity = ? WHERE procurement_id = ?", 
                (payload.status, payload.authorized_quantity, procurement_id)
            )
        elif payload.status == "REJECTED":
            cursor.execute(
                "UPDATE erp_procurement_queue SET status = ? WHERE procurement_id = ?", 
                (payload.status, procurement_id)
            )
        
        # Autonomous Inventory Hook (FULFILLED only)
        if payload.status == "FULFILLED":
            if not db_authorized_qty:
                raise HTTPException(status_code=500, detail="Fatal execution error: FULFILLED triggered without prior authorized_quantity.")
            
            # Phase 1: Increment Master SKU Ledger
            cursor.execute(
                "UPDATE erp_skus SET quantity_on_hand = quantity_on_hand + ? WHERE sku_id = ?",
                (db_authorized_qty, sku_id)
            )

            # Phase 2: Instantiate Serialized Assets
            # Discrete insertion required for physical asset tracking
            for _ in range(db_authorized_qty):
                cursor.execute(
                    "INSERT INTO erp_parts (part_id, sku_id, status) VALUES (lower(hex(randomblob(16))), ?, 'IN_STOCK')",
                    (sku_id,)
                )

            # Phase 3: Mutate Procurement State
            cursor.execute(
                "UPDATE erp_procurement_queue SET status = 'FULFILLED' WHERE procurement_id = ?",
                (procurement_id,)
            )
            
        conn.commit()
        return {"detail": f"Procurement {procurement_id} mathematically fulfilled. Ledger synchronized."}
        
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Procurement Actuation Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# [PHASE 40] Technician Execution Matrix
@api_router.patch("/mwo/{mwo_id}/execute")
def execute_mwo(
    mwo_id: str,
    payload: MWOExecutePayload,
    jwt_payload: dict = Depends(verify_jwt_token)
):
    role = jwt_payload.get("role")
    tech_id = jwt_payload.get("sub")
    
    if role not in ["TECHNICIAN", "TECH", "ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Clearance required to execute work orders.")

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        if payload.action == "START":
            if role in ["ADMINISTRATOR", "ADMIN"]:
                cursor.execute(
                    "UPDATE work_orders SET status = 'IN_PROGRESS', execution_start = ? WHERE mwo_id = ? AND status IN ('ASSIGNED', 'UNASSIGNED', 'PAUSED')",
                    (time.time(), mwo_id)
                )
            else:
                cursor.execute(
                    "UPDATE work_orders SET status = 'IN_PROGRESS', execution_start = ? WHERE mwo_id = ? AND status IN ('ASSIGNED', 'PAUSED') AND assigned_tech = ?",
                    (time.time(), mwo_id, tech_id)
                )

            if cursor.rowcount == 0:
                raise HTTPException(status_code=409, detail="State Conflict: Work order is either already locked, not assigned to you, or invalid state.")
            
            conn.commit()
            return {"status": "success", "message": f"Execution Lock Acquired on {mwo_id}."}

        elif payload.action == "COMPLETE":
            cursor.execute("SELECT execution_start, accumulated_labor_seconds FROM work_orders WHERE mwo_id = ?", (mwo_id,))
            row = cursor.fetchone()
            if not row or not row["execution_start"]:
                raise HTTPException(status_code=400, detail="Cannot complete a work order that hasn't started.")
            
            end_time = time.time()
            start_time = float(row["execution_start"])
            accumulated = row["accumulated_labor_seconds"] or 0.0
            
            total_labor_seconds = accumulated + (end_time - start_time)
            labor_hours = total_labor_seconds / 3600.0

            if role in ["ADMINISTRATOR", "ADMIN"]:
                cursor.execute(
                    "UPDATE work_orders SET status = 'PENDING_REVIEW', execution_end = ?, labor_hours = ?, manual_log = ? WHERE mwo_id = ? AND status = 'IN_PROGRESS'",
                    (end_time, labor_hours, payload.manual_log, mwo_id)
                )
            else:
                cursor.execute(
                    "UPDATE work_orders SET status = 'PENDING_REVIEW', execution_end = ?, labor_hours = ?, manual_log = ? WHERE mwo_id = ? AND status = 'IN_PROGRESS' AND assigned_tech = ?",
                    (end_time, labor_hours, payload.manual_log, mwo_id, tech_id)
                )

            if cursor.rowcount == 0:
                raise HTTPException(status_code=409, detail="State Conflict: Cannot complete. Verification failed.")
            
            conn.commit()
            return {"status": "success", "message": f"Execution completed for {mwo_id}.", "labor_hours": round(labor_hours, 4)}

        elif payload.action == "PAUSE":
            # Calculate elapsed time and add to accumulated_labor_seconds
            cursor.execute("SELECT execution_start, accumulated_labor_seconds FROM work_orders WHERE mwo_id = ?", (mwo_id,))
            row = cursor.fetchone()
            if not row or not row["execution_start"]:
                raise HTTPException(status_code=400, detail="Cannot pause a work order that hasn't started.")
                
            elapsed = time.time() - float(row["execution_start"])
            new_accumulated = (row["accumulated_labor_seconds"] or 0.0) + elapsed
            
            cursor.execute(
                "UPDATE work_orders SET status = 'PAUSED', execution_start = NULL, accumulated_labor_seconds = ? WHERE mwo_id = ? AND status = 'IN_PROGRESS' AND (assigned_tech = ? OR ? IN ('ADMINISTRATOR', 'ADMIN'))",
                (new_accumulated, mwo_id, tech_id, role)
            )
            if cursor.rowcount == 0:
                raise HTTPException(status_code=409, detail="State Conflict: Cannot pause.")
            
            conn.commit()
            return {"status": "success", "message": f"Execution paused on {mwo_id}."}

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Execution Route Error: {e}")
        raise HTTPException(status_code=500, detail="Internal execution error.")
    finally:
        conn.close()

@api_router.get("/system/directive")
def get_system_directive():
    """Exposes the Orchestration Boundary context to external agents."""
    if not GLOBAL_AI_DIRECTIVE_CONTEXT:
        raise HTTPException(status_code=503, detail="Directive context unavailable or not loaded.")
    return {"status": "success", "directive": GLOBAL_AI_DIRECTIVE_CONTEXT}

@api_router.get("/mwo/technicians")
def get_technicians(jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN", "HM"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Clearance required to view technician roster.")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id as user_id, name FROM erp_employees WHERE role = 'TECH' AND is_active = 1")
        rows = [dict(row) for row in cursor.fetchall()]
        return {"status": "success", "data": rows}
    except Exception as e:
        logger.error(f"Failed to fetch technicians: {e}")
        raise HTTPException(status_code=500, detail="Internal server error while fetching technicians.")
    finally:
        conn.close()

@api_router.get("/mwo/hms")
def get_hms(jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Clearance required to view HM roster.")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id as user_id, name FROM erp_employees WHERE role = 'HM' AND is_active = 1")
        rows = [dict(row) for row in cursor.fetchall()]
        return {"status": "success", "data": rows}
    except Exception as e:
        logger.error(f"Failed to fetch hms: {e}")
        raise HTTPException(status_code=500, detail="Internal server error while fetching hms.")
    finally:
        conn.close()

@api_router.get("/mwo")
def get_mwo(limit: int = 50, offset: int = 0, jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    user_id = jwt_payload.get("sub")
    
    if role not in ["ADMINISTRATOR", "ADMIN", "HM"]:
        raise HTTPException(status_code=403, detail="RBAC Violation")
        
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cols = "w.mwo_id, w.status, w.dm_urgency, w.hm_priority, w.description, w.assigned_tech, w.assigned_hm_id, w.manual_log, w.created_at, w.triaged_at, w.execution_start, w.execution_end, w.completed_at, w.start_date, w.equipment_id, e.nomenclature as equipment_nomenclature, w.location_id, w.material_cost, w.archival_pdf_path, w.labor_hours"

        base_query = f"SELECT {cols} FROM work_orders w LEFT JOIN erp_equipment e ON w.equipment_id = e.equipment_id WHERE w.status IN ('UNASSIGNED', 'ASSIGNED', 'PENDING_REVIEW')"
        params = []
        
        if role == "HM":
            base_query += " AND w.assigned_hm_id = ?"
            params.append(user_id)
            
        base_query += " ORDER BY w.execution_start DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])
        
        cursor.execute(base_query, tuple(params))
        rows = [dict(row) for row in cursor.fetchall()]
        return {"status": "success", "data": rows}
    except Exception as e:
        logger.error(f"Failed to fetch work orders: {e}")
        raise HTTPException(status_code=500, detail="Internal server error while fetching work orders.")
    finally:
        if conn:
            conn.close()



# NOTE: reconcile_inventory_ledger() was removed in the Phase 46 cleanup.
# It decremented warehouse_inventory by free-text SKU and had no callers;
# part consumption and SKU ledger math now live in ingest_mwo_consumption
# (POST /work-orders/{mwo_id}/consume) against mwo_consumed_parts.

@api_router.get("/mwo/assigned")
def get_assigned_mwo(
    limit: int = Query(50), 
    offset: int = Query(0), 
    target_tech: str = Query(None),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    user_id = jwt_payload.get("sub")
    role = jwt_payload.get("role")
    
    # RBAC Impersonation Override
    if target_tech and role in ["ADMINISTRATOR", "ADMIN", "HM"]:
        user_id = target_tech
        
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cols = ", ".join(f"w.{c.strip()}" for c in WORK_ORDER_COLS.split(",")) + \
               ", e.nomenclature as equipment_nomenclature, l.name as location_nomenclature"
        cursor.execute(
            f"SELECT {cols} FROM work_orders w LEFT JOIN erp_equipment e ON w.equipment_id = e.equipment_id LEFT JOIN erp_locations l ON w.location_id = l.id WHERE w.assigned_tech = ? AND w.status IN ('ASSIGNED', 'IN_PROGRESS', 'PAUSED', 'PENDING_REVIEW', 'COMPLETED') ORDER BY w.mwo_id DESC LIMIT ? OFFSET ?",
            (user_id, limit, offset)
        )
        rows = [dict(row) for row in cursor.fetchall()]
        return {"status": "success", "data": rows}
    except Exception as e:
        logger.error(f"Failed to fetch assigned work orders: {e}")
        raise HTTPException(status_code=500, detail="Internal server error while fetching assigned work orders.")
    finally:
        if conn:
            conn.close()

class TechCompletePayload(BaseModel):
    resolution_notes: str = Field(..., min_length=1)
    labor_hours: float = Field(..., gt=0)

async def archive_completed_mwo(mwo_data: dict):
    import asyncio
    queue_dir = "/app/archival_queue"
    # Fallback for local Windows execution
    if not os.path.exists("/app") and os.name == 'nt':
        queue_dir = "C:\\app\\archival_queue"
        
    os.makedirs(queue_dir, exist_ok=True)
    mwo_id = mwo_data.get("mwo_id", "UNKNOWN")
    
    archival_payload = {
        "document_type": "MWO_REPORT",
        "payload": mwo_data
    }
    
    tmp_path = os.path.join(queue_dir, f"{mwo_id}.tmp")
    final_path = os.path.join(queue_dir, f"{mwo_id}.json")
    
    def sync_archive():
        with open(tmp_path, 'w') as f:
            json.dump(archival_payload, f, indent=2)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp_path, final_path)
        
    await asyncio.to_thread(sync_archive)

def archive_mwo_pdf_worker(mwo_id: str, resolution_notes: str, labor_hours: float):
    """
    [ASYNC ISOLATED WORKER]
    Target: CPU-Bound PDF Generation & I/O Archival
    Executes strictly off the primary FastAPI thread via BackgroundTasks.
    """
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        logger.info(f"[WORKER ENGAGED] Initiating asynchronous PDF archival for MWO: {mwo_id}")
        
        import time
        current_time = time.time()
        
        # 1. Explicit Data Extraction (Strict Enumeration)
        cursor.execute(
            """
            SELECT w.mwo_id, w.status, w.assigned_tech, w.equipment_id, e.nomenclature as equipment_name, w.description
            FROM work_orders w
            LEFT JOIN erp_equipment e ON w.equipment_id = e.equipment_id
            WHERE w.mwo_id = ?
            """, 
            (mwo_id,)
        )
        row = cursor.fetchone()
        if not row:
            logger.error(f"[WORKER FATAL] MWO {mwo_id} not found during archival extraction.")
            return
            
        mwo_data = dict(row)
        mwo_data["resolution_notes"] = resolution_notes
        mwo_data["labor_hours"] = labor_hours
        mwo_data["completed_at"] = current_time
        
        if not mwo_data:
            logger.error(f"[WORKER FATAL] MWO {mwo_id} not found during archival extraction.")
            return

        # 1b. Extract consumed parts from mwo_consumed_parts (single source of
        # truth for tech consumption). Aliased to the legacy ledger column names
        # the PDF table below expects.
        cursor.execute(
            """
            SELECT c.consumption_id AS transaction_id, c.part_id, c.quantity_consumed,
                   c.logged_by_tech_id AS tech_id, c.consumed_at AS transaction_timestamp
            FROM mwo_consumed_parts c
            WHERE c.mwo_id = ?
            ORDER BY c.consumed_at ASC
            """,
            (mwo_id,)
        )
        consumed_parts = [dict(row) for row in cursor.fetchall()]
        # consumed_at is a float epoch; render it as a readable timestamp.
        for part in consumed_parts:
            ts = part.get("transaction_timestamp")
            if isinstance(ts, (int, float)):
                part["transaction_timestamp"] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts))

        # 2. Native Python PDF Byte-Compilation (fpdf2 Spatial Doctrine)
        pdf = FPDF()
        pdf.add_page()
        
        # Header
        pdf.set_font("helvetica", style="B", size=18)
        pdf.cell(0, 12, text="MAINTENANCE WORK ORDER ARCHIVE", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(4)
        pdf.set_font("helvetica", size=9)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(0, 6, text=f"Generated from sealed ledger transaction", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.set_text_color(0, 0, 0)
        pdf.ln(8)

        # Separator
        pdf.set_draw_color(16, 185, 129)
        pdf.set_line_width(0.5)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(6)

        # MWO Detail Rows
        detail_fields = [
            ("MWO ID", mwo_data["mwo_id"]),
            ("STATUS", mwo_data["status"]),
            ("ASSIGNED TECH", mwo_data["assigned_tech"] or "N/A"),
            ("EQUIPMENT ID", mwo_data["equipment_id"] or "N/A"),
            ("EQUIPMENT DESC", mwo_data["equipment_name"] or "N/A"),
            ("COMPLETED AT", mwo_data["completed_at"] or "N/A"),
            ("LABOR HOURS", str(mwo_data["labor_hours"] or "N/A")),
        ]

        for label, value in detail_fields:
            pdf.set_font("helvetica", style="B", size=10)
            pdf.cell(50, 8, text=f"{label}:", new_x="RIGHT", new_y="TOP")
            pdf.set_font("helvetica", size=10)
            pdf.cell(0, 8, text=str(value), new_x="LMARGIN", new_y="NEXT")

        pdf.ln(4)

        # Description
        if mwo_data["description"]:
            pdf.set_font("helvetica", style="B", size=10)
            pdf.cell(0, 8, text="DESCRIPTION:", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("helvetica", size=10)
            pdf.multi_cell(0, 6, text=str(mwo_data["description"]), new_x="LMARGIN", new_y="NEXT")
            pdf.ln(4)

        # Resolution Notes
        pdf.set_font("helvetica", style="B", size=10)
        pdf.cell(0, 8, text="RESOLUTION NOTES:", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("helvetica", size=10)
        pdf.multi_cell(0, 6, text=str(mwo_data["resolution_notes"] or "N/A"), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(6)

        # Consumed Parts Table
        if consumed_parts:
            pdf.set_draw_color(16, 185, 129)
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
            pdf.ln(4)
            pdf.set_font("helvetica", style="B", size=11)
            pdf.cell(0, 8, text="CONSUMED PARTS LEDGER", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(2)

            # Table header
            pdf.set_font("helvetica", style="B", size=9)
            pdf.cell(35, 7, text="TXN ID", new_x="RIGHT", new_y="TOP")
            pdf.cell(30, 7, text="PART ID", new_x="RIGHT", new_y="TOP")
            pdf.cell(15, 7, text="QTY", new_x="RIGHT", new_y="TOP")
            pdf.cell(30, 7, text="TECH", new_x="RIGHT", new_y="TOP")
            pdf.cell(0, 7, text="TIMESTAMP", new_x="LMARGIN", new_y="NEXT")

            pdf.set_font("helvetica", size=9)
            for part in consumed_parts:
                pdf.cell(35, 6, text=str(part["transaction_id"]), new_x="RIGHT", new_y="TOP")
                pdf.cell(30, 6, text=str(part["part_id"]), new_x="RIGHT", new_y="TOP")
                pdf.cell(15, 6, text=str(part["quantity_consumed"]), new_x="RIGHT", new_y="TOP")
                pdf.cell(30, 6, text=str(part["tech_id"]), new_x="RIGHT", new_y="TOP")
                pdf.cell(0, 6, text=str(part["transaction_timestamp"] or ""), new_x="LMARGIN", new_y="NEXT")

        # 3. Defensive Physical I/O Archival
        directory_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "archives/work_orders"))
        os.makedirs(directory_path, exist_ok=True)
        
        file_path = f"{directory_path}/{mwo_id}.pdf"
        pdf.output(file_path)
        logger.info(f"[WORKER I/O] Byte-compilation physically written to {file_path}")

        # 4. State Mutation (Schema Linkage & Time Telemetry)
        cursor.execute(
            """
            UPDATE work_orders 
            SET archival_pdf_path = ?, status = 'COMPLETED', completed_at = ?, resolution_notes = ?, labor_hours = ? 
            WHERE mwo_id = ?
            """,
            (file_path, current_time, resolution_notes, labor_hours, mwo_id)
        )
        conn.commit()
        logger.info(f"[WORKER SUCCESS] Archival cycle finalized for MWO: {mwo_id}")

    except Exception as e:
        logger.error(f"[WORKER FATAL] PDF Archival failed for MWO {mwo_id}: {e}")
    finally:
        conn.close()

@api_router.get("/mwo/archive/list")
def get_archive_list(jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    user_id = jwt_payload.get("sub")
    
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        cols = "w.mwo_id, w.status, w.description, w.completed_at, w.equipment_id, e.nomenclature as equipment_nomenclature, w.archival_pdf_path"
        base_query = f"SELECT {cols} FROM work_orders w LEFT JOIN erp_equipment e ON w.equipment_id = e.equipment_id WHERE w.status = 'COMPLETED'"
        params = []
        
        if role == "DM":
            base_query += " AND e.department_id = (SELECT department_id FROM erp_employees WHERE id = ?)"
            params.append(user_id)
        elif role == "HM":
            base_query += " AND w.assigned_hm_id = ?"
            params.append(user_id)
        elif role == "TECH":
            base_query += " AND w.assigned_tech = ?"
            params.append(user_id)
        elif role not in ["ADMINISTRATOR", "ADMIN"]:
            raise HTTPException(status_code=403, detail="RBAC Violation: Invalid role clearance.")
            
        base_query += " ORDER BY w.completed_at DESC"
        
        cursor.execute(base_query, tuple(params))
        rows = [dict(row) for row in cursor.fetchall()]
        return {"status": "success", "data": rows}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to fetch archive list: {e}")
        raise HTTPException(status_code=500, detail="Internal server error while fetching archive list.")
    finally:
        conn.close()

@api_router.get("/mwo/{mwo_id}/consumed-parts")
def get_mwo_consumed_parts(mwo_id: str, jwt_payload: dict = Depends(verify_jwt_token)):
    """Read-only list of parts consumed against an MWO (mwo_consumed_parts is the
    single source of truth), with SKU nomenclature resolved for display."""
    conn = get_db_connection()
    try:
        rows = conn.execute("""
            SELECT c.part_id, p.sku_id, s.nomenclature, c.quantity_consumed, c.consumed_at
            FROM mwo_consumed_parts c
            JOIN erp_parts p ON c.part_id = p.part_id
            JOIN erp_skus s ON p.sku_id = s.sku_id
            WHERE c.mwo_id = ?
            ORDER BY c.consumed_at ASC
        """, (mwo_id,)).fetchall()
        return {"status": "success", "data": [dict(r) for r in rows]}
    finally:
        conn.close()

@api_router.get("/mwo/{mwo_id}/archive")
def retrieve_mwo_archive(mwo_id: str, jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    user_id = jwt_payload.get("sub")
    
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        # 1. Explicit Enumeration & RBAC Retrieval
        cursor.execute(
            """
            SELECT assigned_tech, archival_pdf_path 
            FROM work_orders 
            WHERE mwo_id = ?
            """, 
            (mwo_id,)
        )
        record = cursor.fetchone()
        
        if not record:
            raise HTTPException(status_code=404, detail="MWO ledger entry not found.")
            
        if not record['archival_pdf_path']:
            raise HTTPException(status_code=404, detail="Archival payload has not been generated for this MWO.")

        # 2. Cryptographic Role Matrix Authorization
        if role not in ["ADMINISTRATOR", "ADMIN", "HM"]:
            if role == "DM":
                cursor.execute("SELECT department_id FROM erp_employees WHERE id = ?", (user_id,))
                dm_record = cursor.fetchone()
                if not dm_record:
                    raise HTTPException(status_code=403, detail="RBAC Violation: DM profile not found.")
                
                dm_dept = dm_record['department_id']
                cursor.execute(
                    "SELECT e.department_id FROM work_orders w JOIN erp_equipment e ON w.equipment_id = e.equipment_id WHERE w.mwo_id = ?",
                    (mwo_id,)
                )
                eq_record = cursor.fetchone()
                if not eq_record or eq_record['department_id'] != dm_dept:
                    raise HTTPException(status_code=403, detail="RBAC Violation: MWO does not belong to your department.")
            # Strict Tech Validation: Must be the exactly assigned operator
            elif role == "TECH" and record['assigned_tech'] != user_id:
                raise HTTPException(status_code=403, detail="RBAC Violation: You are not authorized to view this structural archive.")
            elif role not in ["TECH", "DM"]:
                 raise HTTPException(status_code=403, detail="RBAC Violation: Invalid role clearance.")

        # 3. Physical I/O Integrity Verification
        file_path = record['archival_pdf_path']
        if not os.path.exists(file_path):
            logger.error(f"[I/O FATAL] Schema points to missing physical volume: {file_path}")
            raise HTTPException(status_code=500, detail="Physical archival volume corrupted or missing.")

        # 4. Asynchronous File Stream Dispatch
        return FileResponse(
            path=file_path, 
            media_type="application/pdf", 
            filename=f"ARCHIVE_{mwo_id}.pdf"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Archive Retrieval Execution Error: {e}")
        raise HTTPException(status_code=500, detail="Internal server error during stream initialization.")
    finally:
        conn.close()

@api_router.post("/mwo/{mwo_id}/complete", status_code=202)
def complete_mwo(
    mwo_id: str, 
    payload: TechCompletePayload, 
    background_tasks: BackgroundTasks, 
    jwt_payload: dict = Depends(verify_jwt_token)
):
    tech_id = jwt_payload.get("sub")
    role = jwt_payload.get("role")
    
    if role not in ["ADMINISTRATOR", "ADMIN", "HM", "TECH"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Clearance required to finalize MWO.")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        
        # 1. RBAC Parity Check + State Validation
        cursor.execute("SELECT status, assigned_tech FROM work_orders WHERE mwo_id = ?", (mwo_id,))
        mwo_record = cursor.fetchone()
        if not mwo_record:
            raise HTTPException(status_code=404, detail="MWO not found.")
        if mwo_record['status'] == 'COMPLETED':
            raise HTTPException(status_code=400, detail="MWO is already finalized.")
        if mwo_record['status'] not in ['IN_PROGRESS', 'PENDING_REVIEW']:
            raise HTTPException(status_code=400, detail=f"State Violation: MWO must be IN_PROGRESS or PENDING_REVIEW to finalize. Current: {mwo_record['status']}")

        # RBAC Enforcement: TECH-tier must be the assigned operator
        if role in ["TECH"] and mwo_record['assigned_tech'] != tech_id:
            raise HTTPException(status_code=403, detail="RBAC Violation: Technician is not the assigned operator for this MWO.")

        # 2. Asynchronous Dispatch: PDF generation & DB completion in background worker
        background_tasks.add_task(archive_mwo_pdf_worker, mwo_id, payload.resolution_notes, payload.labor_hours)
        
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=202, content={
            "status": "success", 
            "message": f"MWO {mwo_id} accepted. PDF archival and finalization dispatched asynchronously."
        })
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"MWO Completion Error: {e}")
        raise HTTPException(status_code=500, detail="Internal server error during MWO termination.")
    finally:
        conn.close()

@api_router.patch("/mwo/{mwo_id}")
async def update_mwo_v2(mwo_id: str, payload: MWOUpdate, background_tasks: BackgroundTasks, rbac_verified_mwo: dict = Depends(verify_rbac_pipeline)):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        current_time = datetime.now(timezone.utc).isoformat()
        update_data = payload.model_dump(exclude_unset=True)

        if not update_data:
            return {"status": "success", "message": "No changes provided"}
        
        # Guard: Terminal State Prerequisites (COMPLETED / APPROVED)
        # All three fields must be non-null and valid before terminal closure.
        target_status = update_data.get("status")
        if target_status in ("COMPLETED", "APPROVED"):
            cursor.execute("SELECT start_date, equipment_id, assigned_tech FROM work_orders WHERE mwo_id = ?", (mwo_id,))
            check_row = cursor.fetchone()
            if not check_row:
                raise HTTPException(status_code=404, detail=f"MWO {mwo_id} not found.")

            missing = []
            current_tech = check_row['assigned_tech']
            if not current_tech or current_tech.strip().upper() in ('', 'UNASSIGNED'):
                missing.append("assigned_tech")
            if not check_row['equipment_id']:
                missing.append("equipment_id")
            if not check_row['start_date']:
                update_data["start_date"] = current_time

            if missing:
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot transition to {target_status}: missing prerequisites [{', '.join(missing)}]."
                )

            if target_status == "COMPLETED":
                update_data["completed_at"] = current_time

        # Guard: Start Date Injection
        if update_data.get("status") == "IN_PROGRESS":
            cursor.execute("SELECT start_date FROM work_orders WHERE mwo_id = ?", (mwo_id,))
            check_row = cursor.fetchone()
            if check_row and not check_row['start_date']:
                update_data["start_date"] = current_time

        # Dynamic SQL Generation
        columns = []
        values = []
        for key, value in update_data.items():
            columns.append(f"{key} = ?")
            values.append(value)
        
        sql = f"UPDATE work_orders SET {', '.join(columns)} WHERE mwo_id = ?"
        values.append(mwo_id)
        
        cursor.execute(sql, tuple(values))
        conn.commit()
        
        background_tasks.add_task(sync_memory_bus)
        
        # Terminal State Actuation: Atomic Drop Protocol
        if update_data.get("status") == "COMPLETED":
            cols = "mwo_id, status, dm_urgency, hm_priority, description, assigned_tech, manual_log, created_at, triaged_at, execution_start, execution_end, completed_at, start_date, equipment_id, location_id, material_cost, archival_pdf_path, labor_hours"
            cursor.execute(f"SELECT {cols} FROM work_orders WHERE mwo_id = ?", (mwo_id,))
            final_row = cursor.fetchone()
            if final_row:
                background_tasks.add_task(archive_completed_mwo, dict(final_row))
                resolution_notes = final_row['manual_log'] or "No manual log provided."
                labor_hours = final_row['labor_hours'] or 0.0
                background_tasks.add_task(archive_mwo_pdf_worker, mwo_id, resolution_notes, labor_hours)
        
        return {"status": "success", "message": "MWO Updated successfully", "patch": update_data}
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as e:
        logger.error(f"Failed to update MWO {mwo_id}: {e}")
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail="Internal server error while updating MWO.")
    finally:
        if conn:
            conn.close()

async def sync_memory_bus():
    import asyncio
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        # Fix 3: Query Safety (COLLATE NOCASE)
        cols = "mwo_id, status, dm_urgency, hm_priority, description, assigned_tech, manual_log, created_at, triaged_at, execution_start, execution_end, completed_at, start_date, equipment_id, location_id, material_cost, archival_pdf_path, labor_hours"
        cursor.execute(f"SELECT {cols} FROM work_orders WHERE status != 'COMPLETED' COLLATE NOCASE")
        raw_rows = [dict(row) for row in cursor.fetchall()]
        
        # Fix 1: The DTO Mapper
        mapped_rows = []
        for row in raw_rows:
            mapped_row = dict(row)
            if 'technician' in mapped_row:
                mapped_row['assigned_tech'] = mapped_row.pop('technician')
            mapped_rows.append(mapped_row)
        
        payload_data = {
            "agent_id": "MWO_Edge_Node",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "artifact_type": "Active_MWO_State",
            "payload": mapped_rows
        }
        
        shared_memory_dir = "/app/shared_memory"
        os.makedirs(shared_memory_dir, exist_ok=True)
        tmp_path = os.path.join(shared_memory_dir, "mwo_state_broadcast.tmp")
        final_path = os.path.join(shared_memory_dir, "mwo_state_broadcast.json")
        
        # Fix 2: Async I/O (Atomic Swap Protocol)
        def sync_write():
            with open(tmp_path, "w") as f:
                json.dump(payload_data, f, indent=2)
                f.flush()
                os.fsync(f.fileno())
            os.replace(tmp_path, final_path)
                
        await asyncio.to_thread(sync_write)
    except Exception as e:
        logger.error(f"Failed to sync memory bus: {e}")
    finally:
        if conn:
            conn.close()

@api_router.post("/mwo/broadcast")
async def broadcast_mwo_state():
    await sync_memory_bus()
    return {"status": "success", "message": "Broadcast written to shared memory."}

# DEPRECATED: targets the defunct erp_maintenance_logs table (absent from the live schema).
# Re-pathed under /orders/legacy/ so /orders/submit can serve the PO module (Back Office Inventory).
@api_router.post("/orders/legacy/submit")
async def submit_order(order: WorkOrderSubmit):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        # Temporal Fix: Strict UTC ISO 8601 string
        reported_at = datetime.now(timezone.utc).isoformat()
        
        cursor.execute(
            """
            INSERT INTO erp_maintenance_logs 
            (reported_by, asset_id, issue_description, status, reported_at) 
            VALUES (?, ?, ?, ?, ?)
            RETURNING *
            """,
            (order.reported_by, order.asset_id, order.issue_description, "PENDING", reported_at)
        )
        
        inserted_row = cursor.fetchone()
        conn.commit()
        
        return {"status": "success", "data": [dict(inserted_row)]}
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Insert Failed: {e}")
        raise HTTPException(status_code=500, detail="Database insertion failed. Transaction rolled back.")
    finally:
        conn.close()

class AssignUpdate(BaseModel):
    order_id: str
    technician_id: str

@api_router.post("/orders/assign")
async def assign_order(payload: AssignUpdate):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        # Optimistic Locking: Only update if not currently assigned
        cursor.execute(
            """
            UPDATE erp_maintenance_logs 
            SET assigned_to = ? 
            WHERE id = ? AND assigned_to IS NULL
            RETURNING *
            """,
            (payload.technician_id, payload.order_id)
        )
        
        # Concurrency Gate
        if cursor.rowcount == 0:
            conn.rollback()
            raise HTTPException(
                status_code=409, 
                detail="Conflict: Order is already assigned or does not exist."
            )
            
        assigned_row = cursor.fetchone()
        conn.commit()
        
        # Return array structure
        return [dict(assigned_row)]
        
    except HTTPException:
        raise  # Re-raise the 409 Conflict intentionally
    except Exception as e:
        conn.rollback()
        logger.error(f"Assign Failed: {e}")
        raise HTTPException(status_code=500, detail="Internal server error during assignment.")
    finally:
        conn.close()



class NewMWO(BaseModel):
    description: str
    equipment_id: str
    location_id: str
    urgency: str
    assigned_tech: Optional[str] = None
    status: Optional[str] = "PENDING_REVIEW"
    impersonated_creator_id: Optional[str] = None

    @field_validator('equipment_id', 'location_id')
    @classmethod
    def reject_concatenated_strings(cls, v: str) -> str:
        if '(' in v or 'Loc:' in v or ')' in v or ' ' in v:
            raise ValueError("Foreign Keys must be absolute IDs, not concatenated display strings.")
        return v

@api_router.post("/mwo")
async def create_mwo(payload: NewMWO, jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    creator_id = jwt_payload.get("sub")
    
    if role not in ["ADMINISTRATOR", "ADMIN", "DM"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Unauthorized identity.")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        final_creator_id = creator_id
        if payload.impersonated_creator_id and role == "DM":
            # Security check: Verify the impersonated user is in the DM's department
            cursor.execute("SELECT department_id FROM erp_employees WHERE id = ?", (creator_id,))
            dm_dept_row = cursor.fetchone()
            cursor.execute("SELECT department_id FROM erp_employees WHERE id = ?", (payload.impersonated_creator_id,))
            impersonated_dept_row = cursor.fetchone()

            if not dm_dept_row or not impersonated_dept_row or dm_dept_row['department_id'] != impersonated_dept_row['department_id']:
                raise HTTPException(status_code=403, detail="RBAC Violation: Cannot impersonate users outside your department.")
            
            final_creator_id = payload.impersonated_creator_id

        current_time = time.time()
        
        # Auto-generate ID autonomously without client input
        cursor.execute("SELECT mwo_id FROM work_orders WHERE mwo_id LIKE 'MWO-%' ORDER BY mwo_id DESC LIMIT 1")
        last_record = cursor.fetchone()
        next_num = 1
        if last_record and last_record['mwo_id']:
            try:
                # e.g. MWO-2026-003 -> 3
                parts = last_record['mwo_id'].split('-')
                if len(parts) >= 3:
                    last_num = int(parts[-1])
                    next_num = last_num + 1
            except ValueError:
                pass
        current_year = datetime.now(timezone.utc).year
        final_mwo_id = f"MWO-{current_year}-{next_num:03d}"
            
        if not final_mwo_id or not final_mwo_id.strip():
            raise ValueError("Generated MWO ID is invalid.")

        cursor.execute("SELECT assigned_hm_id FROM erp_equipment WHERE equipment_id = ?", (payload.equipment_id,))
        eq_row = cursor.fetchone()
        assigned_hm_id = eq_row['assigned_hm_id'] if eq_row else None

        cursor.execute(
            """
            INSERT INTO work_orders 
            (mwo_id, description, equipment_id, location_id, dm_urgency, assigned_tech, assigned_hm_id, status, hm_priority, execution_start, created_by) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (final_mwo_id, payload.description, payload.equipment_id, payload.location_id, payload.urgency, payload.assigned_tech, assigned_hm_id, "UNASSIGNED", "Normal", current_time, final_creator_id)
        )
        conn.commit()
        
        # Fetch the newly created record
        cursor.execute(f"SELECT {WORK_ORDER_COLS} FROM work_orders WHERE mwo_id = ?", (final_mwo_id,))
        new_row = cursor.fetchone()
        
        return {"status": "success", "message": "MWO created successfully", "data": dict(new_row) if new_row else {}}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Failed to create MWO: {e}")
        raise HTTPException(status_code=500, detail=str(e) or "Internal server error while creating MWO.")
    finally:
        conn.close()

import shutil
from pydantic import ValidationError
from typing import Literal

# STRICT SCHEMA ENFORCEMENT
class UserUploadSchema(BaseModel):
    user_id: str = Field(..., min_length=1)
    name: str = Field(..., min_length=1)
    role: Literal['ADMIN', 'HD', 'HM', 'TECH']
    department: str
    reports_to_hm_id: Optional[str] = None

    @field_validator('reports_to_hm_id', mode='before')
    @classmethod
    def empty_string_to_none(cls, v):
        return None if v == "" else v

def process_csv_background(file_path: str):
    conn = get_db_connection()
    # Enforce WAL mode for concurrent read/writes
    conn.execute("PRAGMA journal_mode=WAL;") 
    cursor = conn.cursor()
    
    try:
        rows_processed = 0
        errors = 0
        with open(file_path, mode='r', encoding='utf-8') as f:
            csv_reader = csv.DictReader(f)
            
            for raw_row in csv_reader:
                try:
                    # Rigorous Schema Validation before DB interaction
                    valid_row = UserUploadSchema(**raw_row)
                    
                    cursor.execute("""
                        INSERT INTO users (user_id, name, role, department, reports_to_hm_id)
                        VALUES (?, ?, ?, ?, ?)
                        ON CONFLICT(user_id) DO UPDATE SET
                            name=excluded.name,
                            role=excluded.role,
                            department=excluded.department,
                            reports_to_hm_id=excluded.reports_to_hm_id
                    """, (
                        valid_row.user_id,
                        valid_row.name,
                        valid_row.role,
                        valid_row.department,
                        valid_row.reports_to_hm_id
                    ))
                    rows_processed += 1
                except ValidationError as ve:
                    logger.error(f"Schema violation for row {raw_row}: {ve}")
                    errors += 1
                    continue # Skip invalid rows, do not crash batch
                except Exception as e:
                    logger.error(f"DB Error for row {raw_row}: {e}")
                    errors += 1
                    continue
                    
        conn.commit()
        logger.info(f"[BACKGROUND WORKER] CSV Ingestion Complete. Processed: {rows_processed}, Skipped: {errors}")
    except Exception as e:
        conn.rollback()
        logger.error(f"Critical failure in background CSV processing: {e}")
    finally:
        conn.close()
        # Autonomous cleanup of temp file
        if os.path.exists(file_path):
            os.remove(file_path)


import concurrent.futures

def hash_pin_cpu_bound(pin: str) -> str:
    # Strictly CPU-bound bcrypt hashing
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(pin.encode('utf-8'), salt).decode('utf-8')

def process_employee_csv_background(file_path: str):
    conn = get_db_connection()
    conn.execute("PRAGMA journal_mode=WAL;") 
    cursor = conn.cursor()
    
    try:
        rows_processed = 0
        errors = 0
        valid_records = []
        
        # Phase 1: I/O & Validation
        with open(file_path, mode='r', encoding='utf-8') as f:
            csv_reader = csv.DictReader(f)
            for raw_row in csv_reader:
                try:
                    record = EmployeeIngestionRecord(**raw_row)
                    valid_records.append(record)
                except Exception as ve:
                    logger.error(f"Schema violation for row: {ve}")
                    errors += 1
                    
        # Phase 2: CPU-Bound Hashing offloaded to ThreadPool
        hashed_records = []
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future_to_record = {executor.submit(hash_pin_cpu_bound, r.pin_code): r for r in valid_records}
            for future in concurrent.futures.as_completed(future_to_record):
                record = future_to_record[future]
                try:
                    hashed_pin = future.result()
                    hashed_records.append((
                        record.id,
                        record.name,
                        record.authorization_level,
                        record.pin_code,  # Storing plain pin_code just for legacy structural alignment if required, but wait!
                        hashed_pin,
                        record.is_active
                    ))
                except Exception as e:
                    logger.error(f"Hashing failed for {record.id}: {e}")
                    errors += 1
                    
        # Phase 3: Atomic Batch SQLite Insertion
        if hashed_records:
            cursor.execute("BEGIN IMMEDIATE")
            cursor.executemany("""
                INSERT INTO erp_employees (id, name, authorization_level, pin_code, pin_hash, is_active)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(id) DO UPDATE SET
                    name=excluded.name,
                    authorization_level=excluded.authorization_level,
                    pin_code=excluded.pin_code,
                    pin_hash=excluded.pin_hash,
                    is_active=excluded.is_active
            """, hashed_records)
            conn.commit()
            rows_processed += len(hashed_records)
            
        logger.info(f"[BACKGROUND WORKER] Employee Ingestion Complete. Processed: {rows_processed}, Skipped: {errors}")
    except Exception as e:
        conn.rollback()
        logger.error(f"Critical failure in employee background processing: {e}")
    finally:
        conn.close()
        if os.path.exists(file_path):
            os.remove(file_path)

@api_router.post("/admin/employees/bulk-upload", status_code=202)
async def bulk_upload_employees(background_tasks: BackgroundTasks, jwt_payload: dict = Depends(verify_jwt_token), file: UploadFile = File(...)):
    if jwt_payload.get("role") not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: ADMIN clearance required.")
        
    if not file.filename.endswith((".csv", ".json")):
        raise HTTPException(status_code=400, detail="Strictly CSV/JSON payloads authorized.")
    
    tmp_path = f"/tmp/emp_{uuid.uuid4().hex}_{file.filename}"
    
    try:
        async with aiofiles.open(tmp_path, 'wb') as out_file:
            while content := await file.read(1024 * 1024):  
                await out_file.write(content)
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to write payload to temporary storage.")
    finally:
        await file.close()

    background_tasks.add_task(process_employee_csv_background, tmp_path)
    
    return {"status": "accepted", "message": "Employee payload queued for secure validation and processing."}

def validate_mwo_chunk(raw_chunk: list) -> tuple:
    valid = []
    errors = 0
    for raw_row in raw_chunk:
        try:
            r = MWOIngestionRecord(**raw_row)
            # consumed_sku stays on the model so legacy CSVs still
            # validate, but is no longer bound — Phase 46 moved
            # consumption to mwo_consumed_parts.
            valid.append((
                r.mwo_id, r.status, r.dm_urgency, r.hm_priority,
                r.description, r.assigned_tech,
                r.manual_log,
                r.start_date.isoformat() if r.start_date else None,
                r.equipment_id
            ))
        except Exception:
            errors += 1
    return valid, errors

def process_mwo_csv_background(file_path: str):
    try:
        rows_processed = 0
        errors = 0
        validated_batches = []
        chunk_size = 5000
        
        # Phase 1: CPU-Bound Decoupling with Strictly Streamed I/O
        with open(file_path, mode='r', encoding='utf-8') as f:
            csv_reader = csv.DictReader(f)
            
            with concurrent.futures.ProcessPoolExecutor() as executor:
                futures = []
                while True:
                    # Stream exactly 5000 rows into memory, preventing OOM
                    chunk = list(islice(csv_reader, chunk_size))
                    if not chunk:
                        break
                    futures.append(executor.submit(validate_mwo_chunk, chunk))
                
                for future in concurrent.futures.as_completed(futures):
                    try:
                        valid_batch, batch_errors = future.result()
                        errors += batch_errors
                        if valid_batch:
                            validated_batches.append(valid_batch)
                    except Exception as e:
                        logger.error(f"ProcessPool execution error: {e}")

        # Phase 2 & 3: Deferred DB Connection & Atomic Batch Write
        conn = get_db_connection()
        try:
            conn.execute("PRAGMA journal_mode=WAL;")
            conn.execute("PRAGMA synchronous=NORMAL;")
            
            for batch in validated_batches:
                cursor = conn.cursor()
                cursor.execute("BEGIN IMMEDIATE")
                try:
                    cursor.executemany("""
                        INSERT INTO work_orders (
                            mwo_id, status, dm_urgency, hm_priority, description,
                            assigned_tech, manual_log, start_date, equipment_id
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(mwo_id) DO UPDATE SET
                            status=excluded.status,
                            dm_urgency=excluded.dm_urgency,
                            hm_priority=excluded.hm_priority,
                            description=excluded.description,
                            assigned_tech=excluded.assigned_tech,
                            manual_log=excluded.manual_log,
                            start_date=excluded.start_date,
                            equipment_id=excluded.equipment_id
                    """, batch)
                    conn.commit()
                    rows_processed += len(batch)
                except Exception as e:
                    conn.rollback()
                    logger.error(f"Batch write failure: {e}")
        finally:
            conn.close()
            
        logger.info(f"[BACKGROUND WORKER] Ingestion Complete. Processed: {rows_processed}, Skipped: {errors}")
        
    except Exception as e:
        logger.error(f"Critical failure in MWO processing: {e}")
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)

@api_router.post("/admin/mwo/bulk-upload", status_code=202)
async def bulk_upload_mwos(background_tasks: BackgroundTasks, jwt_payload: dict = Depends(verify_jwt_token), file: UploadFile = File(...)):
    if jwt_payload.get("role") not in ["ADMINISTRATOR", "ADMIN", "DM", "HM"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Insufficient clearance for bulk MWO ingestion.")
        
    if not file.filename.endswith((".csv", ".json")):
        raise HTTPException(status_code=400, detail="Strictly CSV/JSON payloads authorized.")
    
    tmp_path = f"/tmp/mwo_{uuid.uuid4().hex}_{file.filename}"
    
    try:
        async with aiofiles.open(tmp_path, 'wb') as out_file:
            while content := await file.read(1024 * 1024):  
                await out_file.write(content)
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to write payload to temporary storage.")
    finally:
        await file.close()

    background_tasks.add_task(process_mwo_csv_background, tmp_path)
    
    return {"status": "accepted", "message": "MWO payload queued for secure validation and processing."}

@api_router.post("/admin/users/bulk-upload", status_code=202)
async def bulk_upload_users(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Strictly CSV payloads authorized.")
    
    tmp_path = f"/tmp/{file.filename}"
    
    # Asynchronous chunked writing strictly prevents ASGI event loop blocking
    try:
        async with aiofiles.open(tmp_path, 'wb') as out_file:
            while content := await file.read(1024 * 1024):  # 1MB chunks
                await out_file.write(content)
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to write payload to temporary storage.")
    finally:
        await file.close()

    # Offload strictly to background thread
    background_tasks.add_task(process_csv_background, tmp_path)
    
    return {"status": "accepted", "message": "Payload queued for validation and processing."}

@api_router.get("/orders/active")
async def get_active_orders():
    """
    Retrieves all non-closed maintenance work orders, ordered newest first.
    Strictly utilizes the new local SQLite adapter for concurrent stability.
    """
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        query = """
            SELECT 
                id,
                reported_by,
                asset_id,
                issue_description,
                status,
                assigned_to,
                resolved_at,
                reported_at
            FROM erp_maintenance_logs
            WHERE status != 'CLOSED'
            ORDER BY reported_at DESC
        """
        
        cursor.execute(query)
        rows = cursor.fetchall()
        
        orders = []
        for row in rows:
            orders.append({
                "id": row["id"],
                "reported_by": row["reported_by"],
                "asset_id": row["asset_id"],
                "issue_description": row["issue_description"],
                "status": row["status"],
                "assigned_to": row["assigned_to"],
                "resolved_at": row["resolved_at"],
                "reported_at": row["reported_at"]
            })
            
        return {"status": "success", "orders": orders}
        
    except Exception as e:
        logger.error(f"Active orders telemetry query failed: {e}")
        raise HTTPException(status_code=500, detail="Database Error: Unable to fetch active telemetry.")
        
    finally:
        if conn:
            conn.close()


@api_router.get("/admin/users")
async def get_admin_users(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=100),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    # Strict RBAC Enforcement
    actor_role = jwt_payload.get("role")
    if actor_role != "ADMIN":
        raise HTTPException(status_code=403, detail="RBAC Violation: ADMIN clearance required.")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        # Mathematical SQLite Pagination
        offset = (page - 1) * limit
        
        # Enforce Soft Delete Boundary
        cursor.execute("""
            SELECT user_id, name, role, department, reports_to_hm_id, is_active
            FROM users
            WHERE is_active = 1
            ORDER BY user_id ASC
            LIMIT ? OFFSET ?
        """, (limit, offset))
        
        rows = cursor.fetchall()
        
        # C-Optimized Serialization
        users = [{
            "user_id": row["user_id"],
            "name": row["name"],
            "role": row["role"],
            "department": row["department"],
            "reports_to_hm_id": row["reports_to_hm_id"],
            "is_active": row["is_active"]
        } for row in rows]
            
        return users
        
    except Exception as e:
        logger.error(f"Failed to fetch paginated matrix: {e}")
        raise HTTPException(status_code=500, detail="Database Error: Unable to fetch telemetry.")
    finally:
        if conn:
            conn.close()
            
@api_router.get("/admin/users/{user_id}/audit-log")
async def get_user_audit_log(
    user_id: str = Path(..., description="The target user's enterprise ID"),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    actor_role = jwt_payload.get("role")
    
    # Strict Taxonomy Validation
    if actor_role != "ADMIN":

        raise HTTPException(status_code=403, detail="RBAC Violation: ADMIN clearance required.")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT event_id, target_user_id, actor_user_id, action_type, timestamp 
            FROM user_audit_logs 
            WHERE target_user_id = ? 
            ORDER BY timestamp DESC
        """, (user_id,))
        
        rows = cursor.fetchall()
        events = [{"event_id": r["event_id"], "target_user_id": r["target_user_id"], "actor_user_id": r["actor_user_id"], "action": r["action_type"], "timestamp": r["timestamp"]} for r in rows]
            
        return {"target_user_id": user_id, "events": events}
    except Exception as e:
        logger.error(f"Audit log retrieval failed for {user_id}: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve contextual telemetry.")
    finally:
        conn.close()

# --- [RESTRICTED SKU PROCUREMENT] Admin SKU Clearance Management ---

class SkuAssignPayload(BaseModel):
    sku_id: str = Field(..., min_length=1)


def _require_sku_access_admin(jwt_payload: dict):
    if jwt_payload.get("role") not in SKU_ACCESS_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: ADMIN clearance required.")


@api_router.get("/admin/users/{employee_id}/skus")
def get_user_sku_access(employee_id: str = Path(...), jwt_payload: dict = Depends(verify_jwt_token)):
    """Admin view: a user's assigned SKU clearances plus the assignable remainder."""
    _require_sku_access_admin(jwt_payload)
    conn = get_db_connection()
    try:
        if not conn.execute("SELECT 1 FROM erp_employees WHERE id = ?", (employee_id,)).fetchone():
            raise HTTPException(status_code=404, detail=f"Employee {employee_id} not found.")
        assigned = [dict(r) for r in conn.execute("""
            SELECT s.sku_id, s.nomenclature
            FROM erp_employee_sku_access a
            JOIN erp_skus s ON s.sku_id = a.sku_id
            WHERE a.employee_id = ?
            ORDER BY s.nomenclature
        """, (employee_id,)).fetchall()]
        available = [dict(r) for r in conn.execute("""
            SELECT s.sku_id, s.nomenclature
            FROM erp_skus s
            WHERE s.sku_id NOT IN (SELECT sku_id FROM erp_employee_sku_access WHERE employee_id = ?)
            ORDER BY s.nomenclature
        """, (employee_id,)).fetchall()]
        return {"assigned": assigned, "available": available}
    finally:
        conn.close()


@api_router.post("/admin/users/{employee_id}/skus", status_code=201)
def assign_user_sku_access(employee_id: str, payload: SkuAssignPayload, jwt_payload: dict = Depends(verify_jwt_token)):
    """Map a SKU clearance to an employee."""
    _require_sku_access_admin(jwt_payload)
    conn = get_db_connection()
    try:
        if not conn.execute("SELECT 1 FROM erp_employees WHERE id = ?", (employee_id,)).fetchone():
            raise HTTPException(status_code=404, detail=f"Employee {employee_id} not found.")
        if not conn.execute("SELECT 1 FROM erp_skus WHERE sku_id = ?", (payload.sku_id,)).fetchone():
            raise HTTPException(status_code=404, detail=f"SKU {payload.sku_id} not found in master catalog.")
        conn.execute(
            "INSERT OR IGNORE INTO erp_employee_sku_access (employee_id, sku_id) VALUES (?, ?)",
            (employee_id, payload.sku_id)
        )
        conn.commit()
        logger.info(f"[SKU ACCESS] {jwt_payload.get('sub')} granted {payload.sku_id} to {employee_id}.")
        return {"status": "success", "employee_id": employee_id, "sku_id": payload.sku_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"SKU access grant failed: {e}")
        raise HTTPException(status_code=500, detail="SKU clearance assignment failed.")
    finally:
        conn.close()


@api_router.delete("/admin/users/{employee_id}/skus/{sku_id}", status_code=204)
def revoke_user_sku_access(employee_id: str, sku_id: str, jwt_payload: dict = Depends(verify_jwt_token)):
    """Revoke a SKU clearance from an employee."""
    _require_sku_access_admin(jwt_payload)
    conn = get_db_connection()
    try:
        cur = conn.execute(
            "DELETE FROM erp_employee_sku_access WHERE employee_id = ? AND sku_id = ?",
            (employee_id, sku_id)
        )
        conn.commit()
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail=f"No clearance for {sku_id} on {employee_id}.")
        logger.info(f"[SKU ACCESS] {jwt_payload.get('sub')} revoked {sku_id} from {employee_id}.")
        return Response(status_code=204)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"SKU access revoke failed: {e}")
        raise HTTPException(status_code=500, detail="SKU clearance revocation failed.")
    finally:
        conn.close()

@api_router.delete("/admin/users/{user_id}", status_code=204)
async def terminate_user_access(
    user_id: str = Path(..., description="The target user's enterprise ID to terminate"),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    actor_user_id = jwt_payload.get("sub")
    actor_role = jwt_payload.get("role")

    # Strict Taxonomy Validation
    if actor_role != "ADMIN":
        raise HTTPException(status_code=403, detail="RBAC Violation: ADMIN clearance required.")
        
    if actor_user_id == user_id:
        raise HTTPException(status_code=400, detail="Self-termination is architecturally prohibited.")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        # Native Context Manager for Guaranteed Atomicity
        with conn:
            cursor.execute("SELECT is_active FROM users WHERE user_id = ?", (user_id,))
            target_user = cursor.fetchone()
            
            if not target_user:
                raise HTTPException(status_code=404, detail="Target user not found in the matrix.")
            if target_user["is_active"] == 0:
                raise HTTPException(status_code=400, detail="Target user is already structurally terminated.")
                
            cursor.execute("""
                UPDATE users 
                SET is_active = 0, token_version = token_version + 1 
                WHERE user_id = ?
            """, (user_id,))
            
            event_id = str(uuid.uuid4())
            action_type = "TERMINATE_ACCESS"
            
            cursor.execute("""
                INSERT INTO user_audit_logs (event_id, target_user_id, actor_user_id, action_type) 
                VALUES (?, ?, ?, ?)
            """, (event_id, user_id, actor_user_id, action_type))
            
        return None  
        
    except HTTPException:
        raise
    except sqlite3.Error as e:
        logger.error(f"Atomic transaction failed during structural termination: {e}")
        raise HTTPException(status_code=500, detail="Database constraints aborted the structural termination.")
    except Exception as e:
        logger.error(f"Unexpected error during termination sequence: {e}")
        raise HTTPException(status_code=500, detail="An unexpected error aborted the actuation.")
    finally:
        conn.close()

class UserEscalationPayload(BaseModel):
    role: Literal['ADMIN', 'DM', 'HM', 'TECH']
    department: str

@api_router.put("/admin/users/{user_id}/escalate")
async def escalate_user(
    payload: UserEscalationPayload,
    response: Response,
    user_id: str = Path(..., description="The target user's enterprise ID"),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    actor_user_id = jwt_payload.get("sub")
    actor_role = jwt_payload.get("role")

    # Strict Taxonomy Validation
    if actor_role != "ADMIN":
        raise HTTPException(status_code=403, detail="RBAC Violation: ADMIN clearance required.")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE")
        
        cursor.execute("SELECT is_active FROM users WHERE user_id = ?", (user_id,))
        target_user = cursor.fetchone()
        
        if not target_user:
            raise HTTPException(status_code=404, detail="Target user not found in the matrix.")
        if target_user["is_active"] == 0:
            raise HTTPException(status_code=400, detail="Target user is structurally terminated.")
            
        cursor.execute("""
            UPDATE users 
            SET role = ?, department = ?, token_version = token_version + 1 
            WHERE user_id = ?
        """, (payload.role, payload.department, user_id))
        
        event_id = str(uuid.uuid4())
        action_type = "ROLE_ESCALATION"
        
        cursor.execute("""
            INSERT INTO user_audit_logs (event_id, target_user_id, actor_user_id, action_type) 
            VALUES (?, ?, ?, ?)
        """, (event_id, user_id, actor_user_id, action_type))
        
        conn.commit()
        
        if actor_user_id == user_id:
            # Return specific flush signal for synchronous client-side JWT invalidation
            response.status_code = 205
            response.headers["X-Token-Flush"] = "true"
            return None
        else:
            return {"status": "success", "message": "Role escalation successful."}
        
    except HTTPException:
        conn.rollback()
        raise
    except sqlite3.Error as e:
        conn.rollback()
        logger.error(f"Atomic transaction failed during structural escalation: {e}")
        raise HTTPException(status_code=500, detail="Database constraints aborted the structural escalation.")
    except Exception as e:
        conn.rollback()
        logger.error(f"Unexpected error during escalation sequence: {e}")
        raise HTTPException(status_code=500, detail="An unexpected error aborted the actuation.")
    finally:
        conn.close()

# [PHASE 36.1] MWO Queue & Assignment Dispatch

@api_router.get("/work-orders/queue")
def get_work_orders_queue(
    limit: int = Query(25, ge=1, le=100),
    offset: int = Query(0, ge=0),
    status: Optional[str] = Query(None),
    target_dm: Optional[str] = Query(None),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        role = jwt_payload.get("role")
        user_id = jwt_payload.get("sub")
        
        active_user = target_dm if target_dm and role in ["ADMINISTRATOR", "ADMIN"] else user_id
        active_role = "DM" if target_dm and role in ["ADMINISTRATOR", "ADMIN"] else role
        
        query = "SELECT COUNT(*) as total_count FROM work_orders w"
        data_query = "SELECT w.mwo_id, w.status, w.dm_urgency, w.hm_priority, w.equipment_id, w.description, w.created_at FROM work_orders w"
        params = []
        
        if active_role == "DM":
            query += " JOIN erp_equipment e ON w.equipment_id = e.equipment_id WHERE e.department_id = (SELECT department_id FROM erp_employees WHERE id = ?)"
            data_query += " JOIN erp_equipment e ON w.equipment_id = e.equipment_id WHERE e.department_id = (SELECT department_id FROM erp_employees WHERE id = ?)"
            params.append(active_user)
        else:
            query += " WHERE 1=1"
            data_query += " WHERE 1=1"
        
        if status:
            query += " AND w.status = ?"
            data_query += " AND w.status = ?"
            params.append(status)
            
        cursor.execute(query, params)
        count_row = cursor.fetchone()
        total_count = count_row['total_count'] if count_row else 0
        
        data_query += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])
        
        cursor.execute(data_query, params)
        rows = [dict(row) for row in cursor.fetchall()]
        return {"status": "success", "total_count": total_count, "data": rows}
    except Exception as e:
        logger.error(f"Failed to fetch MWO queue: {e}")
        raise HTTPException(status_code=500, detail="Queue synchronization failed.")
    finally:
        conn.close()

class AssignMWOPayload(BaseModel):
    assigned_tech_id: str
    hm_priority: str = "Normal"

@api_router.patch("/mwo/{mwo_id}/assign")
def assign_mwo(
    mwo_id: str,
    payload: AssignMWOPayload,
    jwt_payload: dict = Depends(verify_jwt_token)
):
    role = jwt_payload.get("role")
    user_id = jwt_payload.get("sub")
    
    if role not in ["ADMINISTRATOR", "ADMIN", "HM"]:
        raise HTTPException(status_code=403, detail="RBAC Violation")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        
        cursor.execute("SELECT is_active, department_id FROM erp_employees WHERE id = ? AND role = 'TECH'", (payload.assigned_tech_id,))
        tech_emp = cursor.fetchone()
        
        if not tech_emp:
            raise HTTPException(status_code=404, detail="Technician not found or invalid role.")
            
        if not tech_emp['is_active']:
            raise HTTPException(status_code=400, detail="Structural Violation: Cannot assign to an inactive technician.")
            
        if role == "HM":
            cursor.execute("SELECT department_id FROM erp_employees WHERE id = ?", (user_id,))
            hm_emp = cursor.fetchone()
            if hm_emp and hm_emp['department_id'] != tech_emp['department_id']:
                raise HTTPException(status_code=403, detail="RBAC Violation: Cannot assign a technician outside your departmental isolation boundary.")
            
        triaged_time = time.time()
        
        cursor.execute(
            """
            UPDATE work_orders 
            SET assigned_tech = ?, hm_priority = ?, triaged_at = ?, status = 'ASSIGNED' 
            WHERE mwo_id = ? AND status IN ('UNASSIGNED', 'UNASSIGNED_ESCALATION')
            """,
            (payload.assigned_tech_id, payload.hm_priority, triaged_time, mwo_id)
        )
        
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="MWO not found or already assigned.")
            
        conn.commit()
        return {"status": "success", "message": f"MWO {mwo_id} successfully assigned to {payload.assigned_tech_id}."}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Assignment Actuation Error: {e}")
        raise HTTPException(status_code=500, detail="Assignment failed.")
    finally:
        conn.close()

@api_router.get("/employees")
def get_employees(
    role: Optional[str] = Query(None),
    is_active: Optional[int] = Query(None),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        query = "SELECT id, name, role, is_active FROM erp_employees WHERE 1=1"
        params = []
        if role:
            query += " AND role = ?"
            params.append(role)
        if is_active is not None:
            query += " AND is_active = ?"
            params.append(is_active)
            
        cursor.execute(query, params)
        rows = [dict(row) for row in cursor.fetchall()]
        return {"status": "success", "data": rows}
    finally:
        conn.close()

# Personnel Bulk
@api_router.get("/admin/ingest/personnel/template")
def get_personnel_ingestion_template(jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation.")
    
    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("SELECT name FROM erp_departments")
        departments = [row['name'] for row in c.fetchall()]
        c.execute("SELECT name FROM erp_employees WHERE role='HM'")
        hms = [row['name'] for row in c.fetchall()]
    finally:
        conn.close()
        
    headers = ["name", "role", "pin_code", "department_name", "reports_to_hm_name"]
    return generate_xlsx_template(headers, "personnel_ingestion_template", departments=departments, hms=hms)

@api_router.post("/admin/ingest/personnel/bulk")
async def bulk_ingest_personnel(file: UploadFile = File(...), jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation")
        
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Invalid file format.")
        
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        rows = []
        for row_values in ws.iter_rows(min_row=2, values_only=True):
            if not any(row_values): continue
            row_dict = dict(zip(header_row, row_values))
            rows.append(row_dict)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read XLSX: {e}")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("PRAGMA foreign_keys = ON")
        
        cursor.execute("SELECT id, name FROM erp_departments")
        dep_map = {row['name'].lower().strip(): row['id'] for row in cursor.fetchall()}
        
        cursor.execute("SELECT id, name FROM erp_employees WHERE role = 'HM'")
        hm_map = {row['name'].lower().strip(): row['id'] for row in cursor.fetchall()}

        cursor.execute("BEGIN TRANSACTION")
        
        import bcrypt
        
        # Pass 1: Parse and validate all users, assign IDs, and update hm_map
        parsed_users = []
        for i, row in enumerate(rows):
            row_data = {k.strip(): str(v).strip() if v is not None else None for k, v in row.items()}
            
            name = row_data.get('name')
            prole = row_data.get('role', '').upper()
            pin = row_data.get('pin_code')
            if not name or not prole or not pin:
                raise ValueError(f"Row {i+2}: Missing required fields.")
                
            dep_name = row_data.get('department_name')
            if not dep_name:
                raise ValueError(f"Row {i+2}: Missing department_name")
            if dep_name.lower() not in dep_map:
                new_dep = f"DEP-{uuid.uuid4().hex[:6].upper()}"
                cursor.execute("INSERT INTO erp_departments (id, name) VALUES (?, ?)", (new_dep, dep_name.strip()))
                dep_map[dep_name.lower()] = new_dep
            dep_id = dep_map[dep_name.lower()]
            
            new_id = f"U-{uuid.uuid4().hex[:6].upper()}"
            
            if prole in ['HM', 'DM', 'ADMINISTRATOR', 'ADMIN']:
                hm_map[name.lower().strip()] = new_id
                
            parsed_users.append({
                'row_index': i + 2,
                'id': new_id,
                'name': name,
                'role': prole,
                'pin': pin,
                'department_id': dep_id,
                'reports_to': row_data.get('reports_to_hm_name')
            })

        # Sort so that Managers are inserted before Technicians to satisfy Foreign Key constraints
        parsed_users.sort(key=lambda u: 0 if u['role'] in ['ADMINISTRATOR', 'ADMIN', 'DM', 'HM'] else 1)

        # Pass 2: Resolve HM relationships and insert
        inserted_count = 0
        for user in parsed_users:
            hm_name = user['reports_to']
            hm_id = None
            if hm_name:
                if hm_name.lower().strip() not in hm_map:
                    raise ValueError(f"Row {user['row_index']}: Unknown HM Name '{hm_name}'")
                hm_id = hm_map[hm_name.lower().strip()]
                
            if user['role'] in ['TECH', 'TECHNICIAN'] and not hm_id:
                raise ValueError(f"Row {user['row_index']}: TECH must have a reports_to_hm_name")

            pin_hash = bcrypt.hashpw(user['pin'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            cursor.execute(
                "INSERT INTO erp_employees (id, name, role, pin_hash, is_active, department_id, reports_to_hm_id) VALUES (?, ?, ?, ?, 1, ?, ?)",
                (user['id'], user['name'], user['role'], pin_hash, user['department_id'], hm_id)
            )
            inserted_count += 1
            
        conn.commit()
        return {"status": "success", "message": f"Successfully ingested {inserted_count} personnel."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# Parts Bulk
@api_router.get("/admin/ingest/part/template")
def get_part_ingestion_template(jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation.")
    
    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("SELECT name FROM erp_categories")
        categories = [row['name'] for row in c.fetchall()]
    finally:
        conn.close()
        
    headers = ["nomenclature", "category_name", "quantity_on_hand", "reorder_threshold", "unit_cost"]
    return generate_xlsx_template(headers, "part_ingestion_template", categories=categories)

@api_router.post("/admin/ingest/part/bulk")
async def bulk_ingest_part(file: UploadFile = File(...), jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation")
        
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Invalid file format.")
        
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        rows = []
        for row_values in ws.iter_rows(min_row=2, values_only=True):
            if not any(row_values): continue
            row_dict = dict(zip(header_row, row_values))
            rows.append(row_dict)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read XLSX: {e}")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("PRAGMA foreign_keys = ON")
        
        cursor.execute("SELECT id, name FROM erp_categories")
        cat_map = {row['name'].lower().strip(): row['id'] for row in cursor.fetchall()}

        cursor.execute("BEGIN TRANSACTION")
        
        inserted_count = 0
        for i, row in enumerate(rows):
            row_data = {k.strip(): str(v).strip() if v is not None else None for k, v in row.items()}
            
            cat_name = row_data.get('category_name')
            if not cat_name:
                raise ValueError(f"Row {i+2}: Missing category_name")
            if cat_name.lower() not in cat_map:
                new_cat = f"CAT-{uuid.uuid4().hex[:6].upper()}"
                cursor.execute("INSERT INTO erp_categories (id, name) VALUES (?, ?)", (new_cat, cat_name.strip()))
                cat_map[cat_name.lower()] = new_cat
            cat_id = cat_map[cat_name.lower()]

            new_id = f"PRT-{uuid.uuid4().hex[:6].upper()}"
            
            cursor.execute(
                "INSERT INTO erp_parts (part_id, nomenclature, category_id, quantity_on_hand, reorder_threshold, unit_cost) VALUES (?, ?, ?, ?, ?, ?)",
                (new_id, row_data.get('nomenclature'), cat_id, int(row_data.get('quantity_on_hand', 0)), int(row_data.get('reorder_threshold', 5)), float(row_data.get('unit_cost', 0.0)))
            )
            inserted_count += 1
            
        conn.commit()
        return {"status": "success", "message": f"Successfully ingested {inserted_count} parts."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# --- PHASE 44 SKU INGESTION & LEDGER ---
@api_router.post("/inventory/skus", status_code=201)
def ingest_sku(payload: SKUCreate, jwt_payload: dict = Depends(verify_jwt_token)):
    # HOD clearance: HM operators manage the back-office catalog alongside admins
    if jwt_payload.get("role") not in ["ADMINISTRATOR", "ADMIN", "HM"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: HOD clearance required.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")

        supplier_id = payload.supplier_id
        supplier_created = False
        if payload.new_supplier:
            # Atomic inline registration: supplier insert + SKU insert share this
            # transaction; an SKU failure rolls the supplier back (no orphans).
            sup = payload.new_supplier
            supplier_id = sup.supplier_id
            cursor.execute("SELECT 1 FROM erp_suppliers WHERE supplier_id = ?", (supplier_id,))
            if cursor.fetchone():
                raise HTTPException(status_code=409, detail=f"Supplier {supplier_id} already exists.")
            cursor.execute("""
                INSERT INTO erp_suppliers (supplier_id, name, email, phone, address, default_lead_time_days)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (supplier_id, sup.name, sup.email, sup.phone, sup.address, sup.default_lead_time_days))
            supplier_created = True
        elif supplier_id:
            cursor.execute("SELECT 1 FROM erp_suppliers WHERE supplier_id = ?", (supplier_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=400, detail=f"Structural Violation: supplier {supplier_id} is not registered.")

        cursor.execute("SELECT 1 FROM erp_skus WHERE sku_id = ?", (payload.sku_id,))
        if cursor.fetchone():
            raise HTTPException(status_code=409, detail="SKU already exists.")
        cursor.execute(
            """
            INSERT INTO erp_skus (sku_id, nomenclature, unit_cost, reorder_threshold, quantity_on_hand, supplier_id, min_order_qty)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (payload.sku_id, payload.nomenclature, payload.unit_cost, payload.reorder_threshold, 0, supplier_id,
             max(1, payload.min_order_qty or 1))
        )
        conn.commit()
        if supplier_created:
            logger.info(f"[SUPPLIER] {jwt_payload.get('sub')} atomically registered supplier {supplier_id} with SKU {payload.sku_id}.")
        return {"status": "success", "message": f"SKU {payload.sku_id} ingested.",
                "supplier_id": supplier_id, "supplier_created": supplier_created}
    except sqlite3.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="SKU already exists.")
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database insertion failed: {e}")
    finally:
        conn.close()

@api_router.get("/inventory/skus")
def get_skus(
    limit: int = Query(50, ge=1, le=100),
    offset: int = Query(0, ge=0),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    conn = get_db_connection()
    try:
        # Gate + resolve SKU scope (restricted DM/TECH see only cleared SKUs).
        role, cleared = resolve_procurement_scope(conn, jwt_payload)
        cursor = conn.cursor()

        if cleared is None:
            cursor.execute("SELECT COUNT(*) as count FROM erp_skus")
            total_row = cursor.fetchone()
            total = total_row["count"] if total_row else 0
            cursor.execute("""
                SELECT s.sku_id, s.nomenclature, s.unit_cost, s.reorder_threshold, s.quantity_on_hand,
                       s.supplier_id, sup.name AS supplier_name, s.min_order_qty
                FROM erp_skus s
                LEFT JOIN erp_suppliers sup ON s.supplier_id = sup.supplier_id
                LIMIT ? OFFSET ?
            """, (limit, offset))
        else:
            cleared_list = list(cleared)
            placeholders = ",".join("?" * len(cleared_list))
            cursor.execute(f"SELECT COUNT(*) as count FROM erp_skus WHERE sku_id IN ({placeholders})", cleared_list)
            total_row = cursor.fetchone()
            total = total_row["count"] if total_row else 0
            cursor.execute(f"""
                SELECT s.sku_id, s.nomenclature, s.unit_cost, s.reorder_threshold, s.quantity_on_hand,
                       s.supplier_id, sup.name AS supplier_name, s.min_order_qty
                FROM erp_skus s
                LEFT JOIN erp_suppliers sup ON s.supplier_id = sup.supplier_id
                WHERE s.sku_id IN ({placeholders})
                LIMIT ? OFFSET ?
            """, cleared_list + [limit, offset])

        rows = [dict(row) for row in cursor.fetchall()]
        return {
            "items": rows,
            "total": total,
            "limit": limit,
            "offset": offset
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"SKU Ledger fetch failed: {e}")
        raise HTTPException(status_code=500, detail="Matrix synchronization failed.")
    finally:
        conn.close()

@api_router.post("/inventory/parts", status_code=201)
def ingest_part(payload: PartCreate, token_payload: dict = Depends(verify_jwt_token)):
    part_id = f"PRT-{uuid.uuid4().hex[:8].upper()}"
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN TRANSACTION;")
        
        cursor.execute(
            "INSERT INTO erp_parts (part_id, sku_id, serial_number) VALUES (?, ?, ?)",
            (part_id, payload.sku_id, payload.serial_number)
        )
        
        cursor.execute(
            "UPDATE erp_skus SET quantity_on_hand = quantity_on_hand + 1 WHERE sku_id = ?",
            (payload.sku_id,)
        )
        
        conn.commit()
    except sqlite3.IntegrityError as e:
        conn.rollback()
        error_msg = str(e).lower()
        if "foreign key" in error_msg:
            raise HTTPException(status_code=400, detail=f"Invalid SKU ID: {payload.sku_id} does not exist in the financial catalog.")
        if "unique" in error_msg:
            raise HTTPException(status_code=409, detail="Serial number already exists in the matrix.")
        raise HTTPException(status_code=500, detail="Database integrity violation.")
    finally:
        conn.close()

    return {"status": "success", "message": f"Part {part_id} physically instantiated."}

@api_router.get("/inventory/parts")
def get_paginated_parts(
    limit: int = Query(50, ge=1, le=100),
    offset: int = Query(0, ge=0),
    token_payload: dict = Depends(verify_jwt_token)
):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) as count FROM erp_parts")
        total_row = cursor.fetchone()
        total = total_row["count"] if total_row else 0
        
        cursor.execute("""
            SELECT p.part_id, p.sku_id, s.nomenclature, p.serial_number, p.status 
            FROM erp_parts p
            JOIN erp_skus s ON p.sku_id = s.sku_id
            LIMIT ? OFFSET ?
        """, (limit, offset))
        rows = [dict(row) for row in cursor.fetchall()]
        return {
            "items": rows,
            "total": total,
            "limit": limit,
            "offset": offset
        }
    except Exception as e:
        logger.error(f"Part matrix extraction failed: {e}")
        raise HTTPException(status_code=500, detail="Matrix extraction failed.")
    finally:
        conn.close()

@api_router.post("/inventory/mutate")
def mutate_inventory(payload: InventoryMutation, jwt_payload: dict = Depends(verify_jwt_token)):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN TRANSACTION")
        
        cursor.execute("SELECT quantity_on_hand FROM erp_skus WHERE sku_id = ?", (payload.sku_id,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="SKU not found.")
            
        current_qty = row["quantity_on_hand"]
        if payload.mutation_type == "IN":
            new_qty = current_qty + payload.quantity
        else:
            if current_qty < payload.quantity:
                raise HTTPException(status_code=400, detail="Insufficient quantity_on_hand for OUT mutation.")
            new_qty = current_qty - payload.quantity
            
        cursor.execute("UPDATE erp_skus SET quantity_on_hand = ? WHERE sku_id = ?", (new_qty, payload.sku_id))
        conn.commit()
        return {"status": "success", "quantity_on_hand": new_qty}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Mutation error: {e}")
        raise HTTPException(status_code=500, detail="Internal Execution Error.")
    finally:
        conn.close()

@api_router.get("/users")
def get_paginated_users(
    limit: int = Query(50, ge=1, le=100, description="Strict pagination limit"),
    offset: int = Query(0, ge=0, description="Strict pagination offset"),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        # 1. Execute COUNT() for envelope
        cursor.execute("SELECT COUNT(*) FROM erp_employees")
        total = cursor.fetchone()[0]
        
        # 2. Execute LIMIT/OFFSET extraction
        cursor.execute(
            """
            SELECT id as user_id, name, role, department_id as department, reports_to_hm_id
            FROM erp_employees
            LIMIT ? OFFSET ?
            """,
            (limit, offset)
        )
        rows = [dict(row) for row in cursor.fetchall()]
        
        # 3. Return explicit pagination envelope
        return {
            "items": rows,
            "total": total,
            "limit": limit,
            "offset": offset
        }
    except Exception as e:
        logger.error(f"Failed to fetch users: {e}")
        raise HTTPException(status_code=500, detail="Matrix synchronization failed.")
    finally:
        conn.close()

@api_router.post("/work-orders/{mwo_id}/consume", status_code=201)
def ingest_mwo_consumption(
    mwo_id: str,
    payload: MwoConsumptionPayload,
    token_payload: dict = Depends(verify_jwt_token)
):
    tech_id = token_payload.get("sub")
    if not tech_id:
        raise HTTPException(status_code=401, detail="Unauthorized: Identity Gateway missing sub claim")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        # 2. Enforce Strict Write-Lock to prevent SQLITE_BUSY deadlocks
        cursor.execute("BEGIN IMMEDIATE TRANSACTION;")
        
        # 3. Assert MWO Execution State
        cursor.execute("SELECT status FROM work_orders WHERE mwo_id = ?", (mwo_id,))
        mwo = cursor.fetchone()
        if not mwo or mwo["status"] not in ('ASSIGNED', 'IN_PROGRESS', 'DISPATCHED'):
            raise HTTPException(status_code=400, detail="Invalid or inactive MWO target.")

        import time
        import uuid
        
        # 4. Iterate and Execute Discrete Physical Math
        consumed_ledgers = []
        for part_id in payload.part_ids:
            cursor.execute("SELECT sku_id, status FROM erp_parts WHERE part_id = ?", (part_id,))
            part = cursor.fetchone()
            
            if not part or part["status"] != "IN_STOCK":
                raise HTTPException(status_code=400, detail=f"Part {part_id} is invalid or already consumed.")
                
            consumption_id = f"CNS-{uuid.uuid4().hex[:8].upper()}"
            
            # A. Log Junction Consumption (Strictly QTY 1 for discrete assets)
            cursor.execute(
                """
                INSERT INTO mwo_consumed_parts 
                (consumption_id, mwo_id, part_id, quantity_consumed, consumed_at, logged_by_tech_id) 
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (consumption_id, mwo_id, part_id, 1, time.time(), tech_id)
            )
            
            # B. Mutate Physical Asset Status
            cursor.execute("UPDATE erp_parts SET status = 'CONSUMED' WHERE part_id = ?", (part_id,))
            
            # C. Deduct Exactly 1 from the SKU Ledger
            cursor.execute(
                "UPDATE erp_skus SET quantity_on_hand = quantity_on_hand - 1 WHERE sku_id = ?",
                (part["sku_id"],)
            )
            consumed_ledgers.append(part_id)

        conn.commit()
        return {
            "status": "success",
            "message": f"{len(consumed_ledgers)} discrete parts mathematically consumed against MWO {mwo_id}."
        }
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        print(f"Consumption Execution Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@api_router.get("/admin/procurement", response_model=Dict[str, Any])
def get_procurement_ledger(
    limit: int = Query(50, ge=1, le=100),
    offset: int = Query(0, ge=0),
    jwt_payload: dict = Depends(verify_jwt_token)
):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Clearance required to view procurement ledger.")
        
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        # Phase 1: Total Count Extraction
        cursor.execute("SELECT COUNT(procurement_id) FROM erp_procurement_queue")
        total_records = cursor.fetchone()[0]

        # Phase 2: Explicit Column Pagination Extraction
        query = """
            SELECT 
                p.procurement_id,
                p.part_id,
                s.nomenclature,
                s.quantity_on_hand,
                s.reorder_threshold,
                s.unit_cost,
                p.status,
                p.triggered_at
            FROM 
                erp_procurement_queue p
            JOIN 
                erp_skus s ON p.part_id = s.sku_id
            ORDER BY 
                p.triggered_at DESC
            LIMIT ? OFFSET ?
        """
        cursor.execute(query, (limit, offset))
        rows = [dict(row) for row in cursor.fetchall()]

        # Return strict Unified I/O Serialization Envelope
        return {
            "items": rows,
            "total": total_records,
            "limit": limit,
            "offset": offset
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database execution anomaly: {str(e)}")
    finally:
        conn.close()


# ==========================================
# PHASE 48: MASTER DATA HYDRATION ROUTES
# ==========================================
# Pydantic Models for deterministic Master Data ingestion.
# All models accept an optional client-side `id` to support
# scripted hydration with deterministic primary keys.

class LocationCreate(BaseModel):
    id: Optional[str] = None
    name: str = Field(..., min_length=2, description="Physical location name")

class DepartmentCreate(BaseModel):
    id: Optional[str] = None
    name: str = Field(..., min_length=2, description="Operational department name")

class CategoryCreate(BaseModel):
    id: Optional[str] = None
    name: str = Field(..., min_length=2, description="Equipment/SKU category name")

class Phase48EmployeeCreate(BaseModel):
    id: Optional[str] = None
    name: str = Field(..., min_length=2)
    role: str = Field(..., description="ADMINISTRATOR, ADMIN, HM, TECH")
    pin_code: str = Field(..., min_length=4)
    department_id: str = Field(..., description="FK to erp_departments.id")
    reports_to_hm_id: Optional[str] = None

class Phase48SkuCreate(BaseModel):
    id: Optional[str] = None
    nomenclature: str = Field(..., min_length=2)
    category_id: str = Field(..., description="FK to erp_categories.id")
    unit_cost: float = Field(..., ge=0.0)
    reorder_threshold: int = Field(..., ge=0)

class Phase48EquipmentCreate(BaseModel):
    id: Optional[str] = None
    nomenclature: str = Field(..., min_length=2)
    category_id: str = Field(..., description="FK to erp_categories.id")
    department_id: str = Field(..., description="FK to erp_departments.id")
    location_id: str = Field(..., description="FK to erp_locations.id")
    assigned_hm_id: str = Field(..., description="FK to erp_employees.id")
    status: str = Field(default="ACTIVE")


# --- LEVEL 0 ROUTES ---

@api_router.post("/locations", status_code=201)
def hydrate_location(payload: LocationCreate, jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Administrative clearance required.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        loc_id = payload.id if payload.id else f"LOC-{uuid.uuid4().hex[:6].upper()}"
        cursor.execute("INSERT INTO erp_locations (id, name) VALUES (?, ?)", (loc_id, payload.name))
        conn.commit()
        return {"detail": "Record successfully ingested."}
    except sqlite3.IntegrityError as e:
        conn.rollback()
        logger.error(f"Location IntegrityError: {e}")
        raise HTTPException(status_code=400, detail=f"Structural Violation: {e}")
    except Exception as e:
        conn.rollback()
        logger.error(f"Location Ingestion Error: {e}")
        raise HTTPException(status_code=500, detail=f"Internal Execution Error: {e}")
    finally:
        conn.close()

@api_router.post("/departments", status_code=201)
def hydrate_department(payload: DepartmentCreate, jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Administrative clearance required.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        dep_id = payload.id if payload.id else f"DEP-{uuid.uuid4().hex[:6].upper()}"
        cursor.execute("INSERT INTO erp_departments (id, name) VALUES (?, ?)", (dep_id, payload.name))
        conn.commit()
        return {"detail": "Record successfully ingested."}
    except sqlite3.IntegrityError as e:
        conn.rollback()
        logger.error(f"Department IntegrityError: {e}")
        raise HTTPException(status_code=400, detail=f"Structural Violation: {e}")
    except Exception as e:
        conn.rollback()
        logger.error(f"Department Ingestion Error: {e}")
        raise HTTPException(status_code=500, detail=f"Internal Execution Error: {e}")
    finally:
        conn.close()

@api_router.post("/categories", status_code=201)
def hydrate_category(payload: CategoryCreate, jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Administrative clearance required.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        cat_id = payload.id if payload.id else f"CAT-{uuid.uuid4().hex[:6].upper()}"
        cursor.execute("INSERT INTO erp_categories (id, name) VALUES (?, ?)", (cat_id, payload.name))
        conn.commit()
        return {"detail": "Record successfully ingested."}
    except sqlite3.IntegrityError as e:
        conn.rollback()
        logger.error(f"Category IntegrityError: {e}")
        raise HTTPException(status_code=400, detail=f"Structural Violation: {e}")
    except Exception as e:
        conn.rollback()
        logger.error(f"Category Ingestion Error: {e}")
        raise HTTPException(status_code=500, detail=f"Internal Execution Error: {e}")
    finally:
        conn.close()


class CategoryManagerPayload(BaseModel):
    manager_id: Optional[str] = None


@api_router.put("/admin/categories/{category_id}/manager")
def update_category_manager(category_id: str, payload: CategoryManagerPayload, jwt_payload: dict = Depends(verify_jwt_token)):
    """Admin-only: assign / change / clear the manager of an inventory category."""
    role = jwt_payload.get("role")
    if role not in SKU_ACCESS_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: Admin clearance required.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        cursor.execute("SELECT 1 FROM erp_categories WHERE id = ?", (category_id,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail=f"Category {category_id} not found.")
        if payload.manager_id:
            cursor.execute("SELECT 1 FROM erp_employees WHERE id = ?", (payload.manager_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=400, detail=f"Employee {payload.manager_id} not found.")
        cursor.execute("UPDATE erp_categories SET manager_id = ? WHERE id = ?", (payload.manager_id, category_id))
        conn.commit()
        return {"status": "success", "message": f"Category {category_id} manager updated to {payload.manager_id}."}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Category manager update error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# --- LEVEL 1 ROUTES ---

@api_router.post("/employees", status_code=201)
def hydrate_employee(payload: Phase48EmployeeCreate, jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Administrative clearance required.")
    conn = get_db_connection()
    try:
        import bcrypt
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        emp_id = payload.id if payload.id else f"U-{uuid.uuid4().hex[:6].upper()}"
        pin_hash = bcrypt.hashpw(payload.pin_code.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        cursor.execute(
            "INSERT INTO erp_employees (id, name, role, pin_hash, is_active, department_id, reports_to_hm_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (emp_id, payload.name, payload.role, pin_hash, 1, payload.department_id, payload.reports_to_hm_id)
        )
        conn.commit()
        return {"detail": "Record successfully ingested."}
    except sqlite3.IntegrityError as e:
        conn.rollback()
        logger.error(f"Employee IntegrityError: {e}")
        raise HTTPException(status_code=400, detail=f"Structural Violation: {e}")
    except Exception as e:
        conn.rollback()
        logger.error(f"Employee Ingestion Error: {e}")
        raise HTTPException(status_code=500, detail=f"Internal Execution Error: {e}")
    finally:
        conn.close()

@api_router.post("/skus", status_code=201)
def hydrate_sku(payload: Phase48SkuCreate, jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Administrative clearance required.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        sku_id = payload.id if payload.id else f"SKU-{uuid.uuid4().hex[:6].upper()}"
        cursor.execute(
            "INSERT INTO erp_skus (sku_id, nomenclature, unit_cost, reorder_threshold, quantity_on_hand) VALUES (?, ?, ?, ?, ?)",
            (sku_id, payload.nomenclature, payload.unit_cost, payload.reorder_threshold, 0)
        )
        conn.commit()
        return {"detail": "Record successfully ingested."}
    except sqlite3.IntegrityError as e:
        conn.rollback()
        logger.error(f"SKU IntegrityError: {e}")
        raise HTTPException(status_code=400, detail=f"Structural Violation: {e}")
    except Exception as e:
        conn.rollback()
        logger.error(f"SKU Ingestion Error: {e}")
        raise HTTPException(status_code=500, detail=f"Internal Execution Error: {e}")
    finally:
        conn.close()


# --- LEVEL 2 ROUTES ---

@api_router.post("/equipment", status_code=201)
def hydrate_equipment(payload: Phase48EquipmentCreate, jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    if role not in ["ADMINISTRATOR", "ADMIN"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: Administrative clearance required.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        eq_id = payload.id if payload.id else f"EQ-{uuid.uuid4().hex[:6].upper()}"
        cursor.execute(
            "INSERT INTO erp_equipment (equipment_id, nomenclature, category_id, status, department_id, location_id, assigned_hm_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (eq_id, payload.nomenclature, payload.category_id, payload.status, payload.department_id, payload.location_id, payload.assigned_hm_id)
        )
        conn.commit()
        return {"detail": "Record successfully ingested."}
    except sqlite3.IntegrityError as e:
        conn.rollback()
        logger.error(f"Equipment IntegrityError: {e}")
        raise HTTPException(status_code=400, detail=f"Structural Violation: {e}")
    except Exception as e:
        conn.rollback()
        logger.error(f"Equipment Ingestion Error: {e}")
        raise HTTPException(status_code=500, detail=f"Internal Execution Error: {e}")
    finally:
        conn.close()

@api_router.get("/dm/personnel")
def get_department_personnel(jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    user_id = jwt_payload.get("sub")

    if role not in ["DM", "ADMIN", "ADMINISTRATOR"]:
        raise HTTPException(status_code=403, detail="RBAC Violation: DM clearance required.")

    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        # Global clearance: administrators see the full active roster
        department_id = None
        if role not in ["ADMIN", "ADMINISTRATOR"]:
            cursor.execute("SELECT department_id FROM erp_employees WHERE id = ?", (user_id,))
            user_record = cursor.fetchone()
            if not user_record:
                raise HTTPException(status_code=404, detail="DM record not found.")
            department_id = user_record['department_id']

        if department_id is None:
            # Graceful fallback: DMs without a department assignment (and admins)
            # receive all active personnel instead of an invalid `= NULL` filter.
            cursor.execute("SELECT id, name, role FROM erp_employees WHERE is_active = 1")
        else:
            cursor.execute("SELECT id, name, role FROM erp_employees WHERE department_id = ? AND is_active = 1", (department_id,))
        personnel = [dict(row) for row in cursor.fetchall()]

        return {"status": "success", "data": personnel}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to fetch department personnel: {e}")
        raise HTTPException(status_code=500, detail="Internal server error.")
    finally:
        conn.close()

# =====================================================================
# [BACK OFFICE INVENTORY MODULE] Suppliers / Purchase Orders / Manual Logs
# =====================================================================
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from collections import deque
from typing import List

# --- Inventory Event Bus (SSE) ---
# Bounded in-memory ring buffer; SSE clients poll for events newer than their cursor.
_inventory_events = deque(maxlen=200)
_inventory_event_seq = 0
_inventory_event_lock = __import__("threading").Lock()

def emit_inventory_event(event_type: str, payload: dict):
    global _inventory_event_seq
    with _inventory_event_lock:
        _inventory_event_seq += 1
        _inventory_events.append({
            "id": _inventory_event_seq,
            "type": event_type,
            "payload": payload,
            "ts": datetime.now(timezone.utc).isoformat()
        })

@api_router.get("/notifications/stream")
async def inventory_notification_stream(token: str = Query(...)):
    """SSE feed for inventory/PO events. EventSource cannot set headers, so the JWT rides a query param."""
    if not PUBLIC_KEY:
        raise HTTPException(status_code=503, detail="Cryptographic Gateway Unavailable.")
    try:
        payload = jwt.decode(token, PUBLIC_KEY, algorithms=[ALGORITHM])
        if is_jti_revoked(payload.get("jti")):
            raise HTTPException(status_code=401, detail="Token Revoked.")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=403, detail="Cryptographic Verification Failed.")

    async def event_generator():
        cursor = _inventory_event_seq
        yield "event: connected\ndata: {}\n\n"
        while True:
            with _inventory_event_lock:
                fresh = [e for e in _inventory_events if e["id"] > cursor]
                if fresh:
                    cursor = fresh[-1]["id"]
            for e in fresh:
                yield f"id: {e['id']}\nevent: {e['type']}\ndata: {json.dumps(e)}\n\n"
            await asyncio.sleep(2)

    return StreamingResponse(event_generator(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

# --- Data Models ---

class ManualLogPayload(BaseModel):
    sku_id: str
    direction: str = Field(..., description="IN (stock received) or OUT (write-off / consumption)")
    quantity: int = Field(..., gt=0)
    comment: Optional[str] = Field(None, max_length=500)

    @field_validator('direction', mode='before')
    @classmethod
    def validate_direction(cls, v):
        v = str(v).upper()
        if v not in {"IN", "OUT"}:
            raise ValueError("Structural Violation: direction must be IN or OUT")
        return v

class POItemQuantity(BaseModel):
    sku_id: str
    quantity: int = Field(..., gt=0)

class DraftPOUpdatePayload(BaseModel):
    items: Optional[List[POItemQuantity]] = None
    notes: Optional[str] = Field(None, max_length=1000)
    priority: Optional[int] = Field(None, ge=0, le=1)
    eta_date: Optional[str] = None

class POSubmitPayload(BaseModel):
    po_id: str

class AddDraftItemPayload(BaseModel):
    sku_id: str
    quantity: int = Field(..., gt=0)
    notes: Optional[str] = Field(None, max_length=1000)

class SupplierUpdate(BaseModel):
    name: Optional[str] = Field(None, min_length=2, max_length=200)
    email: Optional[str] = Field(None, max_length=254)
    phone: Optional[str] = Field(None, max_length=50)
    address: Optional[str] = Field(None, max_length=500)
    default_lead_time_days: Optional[int] = Field(None, ge=0, le=365)

    @field_validator('email', mode='before')
    @classmethod
    def validate_email(cls, v):
        if v is None:
            return v
        import re
        v = str(v).strip()
        if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', v):
            raise ValueError("Structural Violation: email must be a valid address (user@domain.tld)")
        return v

    @field_validator('phone', 'address', mode='before')
    @classmethod
    def blank_to_none(cls, v):
        if isinstance(v, str) and not v.strip():
            return None
        return v

class SkuSupplierAssign(BaseModel):
    supplier_id: Optional[str] = None  # null explicitly unassigns

    @field_validator('supplier_id', mode='before')
    @classmethod
    def blank_to_none(cls, v):
        if isinstance(v, str) and not v.strip():
            return None
        return v

class BulkActuationPayload(BaseModel):
    po_ids: List[str] = Field(..., min_length=1)
    action: str = Field(..., description="APPROVE / HOLD / REJECT")
    cfo_notes: Optional[str] = None

    @field_validator('action', mode='before')
    @classmethod
    def validate_action(cls, v):
        v = str(v).upper()
        if v not in {"APPROVE", "HOLD", "REJECT"}:
            raise ValueError("Structural Violation: action must be APPROVE, HOLD or REJECT")
        return v

INVENTORY_OPERATOR_ROLES = ["ADMINISTRATOR", "ADMIN", "HM", "DM", "TECH"]
HOD_ROLES = ["ADMINISTRATOR", "ADMIN", "HM"]
CFO_ROLES = ["CFO", "ADMINISTRATOR", "ADMIN"]

# [RESTRICTED SKU PROCUREMENT] Admin management endpoints clearance.
SKU_ACCESS_ADMIN_ROLES = ["ADMINISTRATOR", "ADMIN"]
# Roles with unrestricted procurement visibility (no SKU clearance gating).
PROCUREMENT_GLOBAL_ROLES = ["ADMINISTRATOR", "ADMIN", "CFO", "HM"]
# Roles permitted into procurement only when they hold >=1 SKU clearance row.
PROCUREMENT_RESTRICTED_ROLES = ["DM", "TECH"]


def get_employee_cleared_skus(conn, employee_id: str) -> set:
    """
    Return the union of an employee's explicit SKU clearances and every SKU in
    the categories they manage. Managing a category grants access to all of its
    SKUs (OR with explicit clearances).
    """
    rows = conn.execute(
        "SELECT sku_id FROM erp_employee_sku_access WHERE employee_id = ?",
        (employee_id,)
    ).fetchall()
    cleared = {r["sku_id"] for r in rows}

    managed_cats = [
        r["id"] for r in conn.execute(
            "SELECT id FROM erp_categories WHERE manager_id = ?", (employee_id,)
        ).fetchall()
    ]
    if managed_cats:
        placeholders = ",".join("?" * len(managed_cats))
        for r in conn.execute(
            f"SELECT sku_id FROM erp_skus WHERE category_id IN ({placeholders})",
            managed_cats
        ).fetchall():
            cleared.add(r["sku_id"])

    return cleared


def verify_employee_sku_access(conn, employee_id: str, sku_id: str, role: str) -> bool:
    """
    True if the role has unrestricted procurement clearance, the employee
    manages the SKU's category, or the employee holds an explicit clearance
    row for this SKU.
    """
    if role in PROCUREMENT_GLOBAL_ROLES:
        return True
    row = conn.execute(
        """
        SELECT 1
        FROM erp_skus s
        JOIN erp_categories c ON s.category_id = c.id
        WHERE s.sku_id = ? AND c.manager_id = ?
        """,
        (sku_id, employee_id)
    ).fetchone()
    if row is not None:
        return True
    row = conn.execute(
        "SELECT 1 FROM erp_employee_sku_access WHERE employee_id = ? AND sku_id = ?",
        (employee_id, sku_id)
    ).fetchone()
    return row is not None


def resolve_procurement_scope(conn, jwt_payload: dict):
    """
    Gate procurement access and resolve the caller's SKU scope.

    Returns (role, cleared_skus) where cleared_skus is None for unrestricted
    (global) roles or a set of authorized sku_ids for restricted (DM/TECH)
    users. Raises 403 for unauthorized roles or restricted users with no
    clearances.
    """
    role = jwt_payload.get("role")
    if role in PROCUREMENT_GLOBAL_ROLES:
        return role, None
    if role in PROCUREMENT_RESTRICTED_ROLES:
        cleared = get_employee_cleared_skus(conn, jwt_payload.get("sub"))
        if not cleared:
            raise HTTPException(status_code=403, detail="RBAC Violation: No procurement SKU clearances assigned.")
        return role, cleared
    raise HTTPException(status_code=403, detail="RBAC Violation: Procurement clearance required.")


def resolve_draft_mutation_scope(conn, jwt_payload: dict):
    """
    Gate draft-PO mutations (save/exclude). Preserves the original HOD-only
    global set — CFO is intentionally excluded from editing drafts — while
    admitting restricted DM/TECH that hold >=1 SKU clearance. Returns
    (role, cleared) where cleared is None for HOD and a set for DM/TECH.
    """
    role = jwt_payload.get("role")
    if role in HOD_ROLES:
        return role, None
    if role in PROCUREMENT_RESTRICTED_ROLES:
        cleared = get_employee_cleared_skus(conn, jwt_payload.get("sub"))
        if not cleared:
            raise HTTPException(status_code=403, detail="RBAC Violation: No procurement SKU clearances assigned.")
        return role, cleared
    raise HTTPException(status_code=403, detail="RBAC Violation: HOD or cleared procurement clearance required.")


def redact_pos_for_scope(pos: list, cleared_skus) -> list:
    """
    Restricted-user PO redaction: keep only POs containing >=1 cleared SKU,
    strip foreign line items, and recompute total_cost from the visible items.
    Unrestricted callers (cleared_skus is None) see the list untouched.
    """
    if cleared_skus is None:
        return pos
    visible = []
    for po in pos:
        items = [i for i in po["items"] if i["sku_id"] in cleared_skus]
        if not items:
            continue
        record = dict(po)
        record["items"] = items
        record["total_cost"] = round(sum(i["quantity"] * i["unit_cost"] for i in items), 2)
        visible.append(record)
    return visible

# --- Threshold Evaluation Worker (Auto-Drafting) ---

def worker_evaluate_sku_threshold(sku_id: str):
    """
    Threshold breach evaluator. When quantity_on_hand <= reorder_threshold,
    appends the SKU to the supplier's open DRAFT purchase order, or synthesizes
    a new draft when none exists. Executes off the HTTP response cycle.
    """
    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("""
            SELECT s.sku_id, s.nomenclature, s.quantity_on_hand, s.reorder_threshold,
                   s.unit_cost, s.supplier_id, s.min_order_qty, sup.default_lead_time_days
            FROM erp_skus s
            LEFT JOIN erp_suppliers sup ON s.supplier_id = sup.supplier_id
            WHERE s.sku_id = ?
        """, (sku_id,))
        sku = c.fetchone()
        if not sku:
            logger.error(f"[PO WORKER] sku_id {sku_id} not found. Terminating evaluation.")
            return
        if sku["quantity_on_hand"] > sku["reorder_threshold"]:
            return
        if not sku["supplier_id"]:
            logger.warning(f"[PO WORKER] {sku_id} breached threshold but has no designated supplier. Skipping auto-draft.")
            return

        # Reorder heuristic: restore to 2x threshold coverage,
        # rounded UP to the SKU's Minimum Order Quantity (MOQ)
        reorder_qty = max((sku["reorder_threshold"] * 2) - sku["quantity_on_hand"], 1)
        moq = sku["min_order_qty"] if sku["min_order_qty"] else 1
        reorder_qty = max(reorder_qty, moq)

        c.execute("BEGIN IMMEDIATE TRANSACTION")
        c.execute("SELECT po_id FROM erp_purchase_orders WHERE supplier_id = ? AND status = 'DRAFT'", (sku["supplier_id"],))
        draft = c.fetchone()

        if draft:
            po_id = draft["po_id"]
            c.execute("""
                INSERT INTO erp_purchase_order_items (po_id, sku_id, quantity, unit_cost)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(po_id, sku_id) DO UPDATE SET quantity = excluded.quantity
            """, (po_id, sku_id, reorder_qty, sku["unit_cost"]))
            conn.commit()
            emit_inventory_event("po_draft_updated", {"po_id": po_id, "sku_id": sku_id, "quantity": reorder_qty})
            logger.warning(f"[PO WORKER] Low stock {sku_id}: appended to existing draft {po_id} (qty {reorder_qty}).")
        else:
            po_id = f"PO-{uuid.uuid4().hex[:8].upper()}"
            lead_days = sku["default_lead_time_days"] if sku["default_lead_time_days"] is not None else 7
            eta = (datetime.now(timezone.utc) + timedelta(days=lead_days)).date().isoformat()
            c.execute("""
                INSERT INTO erp_purchase_orders (po_id, supplier_id, status, priority, eta_date)
                VALUES (?, ?, 'DRAFT', 0, ?)
            """, (po_id, sku["supplier_id"], eta))
            c.execute("""
                INSERT INTO erp_purchase_order_items (po_id, sku_id, quantity, unit_cost)
                VALUES (?, ?, ?, ?)
            """, (po_id, sku_id, reorder_qty, sku["unit_cost"]))
            conn.commit()
            emit_inventory_event("po_draft_created", {"po_id": po_id, "supplier_id": sku["supplier_id"], "sku_id": sku_id, "quantity": reorder_qty})
            logger.warning(f"[PO WORKER] Low stock {sku_id}: new draft {po_id} synthesized for {sku['supplier_id']}.")
    except sqlite3.IntegrityError as e:
        # idx_po_one_draft_per_supplier intercepted a concurrent draft creation race
        conn.rollback()
        logger.info(f"[PO WORKER] Concurrency lock engaged for {sku_id}: {e}")
    except Exception as e:
        conn.rollback()
        logger.error(f"[PO WORKER] Fatal evaluation error for {sku_id}: {e}")
    finally:
        conn.close()

# --- Manual Log Ingestion ---

@api_router.post("/inventory/manual-log", status_code=201)
def ingest_manual_log(payload: ManualLogPayload, background_tasks: BackgroundTasks, jwt_payload: dict = Depends(verify_jwt_token)):
    role = jwt_payload.get("role")
    operator_id = jwt_payload.get("sub")
    if role not in INVENTORY_OPERATOR_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: Inventory operator clearance required.")

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        cursor.execute("SELECT quantity_on_hand FROM erp_skus WHERE sku_id = ?", (payload.sku_id,))
        sku = cursor.fetchone()
        if not sku:
            raise HTTPException(status_code=404, detail=f"SKU {payload.sku_id} not found in master catalog.")

        if payload.direction == "OUT":
            if sku["quantity_on_hand"] < payload.quantity:
                raise HTTPException(status_code=400, detail=f"Structural Violation: Stock-Out of {payload.quantity} exceeds on-hand quantity ({sku['quantity_on_hand']}).")
            cursor.execute("UPDATE erp_skus SET quantity_on_hand = quantity_on_hand - ? WHERE sku_id = ?", (payload.quantity, payload.sku_id))
        else:
            cursor.execute("UPDATE erp_skus SET quantity_on_hand = quantity_on_hand + ? WHERE sku_id = ?", (payload.quantity, payload.sku_id))

        cursor.execute("""
            INSERT INTO erp_inventory_manual_logs (sku_id, direction, quantity, comment, logged_by)
            VALUES (?, ?, ?, ?, ?)
        """, (payload.sku_id, payload.direction, payload.quantity, payload.comment, operator_id))
        cursor.execute("SELECT quantity_on_hand FROM erp_skus WHERE sku_id = ?", (payload.sku_id,))
        new_qty = cursor.fetchone()["quantity_on_hand"]
        conn.commit()

        # Threshold evaluation strictly decoupled from the response cycle
        if payload.direction == "OUT":
            background_tasks.add_task(worker_evaluate_sku_threshold, payload.sku_id)

        return {"status": "success", "sku_id": payload.sku_id, "direction": payload.direction,
                "quantity": payload.quantity, "new_quantity_on_hand": new_qty, "logged_by": operator_id}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Manual log ingestion error: {e}")
        raise HTTPException(status_code=500, detail="Manual log ingestion failed.")
    finally:
        conn.close()

@api_router.get("/inventory/alerts")
def get_inventory_alerts(jwt_payload: dict = Depends(verify_jwt_token)):
    """
    Safety-stock alert feed for the HM dashboard: every SKU at or below its
    reorder threshold, with the open DRAFT PO (if any) that already covers it
    so the UI can deep-link straight to the draft card.
    """
    conn = get_db_connection()
    try:
        # Gate + scope: restricted DM/TECH see alerts only for their cleared SKUs.
        role, cleared = resolve_procurement_scope(conn, jwt_payload)
        scope_clause = ""
        params = []
        if cleared is not None:
            cleared_list = list(cleared)
            placeholders = ",".join("?" * len(cleared_list))
            scope_clause = f" AND s.sku_id IN ({placeholders})"
            params = cleared_list
        rows = conn.execute(f"""
            SELECT s.sku_id, s.nomenclature, s.quantity_on_hand, s.reorder_threshold,
                   s.min_order_qty, s.supplier_id,
                   (SELECT i.po_id
                    FROM erp_purchase_order_items i
                    JOIN erp_purchase_orders po ON po.po_id = i.po_id AND po.status = 'DRAFT'
                    WHERE i.sku_id = s.sku_id
                    LIMIT 1) AS draft_po_id,
                   (SELECT i.po_id
                    FROM erp_purchase_order_items i
                    JOIN erp_purchase_orders po ON po.po_id = i.po_id AND po.status IN ('PENDING_CFO', 'APPROVED', 'HOLD')
                    WHERE i.sku_id = s.sku_id ORDER BY po.created_at DESC LIMIT 1) AS active_po_id,
                   (SELECT po.status
                    FROM erp_purchase_order_items i
                    JOIN erp_purchase_orders po ON po.po_id = i.po_id AND po.status IN ('PENDING_CFO', 'APPROVED', 'HOLD')
                    WHERE i.sku_id = s.sku_id ORDER BY po.created_at DESC LIMIT 1) AS active_po_status
            FROM erp_skus s
            WHERE s.quantity_on_hand <= s.reorder_threshold{scope_clause}
            ORDER BY (s.quantity_on_hand * 1.0) / NULLIF(s.reorder_threshold, 0) ASC, s.sku_id
        """, params).fetchall()
        return {"status": "success", "data": [dict(r) for r in rows]}
    finally:
        conn.close()

@api_router.get("/inventory/skus/search")
def search_skus(q: str = Query("", max_length=100), limit: int = Query(20, ge=1, le=100), jwt_payload: dict = Depends(verify_jwt_token)):
    """SKU autocomplete feed for the manual log widget (maintenance parts + admin supplies)."""
    if jwt_payload.get("role") not in INVENTORY_OPERATOR_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: Inventory operator clearance required.")
    conn = get_db_connection()
    try:
        rows = conn.execute("""
            SELECT s.sku_id, s.nomenclature, s.quantity_on_hand, s.reorder_threshold, s.unit_cost,
                   s.category_id, c.name AS category_name, s.supplier_id, sup.name AS supplier_name
            FROM erp_skus s
            LEFT JOIN erp_categories c ON s.category_id = c.id
            LEFT JOIN erp_suppliers sup ON s.supplier_id = sup.supplier_id
            WHERE s.sku_id LIKE ? OR s.nomenclature LIKE ?
            ORDER BY s.nomenclature LIMIT ?
        """, (f"%{q}%", f"%{q}%", limit)).fetchall()
        return {"status": "success", "data": [dict(r) for r in rows]}
    finally:
        conn.close()

@api_router.get("/inventory/manual-log")
def get_manual_logs(limit: int = Query(50, ge=1, le=200), offset: int = Query(0, ge=0), jwt_payload: dict = Depends(verify_jwt_token)):
    """Zero-trust audit trail of manual stock adjustments."""
    role = jwt_payload.get("role")
    if role not in HOD_ROLES + CFO_ROLES + PROCUREMENT_RESTRICTED_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: HOD clearance required.")
    conn = get_db_connection()
    try:
        # Restricted DM/TECH see only the adjustments they themselves logged;
        # global roles retain full audit visibility.
        own_clause = ""
        params = []
        if role in PROCUREMENT_RESTRICTED_ROLES:
            own_clause = " WHERE l.logged_by = ?"
            params.append(jwt_payload.get("sub"))
        params.extend([limit, offset])
        rows = conn.execute(f"""
            SELECT l.log_id, l.sku_id, s.nomenclature, l.direction, l.quantity, l.comment, l.logged_by, l.logged_at
            FROM erp_inventory_manual_logs l
            LEFT JOIN erp_skus s ON l.sku_id = s.sku_id{own_clause}
            ORDER BY l.log_id DESC LIMIT ? OFFSET ?
        """, params).fetchall()
        return {"status": "success", "data": [dict(r) for r in rows]}
    finally:
        conn.close()

# --- Purchase Order Hydration Helper ---

def _hydrate_purchase_orders(conn, statuses: list):
    placeholders = ",".join("?" * len(statuses))
    pos = conn.execute(f"""
        SELECT po.po_id, po.supplier_id, sup.name AS supplier_name, sup.email AS supplier_email,
               sup.phone AS supplier_phone, sup.address AS supplier_address,
               sup.default_lead_time_days, po.status, po.priority, po.eta_date, po.notes,
               po.cfo_notes, po.created_at, po.submitted_at, po.decided_at
        FROM erp_purchase_orders po
        LEFT JOIN erp_suppliers sup ON po.supplier_id = sup.supplier_id
        WHERE po.status IN ({placeholders})
        ORDER BY po.priority DESC, po.created_at ASC
    """, statuses).fetchall()
    results = []
    for po in pos:
        record = dict(po)
        items = conn.execute("""
            SELECT i.sku_id, s.nomenclature, i.quantity, i.unit_cost,
                   s.quantity_on_hand, s.reorder_threshold, s.min_order_qty
            FROM erp_purchase_order_items i
            LEFT JOIN erp_skus s ON i.sku_id = s.sku_id
            WHERE i.po_id = ? ORDER BY i.sku_id
        """, (po["po_id"],)).fetchall()
        record["items"] = [dict(i) for i in items]
        record["total_cost"] = round(sum(i["quantity"] * i["unit_cost"] for i in record["items"]), 2)
        results.append(record)
    return results

# --- HOD Draft PO Endpoints ---

@api_router.get("/orders/drafts")
def get_draft_orders(jwt_payload: dict = Depends(verify_jwt_token)):
    conn = get_db_connection()
    try:
        role, cleared = resolve_procurement_scope(conn, jwt_payload)
        pos = _hydrate_purchase_orders(conn, ["DRAFT"])
        return {"status": "success", "data": redact_pos_for_scope(pos, cleared)}
    finally:
        conn.close()

@api_router.get("/orders/inbound")
def get_inbound_orders(jwt_payload: dict = Depends(verify_jwt_token)):
    """APPROVED POs awaiting physical receipt, plus PENDING_CFO/HOLD visibility for the HOD."""
    conn = get_db_connection()
    try:
        role, cleared = resolve_procurement_scope(conn, jwt_payload)
        pos = _hydrate_purchase_orders(conn, ["APPROVED", "PENDING_CFO", "HOLD"])
        return {"status": "success", "data": redact_pos_for_scope(pos, cleared)}
    finally:
        conn.close()

# --- Supplier Directory Endpoints ---

@api_router.get("/inventory/suppliers")
def list_suppliers(jwt_payload: dict = Depends(verify_jwt_token)):
    if jwt_payload.get("role") not in INVENTORY_OPERATOR_ROLES + CFO_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: Inventory operator clearance required.")
    conn = get_db_connection()
    try:
        rows = conn.execute("""
            SELECT supplier_id, name, email, phone, address, default_lead_time_days, created_at
            FROM erp_suppliers ORDER BY name
        """).fetchall()
        return {"status": "success", "data": [dict(r) for r in rows]}
    finally:
        conn.close()

@api_router.post("/inventory/suppliers", status_code=201)
def register_supplier(payload: SupplierCreate, jwt_payload: dict = Depends(verify_jwt_token)):
    if jwt_payload.get("role") not in HOD_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: HOD clearance required to register suppliers.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO erp_suppliers (supplier_id, name, email, phone, address, default_lead_time_days)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (payload.supplier_id, payload.name, payload.email, payload.phone, payload.address, payload.default_lead_time_days))
        conn.commit()
        logger.info(f"[SUPPLIER] {jwt_payload.get('sub')} registered supplier {payload.supplier_id} ({payload.name}).")
        return {"status": "success", "supplier_id": payload.supplier_id,
                "detail": f"Supplier {payload.name} registered."}
    except sqlite3.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail=f"Supplier {payload.supplier_id} already exists.")
    except Exception as e:
        conn.rollback()
        logger.error(f"Supplier registration error: {e}")
        raise HTTPException(status_code=500, detail="Supplier registration failed.")
    finally:
        conn.close()

@api_router.put("/inventory/suppliers/{supplier_id}")
def update_supplier(supplier_id: str, payload: SupplierUpdate, jwt_payload: dict = Depends(verify_jwt_token)):
    """HOD supplier lifecycle: amend contact details / lead time. Compulsory fields cannot be blanked."""
    if jwt_payload.get("role") not in HOD_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: HOD clearance required.")
    mutations = {k: v for k, v in payload.model_dump(exclude_unset=True).items()}
    if not mutations:
        raise HTTPException(status_code=400, detail="Structural Violation: no fields supplied for update.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        cursor.execute("SELECT 1 FROM erp_suppliers WHERE supplier_id = ?", (supplier_id,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail=f"Supplier {supplier_id} not found.")
        # Compulsory attributes may be updated but never nulled
        for compulsory in ("name", "email"):
            if compulsory in mutations and mutations[compulsory] is None:
                raise HTTPException(status_code=422, detail=f"Structural Violation: {compulsory} is compulsory and cannot be removed.")
        set_clause = ", ".join(f"{col} = ?" for col in mutations)
        cursor.execute(f"UPDATE erp_suppliers SET {set_clause} WHERE supplier_id = ?",
                       (*mutations.values(), supplier_id))
        conn.commit()
        logger.info(f"[SUPPLIER] {jwt_payload.get('sub')} updated {supplier_id}: {sorted(mutations)}")
        return {"status": "success", "supplier_id": supplier_id, "updated_fields": sorted(mutations)}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Supplier update error: {e}")
        raise HTTPException(status_code=500, detail="Supplier update failed.")
    finally:
        conn.close()

@api_router.put("/inventory/skus/{sku_id}/supplier")
def assign_sku_supplier(sku_id: str, payload: SkuSupplierAssign, background_tasks: BackgroundTasks, jwt_payload: dict = Depends(verify_jwt_token)):
    """HOD supplier reassignment on an existing SKU; supplier_id=null explicitly unassigns."""
    if jwt_payload.get("role") not in HOD_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: HOD clearance required.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        cursor.execute("SELECT 1 FROM erp_skus WHERE sku_id = ?", (sku_id,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail=f"SKU {sku_id} not found.")
        if payload.supplier_id:
            cursor.execute("SELECT 1 FROM erp_suppliers WHERE supplier_id = ?", (payload.supplier_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=400, detail=f"Structural Violation: supplier {payload.supplier_id} is not registered.")
        cursor.execute("UPDATE erp_skus SET supplier_id = ? WHERE sku_id = ?", (payload.supplier_id, sku_id))
        conn.commit()
        logger.info(f"[SUPPLIER] {jwt_payload.get('sub')} reassigned {sku_id} -> {payload.supplier_id or 'UNASSIGNED'}")
        if payload.supplier_id:
            background_tasks.add_task(worker_evaluate_sku_threshold, sku_id)
        return {"status": "success", "sku_id": sku_id, "supplier_id": payload.supplier_id}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"SKU supplier reassignment error: {e}")
        raise HTTPException(status_code=500, detail="SKU supplier reassignment failed.")
    finally:
        conn.close()

@api_router.put("/inventory/skus/{sku_id}")
def update_sku(sku_id: str, payload: SKUUpdate, background_tasks: BackgroundTasks, jwt_payload: dict = Depends(verify_jwt_token)):
    """General HOD update endpoint for existing SKUs."""
    if jwt_payload.get("role") not in HOD_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: HOD clearance required.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        cursor.execute("SELECT 1 FROM erp_skus WHERE sku_id = ?", (sku_id,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail=f"SKU {sku_id} not found.")
        
        if payload.supplier_id:
            cursor.execute("SELECT 1 FROM erp_suppliers WHERE supplier_id = ?", (payload.supplier_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=400, detail=f"Supplier {payload.supplier_id} not found.")
        
        updates = []
        params = []
        for field_name, value in payload.dict(exclude_unset=True).items():
            updates.append(f"{field_name} = ?")
            params.append(value)
        
        if updates:
            params.append(sku_id)
            cursor.execute(f"UPDATE erp_skus SET {', '.join(updates)} WHERE sku_id = ?", params)
            conn.commit()
            
        logger.info(f"[SKU] {jwt_payload.get('sub')} updated SKU {sku_id}: {payload.dict(exclude_unset=True)}")
        background_tasks.add_task(worker_evaluate_sku_threshold, sku_id)
        return {"status": "success", "sku_id": sku_id}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"SKU update failed: {e}")
        raise HTTPException(status_code=500, detail="SKU update failed.")
    finally:
        conn.close()

def _append_sku_to_supplier_draft(conn, sku_id: str, quantity: int, notes):
    """
    Core manual-procurement primitive: append (accumulating) a SKU onto its
    supplier's open DRAFT purchase order, synthesizing a fresh draft when the
    supplier has none open. The caller owns authorization. Commits the unit of
    work and returns (po_id, created, line_quantity).
    """
    cursor = conn.cursor()
    cursor.execute("""
        SELECT s.sku_id, s.unit_cost, s.supplier_id, sup.default_lead_time_days
        FROM erp_skus s
        LEFT JOIN erp_suppliers sup ON s.supplier_id = sup.supplier_id
        WHERE s.sku_id = ?
    """, (sku_id,))
    sku = cursor.fetchone()
    if not sku:
        raise HTTPException(status_code=404, detail=f"SKU {sku_id} not found in master catalog.")
    if not sku["supplier_id"]:
        raise HTTPException(status_code=400, detail=f"Structural Violation: {sku_id} has no designated supplier. Assign one before procuring.")

    cursor.execute("BEGIN IMMEDIATE TRANSACTION")
    cursor.execute("SELECT po_id, notes FROM erp_purchase_orders WHERE supplier_id = ? AND status = 'DRAFT'", (sku["supplier_id"],))
    draft = cursor.fetchone()

    if draft:
        po_id, created = draft["po_id"], False
        cursor.execute("""
            INSERT INTO erp_purchase_order_items (po_id, sku_id, quantity, unit_cost)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(po_id, sku_id) DO UPDATE SET quantity = quantity + excluded.quantity
        """, (po_id, sku_id, quantity, sku["unit_cost"]))
        if notes:
            merged = f"{draft['notes']} | {notes}" if draft["notes"] else notes
            cursor.execute("UPDATE erp_purchase_orders SET notes = ? WHERE po_id = ?", (merged, po_id))
    else:
        po_id, created = f"PO-{uuid.uuid4().hex[:8].upper()}", True
        lead_days = sku["default_lead_time_days"] if sku["default_lead_time_days"] is not None else 7
        eta = (datetime.now(timezone.utc) + timedelta(days=lead_days)).date().isoformat()
        cursor.execute("""
            INSERT INTO erp_purchase_orders (po_id, supplier_id, status, priority, eta_date, notes)
            VALUES (?, ?, 'DRAFT', 0, ?, ?)
        """, (po_id, sku["supplier_id"], eta, notes))
        cursor.execute("""
            INSERT INTO erp_purchase_order_items (po_id, sku_id, quantity, unit_cost)
            VALUES (?, ?, ?, ?)
        """, (po_id, sku_id, quantity, sku["unit_cost"]))

    cursor.execute("SELECT quantity FROM erp_purchase_order_items WHERE po_id = ? AND sku_id = ?", (po_id, sku_id))
    line_qty = cursor.fetchone()["quantity"]
    conn.commit()
    return po_id, created, line_qty


@api_router.post("/orders/drafts/add-item", status_code=201)
def add_draft_item(payload: AddDraftItemPayload, jwt_payload: dict = Depends(verify_jwt_token)):
    """
    Manual procurement: append a SKU to the supplier's open DRAFT purchase order
    (accumulating quantity) regardless of reorder-threshold state. Restricted
    DM/TECH users may only procure SKUs they hold explicit clearance for.
    """
    conn = get_db_connection()
    try:
        role, _cleared = resolve_procurement_scope(conn, jwt_payload)
        requester = jwt_payload.get("sub")
        if not verify_employee_sku_access(conn, requester, payload.sku_id, role):
            raise HTTPException(status_code=403, detail=f"RBAC Violation: No procurement clearance for SKU {payload.sku_id}.")

        po_id, created, line_qty = _append_sku_to_supplier_draft(conn, payload.sku_id, payload.quantity, payload.notes)

        emit_inventory_event("po_draft_created" if created else "po_draft_updated",
                             {"po_id": po_id, "sku_id": payload.sku_id, "quantity": line_qty,
                              "requested_by": requester, "manual": True})
        logger.info(f"[MANUAL PROCURE] {requester} added {payload.quantity}x {payload.sku_id} to {po_id} ({'new draft' if created else 'existing draft'}).")
        return {"status": "success", "po_id": po_id, "created": created,
                "sku_id": payload.sku_id, "line_quantity": line_qty}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Manual procurement error: {e}")
        raise HTTPException(status_code=500, detail="Manual procurement failed.")
    finally:
        conn.close()

@api_router.post("/admin/users/{employee_id}/orders/drafts/add-item", status_code=201)
def admin_add_draft_item_on_behalf(employee_id: str, payload: AddDraftItemPayload, jwt_payload: dict = Depends(verify_jwt_token)):
    """
    Admin-issued draft PO line on behalf of a specific employee. The target
    must be authorized for the SKU (explicit clearance, or a globally-cleared
    role). Attribution records both the beneficiary and the acting admin.
    """
    _require_sku_access_admin(jwt_payload)
    conn = get_db_connection()
    try:
        emp = conn.execute("SELECT role FROM erp_employees WHERE id = ?", (employee_id,)).fetchone()
        if not emp:
            raise HTTPException(status_code=404, detail=f"Employee {employee_id} not found.")
        if not verify_employee_sku_access(conn, employee_id, payload.sku_id, emp["role"]):
            raise HTTPException(status_code=403, detail=f"{employee_id} has no procurement clearance for SKU {payload.sku_id}.")

        po_id, created, line_qty = _append_sku_to_supplier_draft(conn, payload.sku_id, payload.quantity, payload.notes)

        emit_inventory_event("po_draft_created" if created else "po_draft_updated",
                             {"po_id": po_id, "sku_id": payload.sku_id, "quantity": line_qty,
                              "requested_by": employee_id, "issued_by_admin": jwt_payload.get("sub"), "manual": True})
        logger.info(f"[ADMIN PROCURE] {jwt_payload.get('sub')} added {payload.quantity}x {payload.sku_id} to {po_id} on behalf of {employee_id}.")
        return {"status": "success", "po_id": po_id, "created": created,
                "sku_id": payload.sku_id, "line_quantity": line_qty, "on_behalf_of": employee_id}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Admin on-behalf procurement error: {e}")
        raise HTTPException(status_code=500, detail="On-behalf procurement failed.")
    finally:
        conn.close()

@api_router.put("/orders/{po_id}/update")
def update_draft_order(po_id: str, payload: DraftPOUpdatePayload, jwt_payload: dict = Depends(verify_jwt_token)):
    conn = get_db_connection()
    try:
        role, cleared = resolve_draft_mutation_scope(conn, jwt_payload)
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        cursor.execute("SELECT status FROM erp_purchase_orders WHERE po_id = ?", (po_id,))
        po = cursor.fetchone()
        if not po:
            raise HTTPException(status_code=404, detail="Purchase order not found.")
        if po["status"] != "DRAFT":
            raise HTTPException(status_code=400, detail=f"State Violation: Only DRAFT orders are mutable (current: {po['status']}).")

        if cleared is not None:
            # Restricted user: require clearance on this PO and on any SKU touched.
            po_skus = {r["sku_id"] for r in cursor.execute(
                "SELECT sku_id FROM erp_purchase_order_items WHERE po_id = ?", (po_id,)).fetchall()}
            if not (po_skus & cleared):
                raise HTTPException(status_code=403, detail="RBAC Violation: No procurement clearance for this purchase order.")
            for item in (payload.items or []):
                if item.sku_id not in cleared:
                    raise HTTPException(status_code=403, detail=f"RBAC Violation: No procurement clearance for SKU {item.sku_id}.")

        if payload.notes is not None:
            cursor.execute("UPDATE erp_purchase_orders SET notes = ? WHERE po_id = ?", (payload.notes, po_id))
        if payload.priority is not None:
            cursor.execute("UPDATE erp_purchase_orders SET priority = ? WHERE po_id = ?", (payload.priority, po_id))
        if payload.eta_date is not None:
            cursor.execute("UPDATE erp_purchase_orders SET eta_date = ? WHERE po_id = ?", (payload.eta_date, po_id))
        if payload.items:
            for item in payload.items:
                cursor.execute(
                    "UPDATE erp_purchase_order_items SET quantity = ? WHERE po_id = ? AND sku_id = ?",
                    (item.quantity, po_id, item.sku_id)
                )
        conn.commit()
        return {"status": "success", "detail": f"Draft {po_id} updated."}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Draft PO update error: {e}")
        raise HTTPException(status_code=500, detail="Draft update failed.")
    finally:
        conn.close()

@api_router.delete("/orders/{po_id}/items/{sku_id}")
def delete_draft_line_item(po_id: str, sku_id: str, jwt_payload: dict = Depends(verify_jwt_token)):
    """HOD line-item exclusion. Removing the final item dissolves the draft entirely."""
    conn = get_db_connection()
    try:
        role, cleared = resolve_draft_mutation_scope(conn, jwt_payload)
        # Restricted users may only exclude SKUs they are cleared for.
        if cleared is not None and sku_id not in cleared:
            raise HTTPException(status_code=403, detail=f"RBAC Violation: No procurement clearance for SKU {sku_id}.")
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        cursor.execute("SELECT status FROM erp_purchase_orders WHERE po_id = ?", (po_id,))
        po = cursor.fetchone()
        if not po:
            raise HTTPException(status_code=404, detail="Purchase order not found.")
        if po["status"] != "DRAFT":
            raise HTTPException(status_code=400, detail="State Violation: Line items can only be excluded from DRAFT orders.")
        cursor.execute("DELETE FROM erp_purchase_order_items WHERE po_id = ? AND sku_id = ?", (po_id, sku_id))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail=f"Line item {sku_id} not present on {po_id}.")
        cursor.execute("SELECT COUNT(*) AS n FROM erp_purchase_order_items WHERE po_id = ?", (po_id,))
        dissolved = cursor.fetchone()["n"] == 0
        if dissolved:
            cursor.execute("DELETE FROM erp_purchase_orders WHERE po_id = ?", (po_id,))
        conn.commit()
        return {"status": "success", "detail": f"Item {sku_id} excluded.", "po_dissolved": dissolved}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Line item exclusion error: {e}")
        raise HTTPException(status_code=500, detail="Line item exclusion failed.")
    finally:
        conn.close()

@api_router.post("/orders/submit")
def submit_draft_order(payload: POSubmitPayload, jwt_payload: dict = Depends(verify_jwt_token)):
    if jwt_payload.get("role") not in HOD_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: HOD clearance required.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        cursor.execute("SELECT status FROM erp_purchase_orders WHERE po_id = ?", (payload.po_id,))
        po = cursor.fetchone()
        if not po:
            raise HTTPException(status_code=404, detail="Purchase order not found.")
        if po["status"] != "DRAFT":
            raise HTTPException(status_code=400, detail=f"State Violation: Only DRAFT orders can be submitted (current: {po['status']}).")
        cursor.execute("SELECT COUNT(*) AS n FROM erp_purchase_order_items WHERE po_id = ?", (payload.po_id,))
        if cursor.fetchone()["n"] == 0:
            raise HTTPException(status_code=400, detail="Structural Violation: Cannot submit an empty purchase order.")
        cursor.execute(
            "UPDATE erp_purchase_orders SET status = 'PENDING_CFO', submitted_at = ? WHERE po_id = ?",
            (datetime.now(timezone.utc).isoformat(), payload.po_id)
        )
        conn.commit()
        emit_inventory_event("po_submitted", {"po_id": payload.po_id, "submitted_by": jwt_payload.get("sub")})
        return {"status": "success", "detail": f"{payload.po_id} routed to CFO approval queue."}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"PO submission error: {e}")
        raise HTTPException(status_code=500, detail="PO submission failed.")
    finally:
        conn.close()

# --- Transactional PO Email Dispatch ---

PO_EMAIL_TEMPLATE_PATH = os.path.join(_here, "po_email_template.html")
PO_EMAIL_LOG_DIR = os.path.join(_here, "logs", "po_emails")

def dispatch_po_email(po_id: str):
    """
    Renders the branded PO email and dispatches it to the supplier.
    Falls back to a mock SMTP log (logs/po_emails/<po_id>.html) when no SMTP host is configured.
    """
    conn = get_db_connection()
    try:
        pos = [p for p in _hydrate_purchase_orders(conn, ["APPROVED"]) if p["po_id"] == po_id]
    finally:
        conn.close()
    if not pos:
        logger.error(f"[PO EMAIL] {po_id} not found in APPROVED state. Dispatch aborted.")
        return
    po = pos[0]

    try:
        from jinja2 import Template
        with open(PO_EMAIL_TEMPLATE_PATH, "r", encoding="utf-8") as f:
            template = Template(f.read())
        html_body = template.render(po=po, generated_at=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))
    except Exception as e:
        logger.error(f"[PO EMAIL] Template rendering failed for {po_id}: {e}")
        return

    # Mock SMTP log: always persist the rendered artifact for audit
    try:
        os.makedirs(PO_EMAIL_LOG_DIR, exist_ok=True)
        artifact_path = os.path.join(PO_EMAIL_LOG_DIR, f"{po_id}.html")
        with open(artifact_path, "w", encoding="utf-8") as f:
            f.write(html_body)
        logger.info(f"[PO EMAIL] Rendered artifact persisted: {artifact_path}")
    except Exception as e:
        logger.error(f"[PO EMAIL] Failed to persist artifact for {po_id}: {e}")

    smtp_host = os.environ.get("SMTP_HOST")
    if not smtp_host:
        logger.warning(f"[PO EMAIL] SMTP_HOST not configured. {po_id} dispatch to {po['supplier_email']} logged in mock mode only.")
        return

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Purchase Order {po['po_id']} - Meta App Factory ERP"
        msg["From"] = os.environ.get("SMTP_FROM", "procurement@meta-app-factory.example.com")
        msg["To"] = po["supplier_email"]
        msg.attach(MIMEText(html_body, "html"))
        port = int(os.environ.get("SMTP_PORT", "587"))
        with smtplib.SMTP(smtp_host, port, timeout=15) as server:
            if os.environ.get("SMTP_USER"):
                server.starttls()
                server.login(os.environ["SMTP_USER"], os.environ.get("SMTP_PASSWORD", ""))
            server.sendmail(msg["From"], [po["supplier_email"]], msg.as_string())
        logger.info(f"[PO EMAIL] {po_id} dispatched to {po['supplier_email']} via {smtp_host}.")
    except Exception as e:
        logger.error(f"[PO EMAIL] SMTP dispatch failed for {po_id}: {e}")

# --- CSV Report Downloads ---

def _csv_download_response(filename: str, header: list, rows: list) -> Response:
    """Build an Excel-friendly CSV attachment (UTF-8 BOM so Excel detects encoding)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    for r in rows:
        writer.writerow(r)
    content = ("﻿" + buf.getvalue()).encode("utf-8")
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/reports/inventory/download")
def report_inventory_csv(jwt_payload: dict = Depends(verify_jwt_token)):
    """Inventory status CSV. Restricted DM/TECH see only their scoped SKUs;
    HM/CFO/ADMIN see all (gating via resolve_procurement_scope)."""
    conn = get_db_connection()
    try:
        role, cleared = resolve_procurement_scope(conn, jwt_payload)
        scope_clause = ""
        params = []
        if cleared is not None:
            cl = list(cleared)
            scope_clause = f" WHERE s.sku_id IN ({','.join('?' * len(cl))})"
            params = cl
        rows = conn.execute(f"""
            SELECT s.sku_id, s.nomenclature, c.name AS category, s.quantity_on_hand,
                   s.reorder_threshold, s.min_order_qty, sup.name AS supplier
            FROM erp_skus s
            LEFT JOIN erp_categories c ON s.category_id = c.id
            LEFT JOIN erp_suppliers sup ON s.supplier_id = sup.supplier_id
            {scope_clause}
            ORDER BY s.sku_id
        """, params).fetchall()
        data = [[r["sku_id"], r["nomenclature"], r["category"], r["quantity_on_hand"],
                 r["reorder_threshold"], r["min_order_qty"], r["supplier"]] for r in rows]
        return _csv_download_response(
            "inventory_report.csv",
            ["SKU", "Nomenclature", "Category", "On Hand", "Reorder Threshold", "Min Order Qty", "Supplier"],
            data,
        )
    finally:
        conn.close()


_WO_REPORT_HEADER = ["MWO ID", "Status", "DM Urgency", "HM Priority", "Equipment",
                     "Department", "Created By", "Assigned HM", "Assigned Tech",
                     "Created At", "Completed At"]


def _wo_report_rows(conn, where: str, params: list) -> list:
    rows = conn.execute(f"""
        SELECT w.mwo_id, w.status, w.dm_urgency, w.hm_priority, w.equipment_id,
               e.department_id, w.created_by, w.assigned_hm_id, w.assigned_tech,
               w.created_at, w.completed_at
        FROM work_orders w
        LEFT JOIN erp_equipment e ON w.equipment_id = e.equipment_id
        {where}
        ORDER BY w.created_at DESC, w.mwo_id DESC
    """, params).fetchall()
    return [[r["mwo_id"], r["status"], r["dm_urgency"], r["hm_priority"], r["equipment_id"],
             r["department_id"], r["created_by"], r["assigned_hm_id"], r["assigned_tech"],
             r["created_at"], r["completed_at"]] for r in rows]


@api_router.get("/reports/dm-work-orders/download")
def report_dm_work_orders_csv(jwt_payload: dict = Depends(verify_jwt_token)):
    """Work orders for the DM's department. DM scoped to their department; admins see all."""
    role = jwt_payload.get("role")
    if role not in ["DM"] + SKU_ACCESS_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: DM clearance required.")
    conn = get_db_connection()
    try:
        where, params = "", []
        if role == "DM":
            dept_row = conn.execute(
                "SELECT department_id FROM erp_employees WHERE id = ?", (jwt_payload.get("sub"),)
            ).fetchone()
            where = " WHERE e.department_id = ?"
            params = [dept_row["department_id"] if dept_row else None]
        data = _wo_report_rows(conn, where, params)
        return _csv_download_response("dm_work_orders.csv", _WO_REPORT_HEADER, data)
    finally:
        conn.close()


@api_router.get("/reports/hm-work-orders/download")
def report_hm_work_orders_csv(jwt_payload: dict = Depends(verify_jwt_token)):
    """Work orders received by the HM. HM scoped to their assigned orders; admins see all."""
    role = jwt_payload.get("role")
    if role not in ["HM"] + SKU_ACCESS_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: HM clearance required.")
    conn = get_db_connection()
    try:
        where, params = "", []
        if role == "HM":
            where = " WHERE w.assigned_hm_id = ?"
            params = [jwt_payload.get("sub")]
        data = _wo_report_rows(conn, where, params)
        return _csv_download_response("hm_work_orders.csv", _WO_REPORT_HEADER, data)
    finally:
        conn.close()


@api_router.get("/reports/cfo-procurement/download")
def report_cfo_procurement_csv(jwt_payload: dict = Depends(verify_jwt_token)):
    """Purchase-order list for the CFO. CFO/Admin only."""
    if jwt_payload.get("role") not in CFO_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: CFO clearance required.")
    conn = get_db_connection()
    try:
        rows = conn.execute("""
            SELECT po.po_id, po.status, po.priority, sup.name AS supplier, po.eta_date,
                   po.created_at, po.submitted_at, po.decided_at,
                   (SELECT ROUND(SUM(i.quantity * i.unit_cost), 2)
                    FROM erp_purchase_order_items i WHERE i.po_id = po.po_id) AS total_cost
            FROM erp_purchase_orders po
            LEFT JOIN erp_suppliers sup ON po.supplier_id = sup.supplier_id
            ORDER BY po.created_at DESC
        """).fetchall()
        data = [[r["po_id"], r["status"], r["priority"], r["supplier"], r["eta_date"],
                 r["created_at"], r["submitted_at"], r["decided_at"], r["total_cost"]] for r in rows]
        return _csv_download_response(
            "cfo_procurement.csv",
            ["PO ID", "Status", "Priority", "Supplier", "ETA", "Created At", "Submitted At", "Decided At", "Total Cost"],
            data,
        )
    finally:
        conn.close()


# --- CFO Actuation Endpoints ---

@api_router.get("/orders/approvals")
def get_cfo_approval_queue(jwt_payload: dict = Depends(verify_jwt_token)):
    if jwt_payload.get("role") not in CFO_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: CFO clearance required.")
    conn = get_db_connection()
    try:
        # priority DESC bubbles high-priority pulses to the top (enforced in _hydrate)
        return {"status": "success", "data": _hydrate_purchase_orders(conn, ["PENDING_CFO", "HOLD"])}
    finally:
        conn.close()

@api_router.post("/orders/actuate-bulk")
def actuate_orders_bulk(payload: BulkActuationPayload, background_tasks: BackgroundTasks, jwt_payload: dict = Depends(verify_jwt_token)):
    if jwt_payload.get("role") not in CFO_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: CFO clearance required to actuate purchase orders.")

    target_status = {"APPROVE": "APPROVED", "HOLD": "HOLD", "REJECT": "REJECTED"}[payload.action]
    decided_at = datetime.now(timezone.utc).isoformat()
    results = []
    approved_ids = []

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        for po_id in payload.po_ids:
            cursor.execute("SELECT status FROM erp_purchase_orders WHERE po_id = ?", (po_id,))
            po = cursor.fetchone()
            if not po:
                results.append({"po_id": po_id, "result": "NOT_FOUND"})
                continue
            if po["status"] not in ("PENDING_CFO", "HOLD"):
                results.append({"po_id": po_id, "result": f"STATE_VIOLATION ({po['status']})"})
                continue
            if po["status"] == "HOLD" and target_status == "HOLD":
                results.append({"po_id": po_id, "result": "ALREADY_HOLD"})
                continue
            cursor.execute(
                "UPDATE erp_purchase_orders SET status = ?, decided_at = ?, cfo_notes = COALESCE(?, cfo_notes) WHERE po_id = ?",
                (target_status, decided_at, payload.cfo_notes, po_id)
            )
            results.append({"po_id": po_id, "result": target_status})
            if target_status == "APPROVED":
                approved_ids.append(po_id)
        conn.commit()
    except Exception as e:
        conn.rollback()
        logger.error(f"Bulk actuation error: {e}")
        raise HTTPException(status_code=500, detail="Bulk actuation failed.")
    finally:
        conn.close()

    # Supplier email dispatch strictly decoupled from the response cycle
    for po_id in approved_ids:
        background_tasks.add_task(dispatch_po_email, po_id)

    emit_inventory_event("po_decided", {"action": payload.action, "results": results, "decided_by": jwt_payload.get("sub")})
    return {"status": "success", "action": payload.action, "results": results}

@api_router.post("/orders/{po_id}/receive")
def receive_purchase_order(po_id: str, jwt_payload: dict = Depends(verify_jwt_token)):
    """HOD shipment receipt: flips APPROVED -> FULFILLED and increments physical inventory."""
    if jwt_payload.get("role") not in HOD_ROLES:
        raise HTTPException(status_code=403, detail="RBAC Violation: HOD clearance required to receive shipments.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE TRANSACTION")
        cursor.execute("SELECT status FROM erp_purchase_orders WHERE po_id = ?", (po_id,))
        po = cursor.fetchone()
        if not po:
            raise HTTPException(status_code=404, detail="Purchase order not found.")
        if po["status"] != "APPROVED":
            raise HTTPException(status_code=400, detail=f"State Violation: Only APPROVED orders can be received (current: {po['status']}).")
        items = cursor.execute("SELECT sku_id, quantity FROM erp_purchase_order_items WHERE po_id = ?", (po_id,)).fetchall()
        for item in items:
            cursor.execute("UPDATE erp_skus SET quantity_on_hand = quantity_on_hand + ? WHERE sku_id = ?", (item["quantity"], item["sku_id"]))
        cursor.execute("UPDATE erp_purchase_orders SET status = 'FULFILLED' WHERE po_id = ?", (po_id,))
        conn.commit()
        emit_inventory_event("po_fulfilled", {"po_id": po_id, "received_by": jwt_payload.get("sub")})
        return {"status": "success", "detail": f"{po_id} fulfilled. Physical inventory synchronized.", "items_received": len(items)}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"PO receipt error: {e}")
        raise HTTPException(status_code=500, detail="PO receipt failed.")
    finally:
        conn.close()

app.include_router(api_router)

# --- FRONTEND DEPLOYMENT ---
frontend_dist_path = os.path.join(_erp, 'maintenance_frontend', 'dist')

if os.path.exists(frontend_dist_path):
    app.mount("/assets", StaticFiles(directory=os.path.join(frontend_dist_path, "assets")), name="assets")
    
