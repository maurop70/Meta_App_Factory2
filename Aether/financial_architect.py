"""
financial_architect.py -- Aether Financial Controller V2
=========================================================
Meta App Factory | Aether | Antigravity-AI

V2 Executive Polish:
  - Assumptions tab with editable parameters
  - All P&L cells use dynamic cell-reference formulas
  - Embedded Line Chart (12-month revenue growth)
  - Embedded Pie Chart (cost distribution)
  - Inter/Roboto fonts, professional borders, Google Sheets compatible
"""

# ── V3.0 Resilience Integration ──────────────────────────
import os as _os, sys as _sys
_FACTORY_DIR = _os.path.normpath(_os.path.join(_os.path.dirname(_os.path.abspath(__file__)), ".."))
_sys.path.insert(0, _FACTORY_DIR)
try:
    from factory import safe_post
    from local_state_manager import StateManager as _StateManager
    _v3_sm = _StateManager()
    _V3_AVAILABLE = True
except ImportError:
    _V3_AVAILABLE = False
# ── End V3 Integration ──────────────────────────────────

from auto_heal import healed_post, auto_heal, diagnose

def _v3_preflight():
    """V3: Ping Resonance_Watchdog_V3 before execution."""
    if not _V3_AVAILABLE:
        return True
    try:
        import json as _j
        _cfg_path = _os.path.join(_FACTORY_DIR, "resilience_config.json")
        if not _os.path.exists(_cfg_path):
            return True
        with open(_cfg_path) as _f:
            _cfg = _j.load(_f)
        _url = _cfg.get("cloud_health", {}).get("watchdog_url", "")
        if not _url:
            return True
        import requests as _rq
        _r = _rq.get(_url, timeout=5)
        return _r.status_code == 200
    except Exception:
        return False


import os
import sys
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

logger = logging.getLogger("aether.financial")

SCRIPT_DIR = Path(__file__).resolve().parent
FACTORY_DIR = SCRIPT_DIR.parent
sys.path.insert(0, str(FACTORY_DIR))
sys.path.insert(0, str(FACTORY_DIR / "Sentinel_Bridge"))

OUTPUT_DIR = FACTORY_DIR / "data" / "V2_Executive_Reports"
CREDS_PATH = FACTORY_DIR / "utils" / "auth" / "google_creds.json"

try:
    from dotenv import load_dotenv
    load_dotenv(FACTORY_DIR.parent / ".env")
    load_dotenv(FACTORY_DIR / ".env")
except ImportError:
    pass

_pii = None


def _get_pii():
    global _pii
    if _pii is None:
        try:
            from pii_masker import PIIMasker
            _pii = PIIMasker()
        except ImportError:
            pass
    return _pii


def get_google_credentials() -> dict:
    return {
        "client_id": os.getenv("GOOGLE_CLIENT_ID", ""),
        "client_secret": os.getenv("GOOGLE_CLIENT_SECRET", ""),
        "project_id": os.getenv("GOOGLE_PROJECT_ID", ""),
        "redirect_uri": os.getenv("GOOGLE_REDIRECT_URI", ""),
        "creds_file": str(CREDS_PATH) if CREDS_PATH.exists() else None,
        "has_credentials": bool(os.getenv("GOOGLE_CLIENT_ID")),
    }


# ── Styling Constants ────────────────────────────────────

HEADER_FONT = "Inter"
BODY_FONT = "Roboto"
PRIMARY_HEX = "667EEA"
ACCENT_HEX = "10B981"
DARK_HEX = "1E293B"
HEADER_BG_HEX = "EEF2FF"
ACCENT_BG_HEX = "ECFDF5"
WARN_HEX = "EF4444"


class FinancialArchitect:
    """
    V2 Financial Architect — Formula-driven spreadsheets with embedded charts.
    """

    def __init__(self):
        self._pii = _get_pii()
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    def generate_projections(self, config: dict = None,
                             output_name: str = None) -> dict:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, PatternFill, Border, Side, Alignment
            )
            from openpyxl.chart import LineChart, PieChart, Reference
            from openpyxl.chart.label import DataLabelList
            from openpyxl.chart.series import DataPoint
            from openpyxl.utils import get_column_letter
        except ImportError:
            return {"error": "openpyxl not installed", "generated": False}

        config = config or {}
        company = config.get("company_name", "Delegate AI — Antigravity-AI")
        agents = config.get("agents", 18)
        cost_per_agent = config.get("cost_per_agent", 50.0)
        monthly_rev = config.get("monthly_revenue", 100000.0)
        growth = config.get("growth_rate", 0.08)
        fixed_costs = config.get("fixed_costs", 25000.0)
        investment = config.get("investment", 150000.0)

        wb = Workbook()

        # ── Style shortcuts ──────────────────────────────
        hdr_font = Font(name=HEADER_FONT, size=11, bold=True, color=DARK_HEX)
        body_font = Font(name=BODY_FONT, size=10, color=DARK_HEX)
        title_font = Font(name=HEADER_FONT, size=14, bold=True,
                          color=PRIMARY_HEX)
        money_fmt = '$#,##0'
        pct_fmt = '0.0%'
        hdr_fill = PatternFill(start_color=HEADER_BG_HEX,
                               end_color=HEADER_BG_HEX, fill_type="solid")
        accent_fill = PatternFill(start_color=ACCENT_BG_HEX,
                                  end_color=ACCENT_BG_HEX, fill_type="solid")
        thin = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        center = Alignment(horizontal="center", vertical="center")

        def style_hdr(ws, row, cols):
            for c in range(1, cols + 1):
                cl = ws.cell(row=row, column=c)
                cl.font = hdr_font
                cl.fill = hdr_fill
                cl.border = thin
                cl.alignment = center

        def style_cell(ws, r, c, is_money=False, is_pct=False):
            cl = ws.cell(row=r, column=c)
            cl.font = body_font
            cl.border = thin
            cl.alignment = center
            if is_money:
                cl.number_format = money_fmt
            elif is_pct:
                cl.number_format = pct_fmt

        # ═══════════════════════════════════════════════════
        # SHEET 1: ASSUMPTIONS (editable driver tab)
        # ═══════════════════════════════════════════════════
        ws_a = wb.active
        ws_a.title = "Assumptions"

        ws_a.merge_cells("A1:C1")
        ws_a["A1"] = f"{company} — Input Assumptions"
        ws_a["A1"].font = title_font

        ws_a["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws_a["A2"].font = Font(name=BODY_FONT, size=9, italic=True,
                               color="94A3B8")

        # Headers
        for i, h in enumerate(["Parameter", "Value", "Description"], 1):
            ws_a.cell(row=4, column=i, value=h)
        style_hdr(ws_a, 4, 3)

        # Assumption rows (these are the DRIVERS — everything else references these)
        assumptions = [
            ("Agents", agents, "Total specialist agents in V7 Router"),
            ("Cost per Agent ($/mo)", cost_per_agent, "Monthly cost per agent"),
            ("Monthly Revenue (Base)", monthly_rev, "Starting monthly revenue"),
            ("Growth Rate (MoM)", growth, "Month-over-month growth rate"),
            ("Fixed Costs ($/mo)", fixed_costs, "Rent, infra, overhead"),
            ("Startup Investment", investment, "Initial capital deployed"),
            ("Revenue per Agent", None, "=B7/B5"),  # Formula
            ("Total Annual Agent Cost", None, "=B5*B6*12"),  # Formula
        ]

        # Cell map: B5=Agents, B6=Cost, B7=Revenue, B8=Growth, B9=Fixed, B10=Investment
        for i, (param, val, desc) in enumerate(assumptions):
            r = 5 + i
            ws_a.cell(row=r, column=1, value=param).font = body_font
            ws_a.cell(row=r, column=1).border = thin
            if val is not None:
                ws_a.cell(row=r, column=2, value=val)
            else:
                ws_a.cell(row=r, column=2, value=desc)  # Formula
            ws_a.cell(row=r, column=2).font = Font(name=HEADER_FONT, size=12,
                                                    bold=True, color=PRIMARY_HEX)
            ws_a.cell(row=r, column=2).border = thin
            ws_a.cell(row=r, column=2).alignment = center
            ws_a.cell(row=r, column=3, value=desc).font = Font(
                name=BODY_FONT, size=9, color="94A3B8")
            ws_a.cell(row=r, column=3).border = thin

            # Format
            if param.startswith("Cost") or param.startswith("Monthly Rev") or \
               param.startswith("Fixed") or param.startswith("Startup") or \
               param.startswith("Revenue per"):
                ws_a.cell(row=r, column=2).number_format = money_fmt
            if param.startswith("Growth"):
                ws_a.cell(row=r, column=2).number_format = pct_fmt

        ws_a.column_dimensions["A"].width = 30
        ws_a.column_dimensions["B"].width = 20
        ws_a.column_dimensions["C"].width = 40

        # ═══════════════════════════════════════════════════
        # SHEET 2: MONTHLY P&L (100% formula-driven)
        # ═══════════════════════════════════════════════════
        ws_pnl = wb.create_sheet("Monthly P&L")

        headers = ["Month", "Volume Factor", "Revenue", "Agent Costs",
                    "Fixed Costs", "Gross Profit", "Net Profit", "Margin %",
                    "Cumulative Profit"]
        for i, h in enumerate(headers, 1):
            ws_pnl.cell(row=1, column=i, value=h)
        style_hdr(ws_pnl, 1, len(headers))

        for month in range(1, 13):
            r = month + 1

            # A: Month number
            ws_pnl.cell(row=r, column=1, value=month)
            style_cell(ws_pnl, r, 1)

            # B: Volume Factor = (1+Growth)^(Month-1) -> refs Assumptions!B8
            ws_pnl[f"B{r}"] = f"=(1+Assumptions!B8)^(A{r}-1)"
            style_cell(ws_pnl, r, 2)
            ws_pnl.cell(row=r, column=2).number_format = '0.000'

            # C: Revenue = Base Revenue * Volume Factor -> refs Assumptions!B7
            ws_pnl[f"C{r}"] = f"=Assumptions!B7*B{r}"
            style_cell(ws_pnl, r, 3, is_money=True)

            # D: Agent Costs = Agents * Cost/Agent -> refs Assumptions!B5, B6
            ws_pnl[f"D{r}"] = f"=Assumptions!B5*Assumptions!B6"
            style_cell(ws_pnl, r, 4, is_money=True)

            # E: Fixed Costs -> refs Assumptions!B9
            ws_pnl[f"E{r}"] = f"=Assumptions!B9"
            style_cell(ws_pnl, r, 5, is_money=True)

            # F: Gross Profit = Revenue - Agent Costs
            ws_pnl[f"F{r}"] = f"=C{r}-D{r}"
            style_cell(ws_pnl, r, 6, is_money=True)

            # G: Net Profit = Gross - Fixed
            ws_pnl[f"G{r}"] = f"=F{r}-E{r}"
            style_cell(ws_pnl, r, 7, is_money=True)

            # H: Margin % = Net / Revenue
            ws_pnl[f"H{r}"] = f"=IF(C{r}>0,G{r}/C{r},0)"
            style_cell(ws_pnl, r, 8, is_pct=True)

            # I: Cumulative Profit = previous + Net - (Investment in M1)
            if month == 1:
                ws_pnl[f"I{r}"] = f"=G{r}-Assumptions!B10"
            else:
                ws_pnl[f"I{r}"] = f"=I{r-1}+G{r}"
            style_cell(ws_pnl, r, 9, is_money=True)

        # Totals row
        tr = 14
        ws_pnl.cell(row=tr, column=1, value="TOTAL").font = hdr_font
        ws_pnl.cell(row=tr, column=1).fill = accent_fill
        ws_pnl.cell(row=tr, column=1).border = thin
        for col_idx in range(3, 8):  # C through G
            cl = get_column_letter(col_idx)
            ws_pnl[f"{cl}{tr}"] = f"=SUM({cl}2:{cl}13)"
            ws_pnl.cell(row=tr, column=col_idx).font = hdr_font
            ws_pnl.cell(row=tr, column=col_idx).fill = accent_fill
            ws_pnl.cell(row=tr, column=col_idx).border = thin
            ws_pnl.cell(row=tr, column=col_idx).number_format = money_fmt

        # Annual margin
        ws_pnl[f"H{tr}"] = f"=IF(C{tr}>0,G{tr}/C{tr},0)"
        ws_pnl.cell(row=tr, column=8).font = hdr_font
        ws_pnl.cell(row=tr, column=8).fill = accent_fill
        ws_pnl.cell(row=tr, column=8).border = thin
        ws_pnl.cell(row=tr, column=8).number_format = pct_fmt

        for c in range(1, 10):
            ws_pnl.column_dimensions[get_column_letter(c)].width = 16

        # ── LINE CHART: 12-Month Revenue Growth ──────────
        line_chart = LineChart()
        line_chart.title = "12-Month Revenue Growth"
        line_chart.style = 10
        line_chart.y_axis.title = "Revenue ($)"
        line_chart.x_axis.title = "Month"
        line_chart.width = 22
        line_chart.height = 14

        # Revenue data (Column C, rows 2-13)
        rev_data = Reference(ws_pnl, min_col=3, min_row=1, max_row=13)
        rev_cats = Reference(ws_pnl, min_col=1, min_row=2, max_row=13)
        line_chart.add_data(rev_data, titles_from_data=True)
        line_chart.set_categories(rev_cats)

        # Net Profit data (Column G)
        net_data = Reference(ws_pnl, min_col=7, min_row=1, max_row=13)
        line_chart.add_data(net_data, titles_from_data=True)

        # Cumulative Profit (Column I)
        cum_data = Reference(ws_pnl, min_col=9, min_row=1, max_row=13)
        line_chart.add_data(cum_data, titles_from_data=True)

        # Style the series
        s0 = line_chart.series[0]  # Revenue
        s0.graphicalProperties.line.width = 28000
        s1 = line_chart.series[1]  # Net Profit
        s1.graphicalProperties.line.width = 22000
        s2 = line_chart.series[2]  # Cumulative
        s2.graphicalProperties.line.dashStyle = "dash"

        ws_pnl.add_chart(line_chart, "A17")

        # ═══════════════════════════════════════════════════
        # SHEET 3: COST DISTRIBUTION + PIE CHART
        # ═══════════════════════════════════════════════════
        ws_cost = wb.create_sheet("Cost Distribution")

        cost_headers = ["Category", "Monthly Cost", "Annual Cost", "% of Total"]
        for i, h in enumerate(cost_headers, 1):
            ws_cost.cell(row=1, column=i, value=h)
        style_hdr(ws_cost, 1, 4)

        cost_items = [
            ("Agent Costs", "=Assumptions!B5*Assumptions!B6"),
            ("Fixed Costs", "=Assumptions!B9"),
            ("Investment (Amortized 12mo)", "=Assumptions!B10/12"),
        ]

        for i, (cat, formula) in enumerate(cost_items):
            r = i + 2
            ws_cost.cell(row=r, column=1, value=cat).font = body_font
            ws_cost.cell(row=r, column=1).border = thin
            ws_cost[f"B{r}"] = formula
            style_cell(ws_cost, r, 2, is_money=True)
            ws_cost[f"C{r}"] = f"=B{r}*12"
            style_cell(ws_cost, r, 3, is_money=True)

        # Total row
        ws_cost.cell(row=5, column=1, value="TOTAL").font = hdr_font
        ws_cost.cell(row=5, column=1).fill = accent_fill
        ws_cost.cell(row=5, column=1).border = thin
        ws_cost[f"B5"] = "=SUM(B2:B4)"
        ws_cost.cell(row=5, column=2).font = hdr_font
        ws_cost.cell(row=5, column=2).fill = accent_fill
        ws_cost.cell(row=5, column=2).border = thin
        ws_cost.cell(row=5, column=2).number_format = money_fmt
        ws_cost[f"C5"] = "=SUM(C2:C4)"
        ws_cost.cell(row=5, column=3).font = hdr_font
        ws_cost.cell(row=5, column=3).fill = accent_fill
        ws_cost.cell(row=5, column=3).border = thin
        ws_cost.cell(row=5, column=3).number_format = money_fmt

        # Percentage column
        for r in range(2, 5):
            ws_cost[f"D{r}"] = f"=IF(B5>0,B{r}/B5,0)"
            style_cell(ws_cost, r, 4, is_pct=True)

        for c in range(1, 5):
            ws_cost.column_dimensions[get_column_letter(c)].width = 28

        # ── PIE CHART: Cost Distribution ─────────────────
        pie_chart = PieChart()
        pie_chart.title = "Cost Distribution"
        pie_chart.style = 10
        pie_chart.width = 18
        pie_chart.height = 14

        pie_data = Reference(ws_cost, min_col=2, min_row=1, max_row=4)
        pie_cats = Reference(ws_cost, min_col=1, min_row=2, max_row=4)
        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_cats)

        # Data labels
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.showCatName = True

        ws_cost.add_chart(pie_chart, "A8")

        # ═══════════════════════════════════════════════════
        # SHEET 4: SENSITIVITY ANALYSIS
        # ═══════════════════════════════════════════════════
        ws_sens = wb.create_sheet("Sensitivity Analysis")

        ws_sens.merge_cells("A1:E1")
        ws_sens["A1"] = "What-If: Cost Sensitivity Analysis"
        ws_sens["A1"].font = title_font

        sens_headers = ["Scenario", "Agent Cost", "Fixed Cost",
                        "Annual Net Profit", "Margin %"]
        for i, h in enumerate(sens_headers, 1):
            ws_sens.cell(row=3, column=i, value=h)
        style_hdr(ws_sens, 3, 5)

        scenarios = [
            ("Base Case", 1.0, 1.0),
            ("Costs +25%", 1.25, 1.25),
            ("Costs +50%", 1.50, 1.50),
            ("Costs Doubled", 2.0, 2.0),
            ("Costs Tripled", 3.0, 3.0),
            ("Revenue -20%", 1.0, 1.0),  # handled differently
        ]

        for i, (label, agent_mult, fixed_mult) in enumerate(scenarios):
            r = 4 + i
            ws_sens.cell(row=r, column=1, value=label).font = body_font
            ws_sens.cell(row=r, column=1).border = thin

            # Agent Cost = base * multiplier
            ws_sens[f"B{r}"] = f"=Assumptions!B6*{agent_mult}"
            style_cell(ws_sens, r, 2, is_money=True)

            # Fixed Cost
            ws_sens[f"C{r}"] = f"=Assumptions!B9*{fixed_mult}"
            style_cell(ws_sens, r, 3, is_money=True)

            if label == "Revenue -20%":
                # Revenue scenario: reduce revenue, keep costs
                ws_sens[f"D{r}"] = (
                    f"=(Assumptions!B7*0.8*12"
                    f"-Assumptions!B5*Assumptions!B6*12"
                    f"-Assumptions!B9*12)"
                )
            else:
                # Cost scenario: keep revenue, change costs
                ws_sens[f"D{r}"] = (
                    f"=(Assumptions!B7*12"
                    f"-Assumptions!B5*B{r}*12"
                    f"-C{r}*12)"
                )
            style_cell(ws_sens, r, 4, is_money=True)

            # Margin
            if label == "Revenue -20%":
                ws_sens[f"E{r}"] = f"=IF(Assumptions!B7*0.8*12>0,D{r}/(Assumptions!B7*0.8*12),0)"
            else:
                ws_sens[f"E{r}"] = f"=IF(Assumptions!B7*12>0,D{r}/(Assumptions!B7*12),0)"
            style_cell(ws_sens, r, 5, is_pct=True)

        for c in range(1, 6):
            ws_sens.column_dimensions[get_column_letter(c)].width = 22

        # ═══════════════════════════════════════════════════
        # SHEET 5: AGENT ECONOMICS
        # ═══════════════════════════════════════════════════
        ws_ag = wb.create_sheet("Agent Economics")

        agent_headers = ["Agent", "Monthly Cost", "Decisions/Mo",
                         "Cost/Decision", "ROI Contribution"]
        for i, h in enumerate(agent_headers, 1):
            ws_ag.cell(row=1, column=i, value=h)
        style_hdr(ws_ag, 1, len(agent_headers))

        agent_names = [
            "CEO", "CFO", "CTO", "CMO", "Deep Crawler", "The Critic",
            "The Librarian", "Compliance", "Data Architect", "Researcher",
            "Designer", "Presentations", "CX Strategist", "Aether",
            "Delegate AI", "EQ Engine", "GeoTalent Scout", "News Bureau",
        ]
        for idx, name in enumerate(agent_names):
            r = idx + 2
            decisions = 150 + (idx * 20)

            ws_ag.cell(row=r, column=1, value=name)
            # Cost = Assumptions!B6
            ws_ag[f"B{r}"] = "=Assumptions!B6"
            ws_ag.cell(row=r, column=3, value=decisions)
            # Cost/Decision = Cost / Decisions
            ws_ag[f"D{r}"] = f"=IF(C{r}>0,B{r}/C{r},0)"
            # ROI = (Revenue/Agents) / Cost
            ws_ag[f"E{r}"] = f"=IF(B{r}>0,(Assumptions!B7/Assumptions!B5)/B{r},0)"

            for c in range(1, 6):
                is_money = c in (2, 4)
                style_cell(ws_ag, r, c, is_money)

        for c in range(1, 6):
            ws_ag.column_dimensions[get_column_letter(c)].width = 18

        # ═══ Save ════════════════════════════════════════
        filename = output_name or f"Delegate_AI_2026_Projections_V2.xlsx"
        filepath = OUTPUT_DIR / filename
        wb.save(str(filepath))

        safe_path = str(filepath)
        if self._pii:
            safe_path = self._pii.mask(safe_path)
        logger.info("V2 Financial projections generated: %s", safe_path)

        # Compute breakeven for return
        cumulative = -investment
        breakeven_month = None
        for m in range(1, 13):
            rev_m = monthly_rev * ((1 + growth) ** (m - 1))
            net_m = rev_m - (agents * cost_per_agent) - fixed_costs
            cumulative += net_m
            if breakeven_month is None and cumulative >= 0:
                breakeven_month = m

        has_charts = True
        has_formulas = True

        return {
            "generated": True,
            "path": str(filepath),
            "sheets": ["Assumptions", "Monthly P&L", "Cost Distribution",
                        "Sensitivity Analysis", "Agent Economics"],
            "breakeven_month": breakeven_month,
            "agents": agents,
            "has_charts": has_charts,
            "has_formulas": has_formulas,
            "charts": ["12-Month Revenue Growth (Line)",
                       "Cost Distribution (Pie)"],
            "format": "Formula-driven, 16:9 compatible, Google Sheets importable",
        }


# ── Entry Point ──────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s [%(name)s] %(message)s")

    parser = argparse.ArgumentParser(
        description="Aether Financial Architect V2"
    )
    parser.add_argument("--test", action="store_true")
    args = parser.parse_args()

    architect = FinancialArchitect()

    if args.test:
        print("Financial Architect V2 -- Executive Polish")
        print("-" * 50)

        result = architect.generate_projections({
            "company_name": "Delegate AI -- Antigravity-AI",
            "agents": 18,
            "cost_per_agent": 50.0,
            "monthly_revenue": 100000.0,
            "growth_rate": 0.08,
            "fixed_costs": 25000.0,
            "investment": 150000.0,
        })

        print(f"Generated: {result['generated']}")
        print(f"File: {result['path']}")
        print(f"Sheets: {', '.join(result['sheets'])}")
        print(f"Charts: {', '.join(result['charts'])}")
        print(f"Has Formulas: {result['has_formulas']}")
        print(f"Has Charts: {result['has_charts']}")
        print(f"Break-even: Month {result['breakeven_month']}")
        print("\nDone!")
    else:
        print("Use --test to generate V2 projections.")
# V3 AUTO-HEAL ACTIVE
